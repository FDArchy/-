from django.http import HttpResponse, JsonResponse
#Auth: Permissions, groups, users, decorators
from django.contrib.auth.models import User, Permission, Group
from django.contrib.sessions.models import Session

from django.contrib import auth
from django.contrib.auth.decorators import user_passes_test
from django.views.decorators.csrf import csrf_protect
#Shortcuts
from django.shortcuts import render_to_response, redirect
#Request context
from django.template import RequestContext
#Models
from DB.models import Company, Manager, Task, Artist, Event, SiteUser, City, Call, CMSILink, Presentator, PCSLink, PresentatorEvent, TableName, FieldName, ChangeFieldLog, ChangeType, ChatMessage, SiteOptions, SiteUserActivity, SiteUserOptions
from django.db.models import Count, Sum
from django.db import connection
#Exceptions
from django.core.exceptions import ObjectDoesNotExist

#Django:
from django.db import connection
from django.utils import timezone
from django.core.servers.basehttp import FileWrapper
from django.db.models import Q, F, Sum

#Python Libs:
from datetime import datetime, timedelta
import time
from monthdelta import monthdelta

import json, pytz, django



#Дополнительные библиотеки
import functools
import sys, os
import operator
import random
#Сторонние библиотеки
import xlsxwriter
import openpyxl #На продакшн сервере не нужна
from openpyxl import load_workbook

#Свои библиотеки
from .python_funclib import connections as connections
#Свои декораторы:
def user_is_auth_aj_checker(_func):
    """Проверка авторизации через ajax с правильным response"""
    def wrapper(_request):
        if(not user_is_auth(_request)):
            return JsonResponse(create_response("error", "Необходима авторизация. Обновите страницу и заново войдите в свою учетную запись."))
        else:
            mark_activity(_request)
            return  _func(_request)
    return wrapper
def user_is_auth_checker(_func):
    """Проверка авторизации"""
    def wrapper(_request, *args, **kwargs):
        if(not user_is_auth(_request)):
            return redirect("/logon/")
        else:
            mark_activity(_request)
            return  _func(_request, *args, **kwargs)
    return wrapper
def user_is_manager_aj_checker(_func):
    """ajax проверка пользователя на принадлежность к менеджерам с правильным response"""
    def wrapper(_request):
        if(not user_is_manager(_request.user)):
            return JsonResponse(create_response("info", "Действие невозможно, пользователь не является менеджером"))
        else:
            return  _func(_request)
    return wrapper
def user_is_manager_or_admin_aj_checker(_func):
    """ajax проверка пользователя на принадлежность к менеджерам или администраторам с правильным response"""
    def wrapper(_request):
        if(user_is_manager(_request.user) or user_is_admin(_request.user)):
            return  _func(_request)
        else:
            return JsonResponse(create_response("info", "Действие невозможно, пользователь не является ни менеджером ни администратором"))
    return wrapper
def user_is_admin_aj_checker(_func):
    """ajax проверка пользователя на принадлежность к администраторам с правильным response"""
    def wrapper(_request):
        if(user_is_admin(_request.user)):
            return  _func(_request)
        else:
            return JsonResponse(create_response("info", "Действие невозможно, пользователь не является администратором"))
    return wrapper
def user_is_presentator_aj_checker(_func):
    """ajax проверка пользователя на принадлежность к артистам с правильным response"""
    def wrapper(_request):
        if (user_is_presentator(_request.user)):
            return _func(_request)
        else:
            return JsonResponse(
                create_response("info", "Действие невозможно, пользователь не является артистом"))
    return wrapper
#Чекеры
def user_is_presentator(_user):
    """Проверка пользователя на отношение к артистам"""
    siteuser = SiteUser.objects.get(user=_user)
    if (siteuser.type == "p"):
        return True
    else:
        return False
def user_is_manager(_user):
    """Проверка пользователя на отношение к менеджерам"""
    siteuser = SiteUser.objects.get(user=_user)
    if (siteuser.type == "m" or (siteuser.type == "a" and siteuser.options.admin_mode == False)):
        return True
    else:
        return False
def user_is_admin(_user):
    """Проверка пользователя на отношение к администраторам"""
    siteuser = SiteUser.objects.get(user = _user)
    if(siteuser.type == "a" and siteuser.options.admin_mode == True):
        return True
    else:
        return False
def user_is_admin_or_manager(_user):
    """Проверка пользователя на отношение к менеджерам или администраторам"""
    return user_is_admin(_user) or user_is_manager(_user)
def user_is_auth(_request):
    """Проверка пользователя на авторизованность"""
    if(_request.user.is_authenticated()):
        return True
    else:
        return False
def check_manager_allowed_company(_company, _siteuser):
    """Проверка доступности компании менеджеру"""
    try:
        company = Company.objects.get(id = _company)
    except:
        return False
    try:
        site_user = SiteUser.objects.get(id = _siteuser)
    except:
        return False

    user_type = get_current_user_type(_siteuser = site_user.id)
    if(user_type == "admin"):
        return True
    elif(user_type == "manager"):
        user_options = get_user_options_from_db(_user = site_user.user)
        if(user_options["can_show_all_companies"] and user_options["can_show_allien_companies"]):
            return True
        else:
            if(user_options["can_show_all_companies"]):
                if (company.id in get_manager_allowed_companies(get_manager_from_site_user(site_user.id).id)):
                    return True
                else:
                    if(not company.id in CMSILink.objects.exclude(manager__siteuser__id = site_user.id).values_list("company__id", flat=True)):
                        return True
            else:
                if(company.city.id in get_allowed_cities(_site_user = site_user.id, _full_list = True, _id_list=True)):
                    if(user_options["can_show_allien_companies"]):
                        return True
                    else:
                        if (not company.id in CMSILink.objects.exclude(manager__siteuser__id=site_user.id).values_list(
                                "company__id", flat=True)):
                            return True
                        else:
                            return  False
                else:
                    if(company.id in get_manager_allowed_companies(get_manager_from_site_user(site_user.id).id)):
                        return True
                    else:
                        return False
    else:
        return False
#Настройки
#загрузка опций из базы данных в объект
def get_site_options():
    """Получение настроек сайта в виде словаря"""
    options = SiteOptions.objects.all().values()[0]
    return options
def get_user_options_from_db(_request = False, _user = False, _individual = False):
    '''Получение опций пользователя в виде словаря'''
    if(_request):
        siteuser = SiteUser.objects.get(user=_request.user)
    elif(_user):
        siteuser = SiteUser.objects.get(user=_user)
    else:
        return []
    if(_individual):
        options_values = SiteUserOptions.objects.filter(id=siteuser.options.id).values("logout_request", "company_page_calendar", "larger_font", "scrolltop_show", "only_own_tasks_for_admin")[0]
    else:
        options_values = SiteUserOptions.objects.filter(id = siteuser.options.id).values()[0]
        options_values["full_access_cities_list"] =  list(siteuser.options.full_access_cities_list.all().values_list("id", flat=True))
    return options_values
def change_user_option(_request, _option, _value):
    '''Изменение выбранной опции пользователя'''
    siteuser = SiteUser.objects.get(user=_request.user)
    current_option = SiteUserOptions.objects.get(id=siteuser.options.id)
    changed_params = {
        _option: _value
    }
    SiteUserOptions.objects.filter(id = current_option.id).update(**changed_params)
    return True
def change_option_checker(_request, _option, _newvalue):
    '''Проверка опции пользователя и её сохранение при изменении'''
    current_options = get_user_options_from_db(_request)
    if(_newvalue == "" or _newvalue == None):
        return current_options[_option]
    if(current_options[_option] != _newvalue):
        change_user_option(_request, _option, _newvalue)
    return _newvalue
def check_user_permission(_user, _permission):
    """Проверка доступа пользователя - для Bool-полей"""
    option = SiteUserOptions.objects.filter(id=SiteUser.objects.get(user = _user).options.id).values(_permission)[0]
    return option[_permission]

#DESIGN==========================================================================================================================================================================================================================================================================================================================================================================================================================================================
#PAGES
@user_is_auth_checker
@user_passes_test(lambda x: False, login_url='/clients/')
def redirecttomain(request):
    """Редиректор с любого адреса"""
    return redirect('/clients')
@user_is_auth_checker
@user_passes_test(lambda x: False, login_url='/clients/')
def mainpage(request):
    """Редиректор с /main/"""
    return redirect('/clients/')
def logon(request):
    """Страница авторизации"""
    return render_to_response('auth/login.html', {}, context_instance=RequestContext(request))
@user_is_auth_checker
def help(request):
    """Страница помощи"""
    user_type = get_current_user_type(request)
    if(user_type == "admin" or user_type == "manager"):
        options = SiteOptions.objects.all().values("sippoint_login", "sippoint_password")[0]
    else:
        options = []
    edu_link = False
    edu_links = False
    if(user_type == "manager"):
        edu_links = {}
        edu_links["main_tutorial"] = "https://www.youtube.com/embed/o8dl23ir_ps"
        edu_links["tasks_work"] = "https://www.youtube.com/embed/3u3CAVKDMF8"
    elif(user_type == "admin"):
        edu_links = {}
        edu_links["admin"] = "https://www.youtube.com/embed/AAeU7pw0fsk"
        edu_links["manager"] = "https://www.youtube.com/embed/o8dl23ir_ps"
        edu_links["presentator"] = "https://www.youtube.com/embed/0bWwxX4I14Q"
    elif(user_type == "presentator"):
        edu_links = {}
        edu_links["main_tutorial"]= "https://www.youtube.com/embed/0bWwxX4I14Q"
    return  render_to_response('help/help.html', {"links": edu_links, "options": options}, context_instance=RequestContext(request))
#Страницы компаний
@user_is_auth_checker
@user_passes_test(user_is_admin_or_manager, login_url='/events/')
def company_show_page(request, company_number):
    """Вывод страницы компании"""
    try:
        company = Company.objects.get(id = company_number)#Поиск компании в базе
    except:
        return redirect("/")
    user = request.user
    manager = Manager.objects.get(siteuser__user = request.user)
    site_user = Manager.siteuser
    options = get_user_options_from_db(_request = request)
    #Если пользователь не имеет прав для доступа ко всем компаниям и пользователь не администратор, то проверка на принадлежность ему компании
    if(not check_manager_allowed_company(company.id, manager.siteuser.id)):
        return redirect("/")
    if(company.id in get_manager_allowed_companies(manager.id) and not options["admin_mode"]):
        allowed_shows = Artist.objects.filter(id__in=get_allowed_manager_shows_for_company(manager, company.id))
    else:
        allowed_shows = get_allowed_shows_for_city(_company=company.id)
    choosen_show = request.GET.get('artist')
    if choosen_show:
        choosen_show = int(choosen_show)

    dinamic_parameters = {}#Словарь со значениями, указывающими наличие изменений для установки активности кнопок

    allowed_tables = ["Company", "Call", "Task", "Event"]
    allowed_linked_tables = ["Company"]
    if (options["admin_mode"]):
        allowed_actions = ["add", "edit", "remove", "view", "transfer", "transference", "permission"]
        dinamic_parameters["managers"] = CMSILink.objects.filter(company__id = company.id).order_by("manager").distinct().values("manager__siteuser__id", "manager__siteuser__alias")
    else:
        allowed_actions = ["add", "edit", "remove", "view", "transfer", "transference", "permission"]
    logs_count = ChangeFieldLog.objects.filter(
            Q(table_link_id=company.id) | Q(link_to_object_id=company.id),
            Q(link_to_object_table__name__in=allowed_linked_tables) | Q(table__name__in=allowed_tables),
            changeType__name__in=allowed_actions).exclude(whoViewChange__id = manager.siteuser.id).count()
    dinamic_parameters["logs_count"] = logs_count

    allowed_presentators_filter_params = {}
    if(get_current_user_type(request) == "manager"):
        allowed_presentators_filter_params["shows__id__in"] = get_manager_allowed_shows_for_city(request, company.city.id)
    else:
        allowed_presentators_filter_params["shows__id__in"] = get_allowed_shows_for_city(company.city.id)
    allowed_presentators = json.dumps(list(Presentator.objects.filter(id__in = PCSLink.objects.filter(city__id = company.city.id, **allowed_presentators_filter_params).values_list("presentator__id", flat=True)).values("id", "siteuser__alias")))
    return render_to_response('companies/company_page/page_main.html', {'company':company, 'user':site_user, 'shows':allowed_shows, 'choosen_show':choosen_show, 'start_update_not_need': True, 'dinamic_parameters':dinamic_parameters, 'allowed_presentators': allowed_presentators}, context_instance=RequestContext(request))
@user_is_auth_checker
@user_passes_test(user_is_admin_or_manager, login_url='/events/')
def company_show_manager_work_content(request):
    """Вывод работы менеджера - блока с задачами, звонками и шоу, используется костыль - возвращается страница с заполненным контентом"""
    company = Company.objects.get(id = request.POST.get('company'))
    manager = Manager.objects.get(siteuser__user = request.user)
    user_type = get_current_user_type(_request = request)
    all = False
    if(request.POST.get('artist') == '0'):
        artist = get_allowed_shows_for_city(_company=company.id)
        all = True
    else:
        artist = Artist.objects.filter(id = request.POST.get('artist'))

    last_call = Call.objects.filter(company = company, artist__in = artist).exclude(type = "event").order_by("-datetime")[:1]

    task_filter_values = {}
    task_filter_values["company"] = company.id
    task_filter_values["artist__in"] = artist
    if(not user_type == "admin"):
        task_filter_values["manager"] = manager



    last_task = Task.objects.filter(done = True, **task_filter_values).order_by("-doneDateTime")[:1]
    next_task = Task.objects.filter(done = False, **task_filter_values).order_by("datetime")[:1]

    last_event = Event.objects.filter(company = company, startTime__lte = datetime.now(), artist__in = artist).order_by("-startTime")[:1]
    next_event = Event.objects.filter(company = company, startTime__gte = datetime.now(), artist__in = artist).order_by("startTime")[:1]

    events_ids = []

    has_last_event = False
    has_next_event = False

    last_event_status = "error"
    next_event_status = "error"

    if(last_event.count() > 0):
        has_last_event = True
        events_ids.append(last_event[0].id)
    if(next_event.count() > 0):
        has_next_event = True
        events_ids.append(next_event[0].id)
    events_with_fields = Event.objects.filter(id__in = events_ids).values(*events_get_values_list_for_calc_statuses())
    events_statuses = events_fill_events_statuses(events_with_fields, request)
    for event in events_statuses:
        if(has_last_event):
            if(event["id"] == last_event[0].id):
                last_event_status = event["status"]
        if(has_next_event):
            if (event["id"] == next_event[0].id):
                next_event_status = event["status"]

    hasCheckers = {}#Словарь со значениями, указывающими наличие изменений для установки активности кнопок

    if(Call.objects.filter(company = company, artist__in = artist).count() > 0):
        hasCheckers["calls"] = 1
    if(Task.objects.filter(**task_filter_values).count() > 0):
        hasCheckers["tasks"] = 1
    if(Event.objects.filter(company = company, artist__in = artist).count() > 0):
        hasCheckers["events"] = 1

    return render_to_response('companies/company_page/aj_manager_work.html', {"company": company, 'last_call':last_call, 'last_task':last_task, 'next_task':next_task, 'last_event':last_event, 'next_event':next_event, 'all_button':all, 'has_checkers':hasCheckers, 'current_manager':manager, 'last_event_status':last_event_status, 'next_event_status':next_event_status}, context_instance=RequestContext(request))
@user_is_auth_checker
@user_passes_test(user_is_admin_or_manager, login_url='/events/')
def companies_show_page(request):
    """Страница компаний"""
    options = get_user_options_from_db(request)
    return render_to_response('companies/companies_page/page_main.html', options, context_instance=RequestContext(request))
@user_is_auth_checker
@user_passes_test(user_is_admin_or_manager, login_url='/events/')
def companies_full_list_show_page(request):
    """Страница полного списка компаний"""
    options = get_user_options_from_db(request)
    response = {}
    response["options"] = options
    full_list_cities = []
    if(options["admin_mode"] or options["can_show_all_companies"]):
        full_list_cities = City.objects.filter(id__in = get_worked_cities()).values("id", "name")
    else:
        full_list_cities = City.objects.filter(id__in = options["full_access_cities_list"]).values("id", "name")
    response["full_list_cities"] = json.dumps(list(full_list_cities))
    response["full_list_cities_count"] = len(full_list_cities)
    response["full_list_shows"] = json.dumps(list(Artist.objects.filter(id__in=City.objects.filter(id__in = get_allowed_cities(request, _clients_list=True, _full_list=True, _id_list=True)).distinct().values_list("worked_shows__id", flat=True)).values()))
    response["unstandart_load"] = True
    if(response["options"]["admin_mode"]):
        response["managers"] = json.dumps(list(SiteUser.objects.filter((Q(type = "m") | Q(type = "a")), deleted=False).values("id", "alias").order_by("alias")))
    response["choosen_city"] = {}
    response["choosen_city"]["id"] = 0
    response["choousen_show"] = 0
    return render_to_response('companies/full_companies_list_page/page_main.html', response, context_instance=RequestContext(request))
@user_is_auth_checker
def events_show_page(request):
    """Страница мероприятий"""
    try:
        presentator = get_current_presentator(request).id
    except:
        presentator = 0
    return render_to_response('events/page_main.html', {"presentator": presentator, "first_update_not_need": False}, context_instance=RequestContext(request))
@user_is_auth_checker
@user_passes_test(user_is_admin_or_manager, login_url='/events/')
def tasks_show_page(request):
    """Страница задач"""
    return render_to_response('tasks/page_main.html', {"first_update_not_need": False}, context_instance=RequestContext(request))
@user_is_auth_checker
@user_passes_test(user_is_admin, login_url='/clients/')
def control_show_page(request):
    return render_to_response('control/page_main.html', {'start_update_not_need':True}, context_instance=RequestContext(request))
@user_is_auth_checker
@user_passes_test(user_is_admin_or_manager, login_url='/events/')
def stats_show_page(request):
    return render_to_response('stats/page_main.html', {'start_update_not_need':True, 'user_type':get_current_user_type(request)}, context_instance=RequestContext(request))
#Auth:=====================================================================================
def login(request):
    """Авторизация"""
    if request.POST:
        username = request.POST.get('login', '')
        password = request.POST.get('password', '')
        user = auth.authenticate(username = username.strip(), password = password.strip())
        options = get_site_options()
        if user is not None:
            try:
                site_user = SiteUser.objects.get(user = user)
                if(site_user.deleted):
                    return JsonResponse(create_response("error", "Учетные данные неверны. Попробуйте еще раз"))
                if(not site_user.active):
                    return JsonResponse(create_response("error", "Ваша учетная запись заблокирована. Для разблокировки обратитесь к менеджеру сайта (" + options["main_manager_email"]  +")"))
            except:
                return JsonResponse(create_response("error", "Произошла неизвестная ошибка при авторизации. Перезагрузите страницу и попробуйте еще раз, при повторном возникновении обратитесь к администратору сайта"))
            try:
                auth.login(request, user)
                response = {}
                response["user_type"] = get_current_user_type(_user=user)
                return JsonResponse(create_response("data", "", response))
            except:
                return JsonResponse(create_response("error", "Произошла неизвестная ошибка при авторизации. Перезагрузите страницу и попробуйте еще раз, при повторном возникновении обратитесь к администратору сайта"))
        else:
            return JsonResponse(create_response("error", "Учетные данные неверны. Попробуйте еще раз"))
    else:
        return JsonResponse(create_response("error", "Неверный тип запроса авторизации."))
@user_is_auth_aj_checker
def logout(request):
    """Деавторизация"""
    auth.logout(request)
    return redirect('/logon/')


#Companies:=================================================================================
#Company:===================================================================================
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def company_company_data_content(request):
    try:
        company = Company.objects.filter(id=request.POST.get("id")).values()[0]
    except:
        return JsonResponse(create_response("error", "Компания не найдена в базе"))
    manager = get_current_manager(request)
    from_event = bool_convertor_error_json_response(request.POST.get("from_event"))

    allowed_company = check_manager_allowed_company(company["id"], manager.siteuser.id)
    allowed_show = False

    manager_options = get_user_options_from_db(request)
    if ((company["city_id"] in get_manager_allowed_cities(manager.id) and manager_options["can_show_allowed_cities_events"] and from_event)):
        allowed_show = True
    if(not(allowed_company or allowed_show)):
        return  JsonResponse(create_response("info", "Невозможно просмотреть данные чужой компании"))
    return JsonResponse(create_response("data", "", {"city_name": City.objects.get(id = company["city_id"]).name, "cities": list(City.objects.filter(id__in = get_worked_cities()).values("id", "name")), "company": company, "user_type": get_current_user_type(request)}), safe = False)
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def company_company_data_page(request):
    """Обновление страницы компании"""
    result = {}
    company = Company.objects.filter(id=request.POST.get("id"))
    user_type = get_current_user_type(request)
    site_user = get_current_site_user(request)
    if(company.count() == 0):
        return JsonResponse(create_response("error", "Компания не найдена"))
    company_data = list(company.values("id", "ctype", "name", "adress", "telephone", "contacts", "comment", "email", "site"))[0]
    if(user_type == "admin"):
        result['managers'] = list(Manager.objects.filter(id__in=CMSILink.objects.filter(company__id=company[0].id).values("manager__id")).values("siteuser__id", "siteuser__alias"))
    else:
        if(not check_manager_allowed_company(company[0].id, get_current_manager(request).siteuser.id)):
            return JsonResponse(create_response("info", "Невозможно просмотреть данные чужой компании"))
    allowed_tables = ["Company", "Call", "Task", "Event"]
    allowed_linked_tables = ["Company"]
    allowed_actions = []
    if (user_type == "admin"):
        allowed_actions = ["add", "edit", "remove", "view", "transfer", "transfer_back", "permission"]
    else:
        allowed_actions = ["add", "edit", "remove"]
    result['logs_count'] = ChangeFieldLog.objects.filter(
        Q(table_link_id=company[0].id) | Q(link_to_object_id=company[0].id),
        Q(link_to_object_table__name__in=allowed_linked_tables) | Q(table__name__in=allowed_tables),
        changeType__name__in=allowed_actions).exclude(whoViewChange__id = site_user).count()
    result['data'] = company_data
    return JsonResponse(create_response("data", "", result))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def company_manager_work_call_task(request):
    datetime = request.POST.get("current_datetime")
    if(request.POST.get("company")):
        try:
            company = Company.objects.get(id = request.POST.get("company"))
        except:
            return JsonResponse(create_response("error", "Не найдена компания с таким идентификатором"))
    if(request.POST.get("artist")):
        try:
            artist = Artist.objects.get(id = request.POST.get("artist"))
        except:
            return JsonResponse(create_response("error", "Не найдено шоу с таким идентификатором"))
    manager = get_current_manager(request)
    success_text = ""
    if(request.POST.get("task_exe")):
        try:
            task = Task.objects.get(id = request.POST.get("task_id"))
        except:
            return JsonResponse(create_response("error", "Задача по идентификатору не найдена для выполнения"))
        if(Task.objects.filter(id = request.POST.get("task_id"), manager = manager).count() == 0):
            return JsonResponse(create_response("error", "Невозможно выполнить чужую задачу"))
        else:
            try:
                task.done = True
                task.doneDateTime = datetime
                task.save()
            except:
                return JsonResponse(create_response("error", "Ошибка при отметке выполнения задачи"))
        success_text += "Успешная отметка выполнения задачи"
    if(request.POST.get("call")):#Если в запросе есть звонок то добавляем звонок
        if(request.POST.get("new_call")):#Если передан new_call значит это новый звонок
            try:
                call = Call(
                    datetime = datetime,
                    manager = manager,
                    company = company,
                    artist = artist,
                    comment = request.POST.get('call_comment')
                )
                call.save()
                #Если это отзвон по мероприятиям, не записывать его в логгер
                if(not request.POST.get("task_type")):
                    try:
                        write_to_change_log(call.id, "add", "Call", "comment", call.comment, get_current_site_user(request), company.id, "Company")
                    except:
                        return JsonResponse(create_response("error", "Ошибка при добавлении лог файла, обратитесь к администратору"))
            except:
                return JsonResponse(create_response("error", "Ошибка при добавлении звонка"))
            if(request.POST.get("task_type") == "event"):
                call.type = "event"
                call.save()

                event_with_choosen_task_query = Event.objects.filter(Q(taskDay__id = request.POST.get("task_id")) | Q(taskWeek__id = request.POST.get("task_id")) | Q(taskMonth__id = request.POST.get("task_id"))).values("id", "taskDay__id", "taskWeek__id", "taskMonth__id")
                if(event_with_choosen_task_query.count() != 0):
                    event_with_choosen_task = event_with_choosen_task_query[0]
                    event = Event.objects.get(id = event_with_choosen_task["id"])
                    if(event_with_choosen_task["taskDay__id"] == int(request.POST.get("task_id"))):
                        event.callDay = call
                    if (event_with_choosen_task["taskWeek__id"] == int(request.POST.get("task_id"))):
                        event.callWeek = call
                    if (event_with_choosen_task["taskMonth__id"] == int(request.POST.get("task_id"))):
                        event.callMonth = call
                    event.save()
                write_to_change_log(call.id, "add", "Call", "comment", call.comment, get_current_site_user(request),
                                    event.id, "Event")
        if(request.POST.get("task_exe")):
            success_text = ". Успешное добавление звонка"
        else:
            success_text += "Успешное добавление звонка"
    if(request.POST.get("task")):#Если в запросе есть задача то добавляем задачу
        try:
            assigned_by = get_current_site_user(request, _type="object")
            if(get_current_user_type(request) == "admin"):
                manager_id = int_convertor_error_false(request.POST.get("manager_id"))
                if(manager_id):
                    manager = Manager.objects.get(id = manager_id)
                    assigned_by = get_current_site_user(request, _type="object")
            task = Task(
                manager = manager,
                company = company,
                artist = artist,
                datetime = request.POST.get('task_datetime'),
                description = request.POST.get('task_description'),
                comment = request.POST.get('task_comment'),
                done = False,
                assigned_by = assigned_by
            )
            task.save()
        except:
            return JsonResponse(create_response("error", "Ошибка при добавлении задачи"))
        if(request.POST.get("call")):
            success_text += " и задачи"
        else:
            success_text = "Успешное добавление задачи"
        write_to_change_log(task.id, "add", "Task", "description", task.description, get_current_site_user(request), company.id, "Company")
    return JsonResponse(create_response("success", success_text))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def company_remove(request):
    """Удаление компании или компаний"""
    if(request.POST.get('companies')):
        if(not get_user_options_from_db(request)["admin_mode"]):
            return JsonResponse(create_response("error", "Действие запрещено для пользователей, не являющихся администраторами"))
        if(not request.POST.get('remove_confirm') == "true"):
            return JsonResponse(create_response("error", "Удаление не подтверждено. Для подтверждения используйте переключатель"))
        companies = json.loads(request.POST.get('companies'))
        Company.objects.filter(id__in = companies).delete()
        ChangeFieldLog.objects.filter(table__name="Company", table_link_id__in = companies).delete()
        ChangeFieldLog.objects.filter(link_to_object_table__name = "Company", link_to_object_id__in = companies).delete()
        return JsonResponse(create_response("success", "Успешное удаление компаний"))
    else:
        if (not request.POST.get('remove_confirm') == "true"):
            return JsonResponse(create_response("error", "Удаление не подтверждено. Для подтверждения используйте переключатель"))
        try:
            company = Company.objects.get(id = request.POST.get('id'))
        except:
            return JsonResponse(create_response("error", "Не найдена компания с таким идентификатором"))
        if(user_is_admin(request.user)) :#Если пользователь - администратор сайта, значит можно удалять
            text = "Учреждение \"" + company.name + "\" было успешно удалено из базы. Автоматическая переадресация через 5 секунд"
            try:
                ChangeFieldLog.objects.filter(table__name="Company", table_link_id=company.id).delete()
                ChangeFieldLog.objects.filter(link_to_object_table__name="Company",
                                              link_to_object_id=company.id).delete()
                company.delete()
                write_to_change_log(0, "remove", "Company", "name", company.name,
                                    get_current_site_user(request))  # Запись в лог
                if(not mark_users_cities_changed()):
                    JsonResponse(create_response("error", "Ошибка при пересчете списка городов"))
                return JsonResponse(create_response("success", text))
            except:
                return JsonResponse(create_response("error", "Ошибка при удалении учреждения из базы"))
        elif CMSILink.objects.filter(company__id = company.id, manager = get_current_manager(request)).count() > 0:#Если у пользователя доступно учреждение
            if ((datetime.now() - company.dateAdd).total_seconds()/60) < 10.0: #Если со времени добавления компании прошло менее 10 минут, то ее можно удалить
                text = "Учреждение \"" + company.name + "\" было успешно удалено из базы. Автоматическая переадресация через 5 секунд"
                try:
                    ChangeFieldLog.objects.filter(table__name="Company", table_link_id=company.id).delete()
                    ChangeFieldLog.objects.filter(link_to_object_table__name="Company",
                                                  link_to_object_id=company.id).delete()
                    company.delete()
                    write_to_change_log(0, "remove", "Company", "name", company.name,
                                        get_current_site_user(request))  # Запись в лог
                    if(not mark_users_cities_changed(get_current_site_user(request))):
                        JsonResponse(create_response("error", "Ошибка при пересчете списка городов"))
                    return JsonResponse(create_response("success", text))
                except:
                    return JsonResponse(create_response("error", "Ошибка при удалении учреждения из базы"))
            else:
                text = "С момента добавления учреждения прошло более 10 минут, для удаления учреждения свяжитесь с администратором."
                return JsonResponse(create_response("error", text))
        else:
            text = "У вас нет разрешения на удаление данного учреждения, для удаления учреждения свяжитесь с администратором."
            return JsonResponse(create_response("error", text))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def company_edit_or_add(request):
    if(request.POST.get("id") == "False"):
        request_company_id = 0
    else:
        request_company_id = request.POST.get("id")

    company = Company.objects.filter(id = request_company_id)
    company_data = json.loads(request.POST.get("changed"))
    if company.count() == 0:
        manager = get_current_manager(request)
        shows = company_data["shows"]
        company_data.pop('shows', None)
        company_data = convert_ids_to_objects("Company", company_data)
        new = Company(**company_data)
        new.whoAdd = get_current_site_user(_request=request, _type="object")
        new.save()
        for show in shows:
            try:
                cmsilink = CMSILink.objects.get(manager = manager, show = Artist.objects.get(id = show))
                cmsilink.company.add(new)
            except:
                try:
                    newCmsilink = CMSILink(manager = manager, show = Artist.objects.get(id = show))
                    newCmsilink.save()
                    newCmsilink.company.add(new)
                except:
                    JsonResponse(create_response("error", "Компания добавлена, но не удалось осуществить привязку, обратитесь к администратору"))
        try:
            mark_users_cities_changed(get_site_user_from_manager(manager.id))
        except:
            JsonResponse(create_response("error",
                                         "Компания добавлена и привязана, но не удалось осуществить обновление текущих городов менеджера, обратитесь к администратору"))
        return JsonResponse(create_response("success", "Компания успешно добавлена и привязана"))
    else:
        old_company_data = company.values()[0]
        new_city = 0
        if("city" in company_data):
            new_city = company_data.pop("city")

        for key in company_data:
            if(company_data[key] != old_company_data[key]):
                write_to_change_log(old_company_data["id"], "edit", "Company", key, old_company_data[key], get_current_site_user(request))
        if(new_city != 0):
            company_data["city"] = new_city
        company_data = convert_ids_to_objects("Company", company_data)
        company.update(**company_data)
        return JsonResponse(create_response("success", "Компания успешно отредактирована"))


@user_is_auth_aj_checker
@user_is_admin_aj_checker
def company_managers(request):
    """Получение списка менеджеров компании"""
    try:
        company = Company.objects.get(id = request.POST.get("company_id"))
    except:
        return JsonResponse(create_response("error", "Не найдена компания с таким идентификатором"))
    result = {}
    dict_list = []
    managers = Manager.objects.filter(id__in = CMSILink.objects.filter(company__id = company.id).values("manager__id")).values("siteuser__id")

    site_users = SiteUser.objects.filter(id__in = managers).values("id", "alias")
    shows = CMSILink.objects.filter(company__id = company.id).values("manager__siteuser__id", "show__id", "show__color", "show__name")
    for site_user in site_users:
        dict_item = {}
        dict_item["site_user"] = site_user
        dict_item["shows"] = []
        for show in shows:
            if(show["manager__siteuser__id"] == site_user["id"]):
                dict_item["shows"].append(show)
        last_log_event = list(ChangeFieldLog.objects.filter(Q(table_link_id=company.id) | Q(link_to_object_id=company.id), Q(
        link_to_object_table__name__in=["Company"]) | Q(table__name__in=["Company", "Task", "Call", "Event"]), changeType__name__in = ["add", "edit", "remove"], whoChange__id = site_user["id"]).order_by(
        "-datetime").values("id", "datetime", "whoChange__id", "whoChange__alias", "field__verbose", "value", "table__name", "changeType__name", "changeType__verbose",
        "link_to_object_table__name", "table_link_id", "link_to_object_id"))
        if(last_log_event.__len__() > 0):
            dict_item["last_log_event"]  = last_log_event[0]
        else:
            dict_item["last_log_event"] = ""
        dict_list.append(dict_item)
    result["managers"] = dict_list
    return JsonResponse(create_response("data", "", result))
#############################################################################################
#CompaniesList:==============================================================================
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def companies_list(request):
    #"""Получение списка учреждений (при переданном в request-post: paginator, возвращает число страниц)"""
    elements_on_page = int(change_option_checker(request, "clients_page_companies_count", request.POST.get("companies_count")))
    sort_by = change_option_checker(request, "clients_page_order_by", request.POST.get("sort_by"))
    admin_mode = get_user_options_from_db(request)["admin_mode"]
    user_type = get_current_user_type(_request=request)
    filter_type = request.POST.get("filter_type")

    paginator = request.POST.get("paginator")
    current_page = request.POST.get("page")
    if(current_page):
        current_page = int(current_page)
    city_id = int_convertor_error_0(request.POST.get("city_id"))
    show_id = int_convertor_error_0(request.POST.get("show_id"))
    search_string = request.POST.get("search")
    manager = Manager.objects.get(siteuser__user=request.user)

    companies_filter_params = {}  # Словарь kwargs с параметрами ддя фильтрации учреждений
    companies_filter_params["manager__id"] = manager.id  # Добавление параметра фильтрации по менеджеру
    if (city_id != 0):
        companies_filter_params["company__city__id"] = city_id  # Добавление параметра фильтрации по городу
    if (show_id != 0):
        companies_filter_params["show__id"] = show_id  # Добавление параметра фильтрации по шоу
    if(filter_type) == "any":
        filter_type = ""
    if(paginator):
        if(admin_mode):
            if(city_id != 0):
                companies_count = Company.objects.filter(Q(city_id=city_id), Q(name__icontains=search_string) | Q(
                    adress__icontains=search_string), Q(ctype__icontains=filter_type)).count()
            else:
                allowed_cities_list = get_worked_cities()
                if(show_id != 0):
                    allowed_cities_by_shows = set(
                        City.objects.filter(enabled=True,
                                            worked_shows__id=show_id).distinct().values_list(
                            "id", flat=True))
                    allowed_cities_list = list(set(allowed_cities_list) & allowed_cities_by_shows)
                companies_count = Company.objects.filter(Q(city__id__in = allowed_cities_list), Q(name__icontains=search_string) | Q(
                    adress__icontains=search_string), Q(ctype__icontains=filter_type)).count()

        else:
            companies_count = Company.objects.filter(Q(id__in=CMSILink.objects.filter(**companies_filter_params).values("company__id").distinct()), Q(name__icontains=search_string) | Q(adress__icontains=search_string), Q(ctype__icontains=filter_type)).count()
        pages_count = calc_pages_count(elements_on_page, companies_count)
        return JsonResponse(create_response("data", "", {"page_count": pages_count}))

    result = {}
    filtred_companies_list = []#Список компаний

    company_filter_values = {}#Параметры фильтрации для компаний

    #Список выводимых параметров
    params_list = {"id":"Код", "type":"Тип", "name":"Название", "city_name":"Город", "adress":"Адрес", "last_call":"Звонок", "last_task":"Задача", "last_event":"Мероприятие", "is_have_contacts":"Контакты указаны", "is_have_telephone":"Телефоны указаны", "is_have_comment":"Комментарии указаны", "is_have_site":"Сайт указан", "is_have_email":"Почта указана"}

    #Заполнение фидьтрующего компании списка:
    if city_id != 0:
        company_filter_values["company__city__id"] = city_id
    if show_id != 0:
        company_filter_values["show__id"] = show_id
    if search_string == None:
        search_string = ""

    company_filter_values["manager__id"] = manager.id

    #Вычисление номеров элементов в списке:
    last_index = current_page * elements_on_page
    first_index = last_index - elements_on_page
    companies = []
    #Получение списка выбранных компаний, задач и мероприятий:
    #Если есть сортировка, одни условия выбора, иначе - другие
    if(sort_by == "id" or sort_by == "ctype" or sort_by == "name" or sort_by == "city__name" or sort_by == "adress"):
        if(admin_mode):
            if (city_id != 0):
                companies = Company.objects.filter(Q(city_id=city_id), Q(name__icontains=search_string) | Q(
                    adress__icontains=search_string), Q(ctype__icontains=filter_type)).order_by(sort_by)[first_index:last_index]
            else:
                allowed_cities_list = get_worked_cities()
                if (show_id != 0):
                    allowed_cities_by_shows = set(
                        City.objects.filter(enabled=True,
                                            worked_shows__id=show_id).distinct().values_list(
                            "id", flat=True))
                    allowed_cities_list = list(set(allowed_cities_list) & allowed_cities_by_shows)
                companies = Company.objects.filter(Q(city__id__in = allowed_cities_list), Q(name__icontains=search_string) | Q(
                    adress__icontains=search_string), Q(ctype__icontains=filter_type)).order_by(sort_by)[first_index:last_index]
        else:
            companies = Company.objects.filter(Q(id__in = CMSILink.objects.filter(**company_filter_values).values("company__id").distinct()), Q(name__icontains = search_string) | Q(adress__icontains = search_string), Q(ctype__icontains=filter_type)).order_by(sort_by)[first_index:last_index]
    else:
        #Сортировка по звонкам
        if(sort_by == "call"):
            #Все выбранные компании
            if(admin_mode):
                if (city_id != 0):
                    companies = Company.objects.filter(Q(city_id=city_id), Q(name__icontains=search_string) | Q(
                        adress__icontains=search_string), Q(ctype__icontains=filter_type)).order_by(sort_by).values_list(
                    "id", flat=True)
                else:
                    companies = Company.objects.filter(Q(name__icontains=search_string) | Q(
                        adress__icontains=search_string), Q(ctype__icontains=filter_type)).order_by("city").values_list(
                    "id", flat=True)
            else:
                companies = Company.objects.filter(
                    Q(id__in=CMSILink.objects.filter(**company_filter_values).values("company__id").distinct()),
                    Q(name__icontains=search_string) | Q(adress__icontains=search_string),
                    Q(ctype__icontains=filter_type)).order_by("city").values_list(
                    "id", flat=True)
            #Из за ограничений базы данных объекты добавляются по 900 штук
            begin = 0
            end = 900
            #Поиск всех звонков и id компаний со звонками
            show_filter_values = {}
            if(show_id != 0):
                show_filter_values["artist__id"] = show_id
            companies_with_calls = list(Call.objects.filter(company__id__in=companies[begin:end], type = "company", **show_filter_values).values("company__id", "datetime").distinct())
            companies_with_calls_ids_only = list(Call.objects.filter(company__id__in=companies[begin:end], type = "company", **show_filter_values).values("company__id").distinct())

            if(len(companies) > 900):
                while (end < len(companies)):
                    begin += 900
                    end += 900
                    companies_with_calls.extend(list(Call.objects.filter(company__id__in=companies[begin : end], type="company").distinct().values("company__id", "datetime")))
                    companies_with_calls_ids_only.extend(list(Call.objects.filter(company__id__in=companies[begin:end], type="company").values("company__id").distinct()))
            #Сортировка компаний со звонками по дате
            companies_with_calls = sorted(companies_with_calls, key=lambda x: x["datetime"])
            sorted_companies_list = []

            #Вычленение идентификаторов с сортировкой
            for company in companies_with_calls:
                for id in companies_with_calls_ids_only:
                    if(id["company__id"] == company["company__id"]):
                        if(not id["company__id"] in sorted_companies_list):
                            sorted_companies_list.append(id["company__id"])
                        else:
                            sorted_companies_list.remove(id["company__id"])
                            sorted_companies_list.append(id["company__id"])
            #Идентификаторы компаний со звонками
            companies_with_calls_ids = [x["company__id"] for x in companies_with_calls]
            # Идентификаторы компаний без звонков
            companies_without_calls = list(set(companies) - set(sorted_companies_list))
            if(last_index > len(sorted_companies_list)):
                if(first_index > len(sorted_companies_list)):
                    without_first_index = first_index - len(sorted_companies_list)
                    without_last_index = last_index - len(sorted_companies_list)
                    companies = companies_without_calls[without_first_index:without_last_index]
                else:

                    with_last_index = len(sorted_companies_list)
                    companies = sorted_companies_list[first_index:with_last_index]

                    without_first_index = 0
                    without_last_index = last_index - with_last_index
                    companies.extend(companies_without_calls[without_first_index:without_last_index])
            else:
                companies = sorted_companies_list[first_index:last_index]

            presorted_companies = Company.objects.filter(id__in = companies)
            presorted_companies = dict([(obj.id, obj) for obj in presorted_companies])
            presorted_companies = [presorted_companies[id] for id in companies]
            companies = presorted_companies
        elif(sort_by == "task"):
            # Все выбранные компании
            if (admin_mode):
                if (city_id != 0):
                    companies = Company.objects.filter(Q(city_id=city_id), Q(name__icontains=search_string) | Q(
                        adress__icontains=search_string), Q(ctype__icontains=filter_type)).order_by(
                        sort_by).values_list(
                        "id", flat=True)
                else:
                    companies = Company.objects.filter(Q(name__icontains=search_string) | Q(
                        adress__icontains=search_string), Q(ctype__icontains=filter_type)).order_by("city").values_list(
                        "id", flat=True)
            else:
                companies = Company.objects.filter(
                    Q(id__in=CMSILink.objects.filter(**company_filter_values).values("company__id").distinct()),
                    Q(name__icontains=search_string) | Q(adress__icontains=search_string),
                    Q(ctype__icontains=filter_type)).order_by("city").values_list(
                    "id", flat=True)
            # Из за ограничений базы данных объекты добавляются по 900 штук
            begin = 0
            end = 900
            # Поиск всех звонков и id компаний со звонками
            task_filter_values = {}
            if (show_id != "0"):
                task_filter_values["artist__id"] = show_id
            if (not admin_mode):
                task_filter_values["manager"] = get_current_manager(request)
            companies_with_tasks = list(
                Task.objects.filter(company__id__in=companies[begin:end], **task_filter_values).values("company__id",
                                                                                 "datetime").distinct())
            companies_with_tasks_ids_only = list(
                Task.objects.filter(company__id__in=companies[begin:end], **task_filter_values).values("company__id").distinct())


            if (len(companies) > 900):
                while (end < len(companies)):
                    begin += 900
                    end += 900
                    companies_with_tasks.extend(list(
                        Task.objects.filter(company__id__in=companies[begin: end], **task_filter_values).distinct().values("company__id",
                                                                                                     "datetime")))
                    companies_with_tasks_ids_only.extend(list(
                        Task.objects.filter(company__id__in=companies[begin:end], **task_filter_values).values("company__id").distinct()))
            # Сортировка компаний со звонками по дате
            companies_with_tasks = sorted(companies_with_tasks, key=lambda x: x["datetime"])
            sorted_companies_list = []

            # Вычленение идентификаторов с сортировкой
            for company in companies_with_tasks:
                for id in companies_with_tasks_ids_only:
                    if (id["company__id"] == company["company__id"]):
                        if (not id["company__id"] in sorted_companies_list):
                            sorted_companies_list.append(id["company__id"])
                        else:
                            sorted_companies_list.remove(id["company__id"])
                            sorted_companies_list.append(id["company__id"])
            # Идентификаторы компаний со звонками
            companies_with_tasks_ids = [x["company__id"] for x in companies_with_tasks]
            # Идентификаторы компаний без звонков
            companies_without_tasks = list(set(companies) - set(sorted_companies_list))
            if (last_index > len(sorted_companies_list)):
                if (first_index > len(sorted_companies_list)):
                    without_first_index = first_index - len(sorted_companies_list)
                    without_last_index = last_index - len(sorted_companies_list)
                    companies = companies_without_tasks[without_first_index:without_last_index]
                else:

                    with_last_index = len(sorted_companies_list)
                    companies = sorted_companies_list[first_index:with_last_index]

                    without_first_index = 0
                    without_last_index = last_index - with_last_index
                    companies.extend(companies_without_tasks[without_first_index:without_last_index])
            else:
                companies = sorted_companies_list[first_index:last_index]

            presorted_companies = Company.objects.filter(id__in=companies)
            presorted_companies = dict([(obj.id, obj) for obj in presorted_companies])
            presorted_companies = [presorted_companies[id] for id in companies]
            companies = presorted_companies
        elif(sort_by == "event"):
            pass

    calls_filter_values = {}
    tasks_filter_values = {}
    events_filter_values = {}
    if(not admin_mode):
        tasks_filter_values["manager__id"] = manager.id
    if (show_id != 0):
        calls_filter_values["artist__id"] = show_id
        tasks_filter_values["artist__id"] = show_id
        events_filter_values["artist__id"] = show_id
        allowed_shows = [show_id, ]
    else:
        if (admin_mode):
            allowed_shows = Artist.objects.all().values_list("id", flat=True)
        else:
            getted_links = CMSILink.objects.filter(manager=manager, company__id__in = companies.values_list("id", flat=True)).values("show__id", "company__id").distinct()

    companies_ids = companies.values_list("id", flat=True)

    calls_all = list(Call.objects.filter(company__id__in = companies_ids, type="company", **calls_filter_values).order_by("-datetime").values("id", "company__id", "artist__id", "artist__name", "artist__color", "manager__siteuser__id", "manager__siteuser__alias", "datetime", "comment"))
    tasks_all = list(Task.objects.filter(company__id__in=companies_ids, **tasks_filter_values).values("id", "company__id", "artist__id", "artist__name", "artist__color", "manager__siteuser__id", "manager__siteuser__alias", "datetime", "done", "doneDateTime").order_by("-datetime"))
    events_all = list(Event.objects.filter(company__id__in=companies_ids, **events_filter_values).order_by("startTime").values("id", "manager__id", "company__id", "artist__id", "artist__name", "manager__siteuser__id", "manager__siteuser__alias", "artist__color", "startTime", "statsdt", "removed", "crashBool", "sumTransfered", "resultSum"))

    current_date_time = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    for company in companies:
        current_company_id = company.id
        if(not admin_mode and show_id == 0):
            allowed_shows = []
            for link in getted_links:
                if(link["company__id"] == current_company_id):
                    allowed_shows.append(link["show__id"])
        dict_item = {}
        for param in params_list:
            if(param == "id"):
                dict_item["id"] = company.id
            elif(param == "type"):
                dict_item["type"] = company.ctype
            elif(param == "name"):
                dict_item[param] = company.name
            elif(param == "city_name"):
                dict_item["city_name"] = company.city.name
            elif(param == "adress"):
                dict_item["adress"] = company.adress

            elif(param == "last_call"):
                choosen_company_calls = list(filter(lambda x: (x["company__id"] == current_company_id and x["artist__id"] in allowed_shows), calls_all))
                if(len(choosen_company_calls) != 0):
                    dict_item["last_call"] = choosen_company_calls[0]
                    dict_item["last_call"]["datetime"] = convert_datetime_to_javascript(dict_item["last_call"]["datetime"])
                    dict_item["call_comment"] = choosen_company_calls[0]["comment"]
                else:
                    dict_item["last_call"] = ""
                    dict_item["call_comment"] = ""
            elif(param == "last_task"):
                choosen_company_tasks = list(filter(lambda x: (x["company__id"] == current_company_id and x["artist__id"] in allowed_shows),tasks_all))
                if(len(choosen_company_tasks) != 0):
                    done_tasks = sorted(filter(lambda x: x["done"] == True, choosen_company_tasks), key=lambda y: y["doneDateTime"], reverse=True)
                    undone_tasks = sorted(filter(lambda x: x["done"] == False, choosen_company_tasks), key=lambda y: y["datetime"])

                    last_task = False
                    if(len(undone_tasks) > 0):
                        last_task = undone_tasks[0]
                    elif(len(done_tasks) > 0):
                        last_task = done_tasks[0]
                    if (last_task):
                        dict_item["last_task"] = last_task
                        if(last_task["done"]):
                            dict_item["last_task"]["datetime"] = convert_datetime_to_javascript(last_task["doneDateTime"])
                        else:
                            dict_item["last_task"]["datetime"] = convert_datetime_to_javascript(last_task["datetime"])
                    else:
                        dict_item["last_task"] = ""
                else:
                    dict_item["last_task"] = ""
            elif(param == "last_event"):
                choosen_company_events = list(filter(lambda x: (x["company__id"] == current_company_id and x["artist__id"] in allowed_shows), events_all))
                if (len(choosen_company_events) != 0):
                    before_today_events = list(sorted(filter(lambda x: x["startTime"] < current_date_time, choosen_company_events), key=lambda y: y["startTime"], reverse=True))
                    after_today_events = list(sorted(filter(lambda x: x["startTime"] > current_date_time, choosen_company_events), key=lambda y: y["startTime"]))

                    last_event = False
                    if (len(after_today_events) > 0):
                        last_event = after_today_events[0]
                    elif (len(before_today_events) > 0):
                        last_event = before_today_events[0]

                    if (last_event):
                        dict_item["last_event"] = last_event
                        dict_item["last_event"]["datetime"] = convert_datetime_to_javascript(last_event["startTime"])
                        dict_item["last_event"]["status"] = events_fill_events_statuses([last_event, ], request)[0]["status"]

                        dict_item["last_event"]["datetime"] = convert_datetime_to_javascript(
                        dict_item["last_event"]["startTime"])
                        dict_item["last_event"].pop("startTime")
                        dict_item["last_event"].pop("resultSum")
                        dict_item["last_event"].pop("sumTransfered")
                    else:
                        dict_item["last_event"] = ""


                else:
                    dict_item["last_event"] = ""

            elif(param == "is_have_email"):
                if company.email == None or company.email == "":
                    dict_item[param] = False
                else:
                    dict_item[param] = True
            elif(param == "is_have_site"):
                if company.site == None or company.site == "":
                    dict_item[param] = False
                else:
                    dict_item[param] = True
            elif(param == "is_have_contacts"):
                if company.contacts == None or company.contacts == "":
                    dict_item[param] = False
                else:
                    dict_item[param] = True
            elif(param == "is_have_telephone"):
                if company.telephone == None or company.telephone== "":
                    dict_item[param] = False
                else:
                    dict_item[param] = True
            elif(param == "is_have_comment"):
                if company.comment == None or company.comment== "":
                    dict_item[param] = False
                else:
                    dict_item[param] = True

        result["companies"] = filtred_companies_list
        filtred_companies_list.append(dict_item)

    result["manager__siteuser__id"] = SiteUser.objects.get(user = request.user).id
    return JsonResponse(create_response("data", "", result))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def companies_full_list(request):
    #"""Получение списка учреждений (при переданном в request-post: paginator, возвращает число страниц)"""
    options = get_user_options_from_db(request)
    elements_on_page = int(change_option_checker(request, "full_list_page_companies_count", request.POST.get("companies_count")))
    sort_by = request.POST.get("sort_by")
    filter_type = request.POST.get("filter_type")
    only_own = request.POST.get('only_own')
    only_free = request.POST.get('only_free')
    only_busy = request.POST.get('only_busy')

    if(request.POST.get("manager")):
        siteuser_id = request.POST.get("manager")
    else:
        siteuser_id = '0'

    allowed_cities_list = []
    if(options["admin_mode"] or options["can_show_all_companies"]):
        allowed_cities_list = get_worked_cities()
        if (request.POST.get("show_id") != "0"):
            allowed_cities_by_shows = set(
                City.objects.filter(enabled=True, worked_shows__id=request.POST.get("show_id")).distinct().values_list(
                    "id", flat=True))
            allowed_cities_list = list(set(allowed_cities_list) & allowed_cities_by_shows)
    else:
        allowed_cities_list = get_allowed_cities(request, _full_list=True, _id_list=True)
        if(request.POST.get("show_id") != "0"):
            allowed_cities_by_shows = set(City.objects.filter(enabled=True, worked_shows__id = request.POST.get("show_id")).distinct().values_list("id", flat=True))
            allowed_cities_list = list(set(allowed_cities_list) & allowed_cities_by_shows)
    paginator = request.POST.get("paginator")
    current_page = request.POST.get("page")
    if(current_page):
        current_page = int(current_page)
    city_id = request.POST.get("city_id")
    if (city_id != '0' and not int(city_id) in allowed_cities_list):
        return JsonResponse(create_response("error", "Этот город не доступен для просмотра"))
    show_id = request.POST.get("show_id")
    search_string = request.POST.get("search")
    manager = Manager.objects.get(siteuser__user=request.user)

    companies_filter_params = {}  # Словарь kwargs с параметрами для фильтрации учреждений
    if(only_own == "true"):
        companies_filter_params["manager__id"] = manager.id  # Добавление параметра фильтрации по менеджеру
    if (request.POST.get("city_id") != "0" and request.POST.get("city_id") != None):
        companies_filter_params["company__city__id"] = request.POST.get("city_id")  # Добавление параметра фильтрации по городу
    else:
        companies_filter_params["company__city__id__in"] = allowed_cities_list
    if (request.POST.get("show_id") != "0" and request.POST.get("show_id") != None):
        companies_filter_params["show__id"] = request.POST.get("show_id")  # Добавление параметра фильтрации по шоу
    if(filter_type) == "any":
        filter_type = ""
    if(siteuser_id != '0'):
        companies_filter_params["manager__id"] = get_manager_from_site_user(siteuser_id).id
    if(paginator):
        if(only_own == "true"):
            companies_count = Company.objects.filter(
                Q(id__in=CMSILink.objects.filter(**companies_filter_params).values("company__id").distinct()),
                Q(name__icontains=search_string) | Q(adress__icontains=search_string),
                Q(ctype__icontains=filter_type)).count()
        elif(only_free == "true"):
            busy_companies_count = Company.objects.filter(
                Q(id__in=CMSILink.objects.filter(**companies_filter_params).values("company__id").distinct()),
                Q(name__icontains=search_string) | Q(adress__icontains=search_string),
                Q(ctype__icontains=filter_type)).count()
            if (city_id != "0"):
                all_companies_count = Company.objects.filter(Q(city_id=city_id), Q(name__icontains=search_string) | Q(
                    adress__icontains=search_string), Q(ctype__icontains=filter_type)).count()
            else:
                all_companies_count = Company.objects.filter(Q(city__id__in = allowed_cities_list), Q(name__icontains=search_string) | Q(
                    adress__icontains=search_string), Q(ctype__icontains=filter_type)).count()
            companies_count = all_companies_count - busy_companies_count
        elif(only_busy == "true"):
            companies_count = Company.objects.filter(
                    Q(id__in=CMSILink.objects.filter(**companies_filter_params).values(
                        "company__id").distinct()),
                    Q(name__icontains=search_string) | Q(adress__icontains=search_string),
                    Q(ctype__icontains=filter_type)).count()

        else:
            if(siteuser_id != '0'):
                companies_count = Company.objects.filter(
                    Q(id__in=CMSILink.objects.filter(**companies_filter_params).values("company__id").distinct()),
                    Q(name__icontains=search_string) | Q(adress__icontains=search_string),
                    Q(ctype__icontains=filter_type)).count()
            else:
                if (city_id != "0"):
                    companies_count = Company.objects.filter(Q(city_id=city_id), Q(name__icontains=search_string) | Q(
                        adress__icontains=search_string), Q(ctype__icontains=filter_type)).count()
                else:
                    companies_count = Company.objects.filter(Q(city__id__in = allowed_cities_list), Q(name__icontains=search_string) | Q(
                    adress__icontains=search_string), Q(ctype__icontains=filter_type)).count()
        pages_count = calc_pages_count(elements_on_page, companies_count)
        return JsonResponse(create_response("data", "", {"page_count": pages_count}))

    counters = {}
    counters_all_filter_params = {}
    counters_busy_filter_params = {}
    if(city_id == '0'):
        counters_all_filter_params["city__id__in"] = allowed_cities_list
        counters_busy_filter_params["company__city__id__in"] = allowed_cities_list
    else:
        counters_all_filter_params["city__id"] = city_id
        counters_busy_filter_params["company__city__id"] = city_id
    if (show_id != '0'):
        counters_busy_filter_params["show__id"] = show_id

    counters["all"] = Company.objects.filter(**counters_all_filter_params).count()
    counters["busy"] = Company.objects.filter(id__in = CMSILink.objects.filter(**counters_busy_filter_params).values_list("company__id", flat=True).distinct()).count()
    counters["free"] = counters["all"] - counters["busy"]
    result = {}
    filtred_companies_list = []#Список компаний

    company_filter_values = {}#Параметры фильтрации для компаний

    #Список выводимых параметров
    params_list = {"id":"Код", "name":"Название", "city_name":"Город", "adress":"Адрес", "comment": "Заметка", "last_call": "Последний звонок"}

    #Заполнение фидьтрующего компании списка:
    if city_id != "0":
        company_filter_values["company__city__id"] = city_id
    if show_id != "0":
        company_filter_values["show__id"] = show_id
    if search_string == None:
        search_string = ""

    #Вычисление номеров элементов в списке:
    last_index = current_page * elements_on_page
    first_index = last_index - elements_on_page

    companies = []
    #Получение списка выбранных компаний, задач и мероприятий:
    #Если есть сортировка, одни условия выбора, иначе - другие
    if(sort_by == "id" or sort_by == "city__name"):
        if(only_own == "true"):
            companies = Company.objects.filter(
                Q(id__in=CMSILink.objects.filter(**companies_filter_params).values("company__id").distinct()),
                Q(name__icontains=search_string) | Q(adress__icontains=search_string),
                Q(ctype__icontains=filter_type))
        elif (only_busy == "true"):
            companies = Company.objects.filter(
                    Q(id__in=CMSILink.objects.filter(**companies_filter_params).values(
                        "company__id").distinct()),
                    Q(name__icontains=search_string) | Q(adress__icontains=search_string),
                    Q(ctype__icontains=filter_type))
        else:
            if (siteuser_id != '0'):
                companies_filter_params["manager__id"] = get_manager_from_site_user(siteuser_id).id
                companies = Company.objects.filter(
                    Q(id__in=CMSILink.objects.filter(**companies_filter_params).values("company__id").distinct()),
                    Q(name__icontains=search_string) | Q(adress__icontains=search_string),
                    Q(ctype__icontains=filter_type))
            elif(only_free == "true"):
                if (city_id != "0"):
                    all_companies = Company.objects.filter(Q(city_id=city_id), Q(name__icontains=search_string) | Q(
                        adress__icontains=search_string), Q(ctype__icontains=filter_type)).order_by(sort_by).values_list("id", flat=True)
                else:
                    all_companies = Company.objects.filter(Q(city__id__in = allowed_cities_list), Q(name__icontains=search_string) | Q(
                        adress__icontains=search_string), Q(ctype__icontains=filter_type)).order_by(sort_by).values_list("id", flat=True)
                busy_companies = Company.objects.filter(
                    Q(id__in=CMSILink.objects.filter(**companies_filter_params).values("company__id").distinct()),
                    Q(name__icontains=search_string) | Q(adress__icontains=search_string),
                    Q(ctype__icontains=filter_type)).values_list("id", flat=True)
                free_companies = list(set(all_companies) - set(busy_companies))
                companies = Company.objects.filter(id__in = free_companies[first_index:last_index])
            else:
                if (city_id != "0"):
                    companies = Company.objects.filter(Q(city_id=city_id), Q(name__icontains=search_string) | Q(
                        adress__icontains=search_string), Q(ctype__icontains=filter_type)).order_by(sort_by)
                else:
                    companies = Company.objects.filter(Q(city__id__in = allowed_cities_list), Q(name__icontains=search_string) | Q(
                        adress__icontains=search_string), Q(ctype__icontains=filter_type)).order_by(sort_by)

        if(not (only_free == "true")):
            companies = companies[first_index: last_index]
        else:
            companies = companies[0: elements_on_page]

    calls_filter_values = {}
    if (show_id != '0'):
        calls_filter_values["artist__id"] = show_id
        allowed_shows = [int_convertor_error_0(show_id), ]
    else:
        allowed_shows = Artist.objects.all().values_list("id", flat=True)

    companies_ids = companies.values_list("id", flat=True)

    calls_all = list(Call.objects.filter(company__id__in = companies_ids, type="company", **calls_filter_values).order_by("-datetime").values("id", "company__id", "artist__id", "artist__name", "artist__color", "manager__siteuser__id", "manager__siteuser__alias", "datetime", "comment"))

    for company in companies:
        current_company_id = company.id

        dict_item = {}

        manager_work_filter_values = {}
        if(show_id != '0'):
            manager_work_filter_values["artist__id"] = show_id
        else:
            allowed_shows = Artist.objects.all().values_list("id", flat = True)
            manager_work_filter_values["artist__in"] = allowed_shows

        manager_work_filter_values["company"] = company

        for param in params_list:
            if(param == "id"):
                dict_item["id"] = company.id
            if(param == "type"):
                dict_item["type"] = company.ctype
            if(param == "name"):
                dict_item[param] = company.name
            if(param == "city_name"):
                dict_item["city_name"] = company.city.name
            if (param == "adress"):
                dict_item["adress"] = company.adress
            if(param == "comment"):
                dict_item["comment"] = company.comment
            if (param == "last_call"):
                choosen_company_calls = list(filter(lambda x: (x["company__id"] == current_company_id and x["artist__id"] in allowed_shows), calls_all))
                if (len(choosen_company_calls) != 0):
                    dict_item["last_call"] = choosen_company_calls[0]
                    dict_item["last_call"]["datetime"] = convert_datetime_to_javascript(dict_item["last_call"]["datetime"])
                    dict_item["call_comment"] = choosen_company_calls[0]["comment"]
                else:
                    dict_item["last_call"] = ""
                    dict_item["call_comment"] = ""
        cms_list_search_params = {}
        cms_list_search_params["company__id"] = company.id
        if(show_id != '0'):
            cms_list_search_params["show__id"] = show_id
        unsorted_cms_list = list(CMSILink.objects.filter(**cms_list_search_params).values("manager__siteuser__id", "manager__siteuser__alias", "show__id", "show__color", "show__name").distinct())
        managers = []
        for cms in unsorted_cms_list:
            if not cms["manager__siteuser__id"] in managers:
                managers.append(cms["manager__siteuser__id"])
        sorted_managers_list = []
        for manager in managers:
            sorted_manager = {"id": manager}
            sorted_manager["shows"] = []
            for cms in unsorted_cms_list:
                manager_show_data = {}
                if(cms["manager__siteuser__id"] == manager):
                    if not "name" in sorted_manager:
                        sorted_manager["name"] = cms["manager__siteuser__alias"]
                    manager_show_data["show__id"] = cms["show__id"]
                    manager_show_data["show__color"] = cms["show__color"]
                    manager_show_data["show__name"] = cms["show__name"]
                    sorted_manager["shows"].append(manager_show_data)
            sorted_managers_list.append(sorted_manager)


        dict_item["managers"] = sorted_managers_list
        last_event = ChangeFieldLog.objects.filter(Q(link_to_object_id=company.id) | Q(table_link_id = company.id), Q(link_to_object_table__name = "Company") | Q(table__name__in = ["Company", "Call", "Task", "Event"]), changeType__name__in = ["add", "edit", "remove"]).order_by("-datetime").values("whoChange__id", "whoChange__alias", "datetime", "table__name", "table_link_id")
        if(last_event.count() == 0):
            dict_item["last_event"] = False
        else:
            dict_item["last_event"] = last_event[0]
            if(dict_item["last_event"]["table__name"] in ["Call", "Task", "Event"]):
                try:
                    if(dict_item["last_event"]["table__name"] == "Call"):
                        last_event_show = Artist.objects.get(id=Call.objects.get(id = dict_item["last_event"]["table_link_id"]).artist.id)
                    elif(dict_item["last_event"]["table__name"] == "Task"):
                        last_event_show = Artist.objects.get(id=Task.objects.get(id=dict_item["last_event"]["table_link_id"]).artist.id)
                    elif (dict_item["last_event"]["table__name"] == "Event"):
                        last_event_show = Artist.objects.get(id=Event.objects.get(id=dict_item["last_event"]["table_link_id"]).artist.id)
                    dict_item["last_event"]["show"] = {}
                    dict_item["last_event"]["show"]["color"] = last_event_show.color
                    dict_item["last_event"]["show"]["name"] = last_event_show.name
                except:
                    pass
        filtred_companies_list.append(dict_item)

    result["companies"] = filtred_companies_list
    result["manager__siteuser__id"] = SiteUser.objects.get(user = request.user).id
    result["counters"] = counters
    return JsonResponse(create_response("data", "", result))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def companies_full_list_make_companies_free(request):
    """Освобождение выбранного списка компаний от менеджеров"""
    try:
        companies_ids_list = json.loads(request.POST.get("companies"))
    except:
        return JsonResponse(create_response("error", "Ошибка при разборе полученных компаний"))

    options = get_user_options_from_db(request)

    cmslink_filter_params = {}
    if(request.POST.get("show_id") != '0'):
        cmslink_filter_params["show__id"] = request.POST.get("show_id")
    if(options["admin_mode"]):
        if(request.POST.get("siteuser_id") != '0'):
            cmslink_filter_params["manager__id"] = get_manager_from_site_user(request.POST.get("siteuser_id")).id
    else:
        cmslink_filter_params["manager__id"] = get_current_manager(request).id

    cmsi_links = CMSILink.objects.filter(company__id__in = companies_ids_list, **cmslink_filter_params).distinct()
    cmsi_links_managers = list(CMSILink.objects.filter(company__id__in = companies_ids_list, **cmslink_filter_params).values_list("manager__id", flat=True).distinct())
    if(cmsi_links.count() == 0):
        if(options["admin_mode"]):
            return JsonResponse(create_response("info", "Выбранные Вами компании не принадлежат ни одному из менеджеров"))
        else:
            return JsonResponse(create_response("info", "Выбранные Вами компании не принадлежат вам, невозможно освободить"))
    try:
        for link in cmsi_links:
            link.company.remove(*Company.objects.filter(id__in = companies_ids_list))
            Task.objects.filter(company__id__in = companies_ids_list, artist__id = link.show.id, manager__id = link.manager.id).delete()
    except:
        return JsonResponse(create_response("error", "Неизвестная ошибка при освобождении компаний, обратитесь к администратору"))
    try:
        for manager in cmsi_links_managers:
            mark_users_cities_changed(get_site_user_from_manager(manager).id)
    except:
        return JsonResponse(create_response("error", "Ошибка при пересчете списка городов"))
    return JsonResponse(create_response("success", "Выбранные компании успешно освобождены"))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def companies_full_list_transfer_companies_to(request):
    """Добавление выбранных компаний пользователю"""
    try:
        companies_ids = json.loads(request.POST.get("companies"))
        shows_ids = json.loads(request.POST.get("shows"))
        siteuser_id = request.POST.get("siteuser_id")
        if (request.POST.get("make_free") == "true"):
            make_free = True
        else:
            make_free = False
    except:
        JsonResponse(create_response("error", "Ошибка при разборе полученных данных"))
    companies = Company.objects.filter(id__in = companies_ids)
    shows = Artist.objects.filter(id__in = shows_ids)

    options = get_user_options_from_db(request)
    success_response_text = "Успешное добавление выбранных компаний пользователю"
    if (options["admin_mode"]):
        if(make_free):
            cmsi_links = CMSILink.objects.filter(company__id__in=companies_ids, show__id__in = shows_ids).distinct()
            try:
                for link in cmsi_links:
                    link.company.remove(*Company.objects.filter(id__in=companies_ids))
            except:
                return JsonResponse(create_response("error",
                                                    "Неизвестная ошибка при освобождении компаний, обратитесь к администратору"))
        manager = get_manager_from_site_user(siteuser_id)
    else:
        if(not options["can_add_companies_to_self"]):
            JsonResponse(create_response("error", "Нет доступа на добавление себе компаний"))
        manager = get_current_manager(request)
        excluded_companies_ids = CMSILink.objects.filter(show__id__in = shows_ids, company__id__in = companies_ids).exclude(manager = manager).values_list("company__id", flat=True).distinct()
        if(len(excluded_companies_ids) != 0):
            success_response_text = "Добавлены только свободные компании"
        companies = companies.exclude(id__in = excluded_companies_ids)
    have_changes = False
    for show in shows_ids:
        checker_links = CMSILink.objects.filter(show__id = show, manager = manager, company__id__in = companies_ids)
        if(not checker_links.count() == len(companies_ids)):
            have_changes = True
            break
    if(not have_changes):
        return JsonResponse(create_response("info", "Выбранные компании уже добавлены пользователю с выбранными шоу"))
    for show in shows:
        try:
            if(CMSILink.objects.filter(manager = manager, show=show).count() == 0):
                link = CMSILink(
                    manager = manager,
                    show = show
                )
                link.save()
                link.company.add(*companies)
            else:
                link = CMSILink.objects.get(manager = manager, show=show)
                link.company.add(*companies)
        except:
            return JsonResponse(create_response("error", "Ошибка при выдаче доступов пользователю"))
    if(not mark_users_cities_changed(get_site_user_from_manager(manager.id).id)):
        return JsonResponse(create_response("error", "Ошибка при пересчете списка городов"))
    return JsonResponse(create_response("success", success_response_text))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def companies_full_list_distribute_city(request):
    """Распределение города между менеджерами"""
    try:
        city_id = request.POST.get('city')
        shows = json.loads(request.POST.get('shows'))
        siteusers = json.loads(request.POST.get('siteusers'))
        if(request.POST.get("all_city") == "true"):
            all_city = True
        else:
            all_city = False
        companies_count = int(request.POST.get('companies_count'))
    except:
        JsonResponse(create_response("error", "Ошибка при разборе полученных данных"))

    all_companies_count = len(Company.objects.filter(city__id=city_id).values_list("id", flat=True))
    busy_companies_count = Company.objects.filter(id__in=CMSILink.objects.filter(show__id__in=shows, company__city__id=city_id).values_list("company__id", flat=True).distinct()).count()
    free_companies_count = all_companies_count - busy_companies_count
    if(not all_city):
        if((free_companies_count - (len(siteusers) * companies_count)) < 0):
            company_count_for_manager = int(free_companies_count / (len(siteusers)))
            company_remainder_count = int(free_companies_count % (len(siteusers)))
        else:
            company_count_for_manager = companies_count
            company_remainder_count = 0
    else:
        company_count_for_manager = int(free_companies_count / (len(siteusers)))
        company_remainder_count = int(free_companies_count % (len(siteusers)))

    free_companies = Company.objects.filter(city__id = city_id).exclude(id__in=CMSILink.objects.filter(show__id__in=shows, company__city__id=city_id).values_list("company__id", flat=True).distinct()).values_list("id", flat=True)
    count_low_range = 0
    count_high_range = company_count_for_manager
    maked_links_text = ""
    counter = 0
    for siteuser in siteusers:
        if(make_cms_link(siteuser, shows, free_companies[count_low_range: count_high_range])):
            maked_links_text += "| Создана связь " + SiteUser.objects.get(id = siteuser).name + ", шоу: " + str(list(Artist.objects.filter(id__in = shows).values_list("name", flat=True))) + ", компаний: " + str(len(free_companies[count_low_range: count_high_range]))
            if(counter == (len(siteusers) - 2)):
                count_high_range += company_remainder_count
            counter += 1
        else:
            return JsonResponse(create_response("error", "Ошибка при создании связи менеджер-шоу-компании. Не все связи созданы. Созданные связи:" + maked_links_text))
    try:
        for siteuser in siteusers:
            mark_users_cities_changed(siteuser)
    except:
        return JsonResponse(create_response("error", "Ошибка при пересчете списка городов"))
    return JsonResponse(create_response("success", maked_links_text))
#############################################################################################
#Calls:======================================================================================
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def calls_list(request):
    elements_on_page = 13

    calls_filter_values = {}

    if (request.POST.get("stats")):
        calls_filter_values["statsdt__gte"] = request.POST.get("from")
        calls_filter_values["statsdt__lte"] = request.POST.get("to")
        if(request.POST.get("city__id") != '0'):
            calls_filter_values["company__city__id"] = request.POST.get("city__id")
        if(request.POST.get("show_id") != '0'):
            calls_filter_values["artist__id"] = request.POST.get("show_id")
        calls_filter_values["manager__id"] = get_manager_from_site_user(request.POST.get("siteuser")).id
    else:
        # Если пользователь администратор, выводятся все звонки, в ином случае только по доступному шоу
        calls_filter_values["company"] = request.POST.get("company_id")
        if (not user_is_admin(request.user)):
            calls_filter_values["artist__in"] = get_allowed_manager_shows_for_company(
                get_manager_from_user(request.user), calls_filter_values["company"])
        else:
            calls_filter_values["artist__in"] = get_allowed_shows_for_city(_company=calls_filter_values["company"])
        show_id = request.POST.get("show_id")
        if (show_id != "0"):
            calls_filter_values["artist"] = show_id

    calls_filter_values["type"] = "company"
    if(request.POST.get("paginator")):
        pages_count = calc_pages_count(elements_on_page,  Call.objects.filter(**calls_filter_values).count())
        return JsonResponse(create_response("data", "", {"page_count": pages_count}))
    else:
        current_page = request.POST.get("page")
        if (current_page):
            current_page = int(current_page)
        last_index = current_page * elements_on_page
        first_index = last_index - elements_on_page
        calls = list(Call.objects.filter(**calls_filter_values).order_by("-datetime").values("id", "artist__name", "manager__id", "artist__id", "manager__siteuser__alias", "datetime", "comment", "artist__color"))[first_index:last_index]
        current_manager = get_current_manager(request).id
        for call in calls:
            if(call["manager__id"]) == current_manager:
                call["own"] = True
            else:
                call["own"] = False
            call.pop("manager__id")
    return JsonResponse(create_response("data", "", calls))
@user_is_auth_aj_checker
def call_get(request):
    id = request.POST.get('id')
    current_user_type = get_current_user_type(request)
    try:
        call = Call.objects.filter(id = id).values("id", "company__id", "company__name", "company__id", "manager__siteuser__id", "manager__siteuser__alias", "datetime", "artist__name", "artist__color", "comment")[0]
    except:
        return JsonResponse(create_response("error", "Звонок с таким идентификатором не найден"))
    call_object = Call.objects.get(id = request.POST.get('id'))

    event_call = False
    if(Event.objects.filter(Q(callDay__id = call_object.id) | Q(callWeek__id = call_object.id) | Q(callMonth__id = call_object.id)).count() != 0):
        event_call = True

    if (current_user_type == "presentator" and not event_call):
        return JsonResponse(create_response("info", "Артистам запрещено просматривать звонки к учреждениям"))
    elif(current_user_type == "manager"):
        if (call_object.manager.id != get_current_manager(request).id and call_object.type == "event"):
            return JsonResponse(create_response("info", "Невозможно просматривать чужие отзвоны по мероприятиям"))
    if(user_is_admin(request.user)):
        call["admin_mode"] = user_is_admin(request.user)
    return JsonResponse(create_response("data", "", call))
#Tasks
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def tasks_list(request):
    user_type = get_current_user_type(request)
    elements_on_page = 13
    paginator = request.POST.get("paginator")

    tasks_filter_values = {}

    if (request.POST.get("stats")):
        tasks_filter_values["statsdt__gte"] = request.POST.get("from")
        tasks_filter_values["statsdt__lte"] = request.POST.get("to")
        if(request.POST.get("city__id") != '0'):
            tasks_filter_values["company__city__id"] = request.POST.get("city__id")
        if(request.POST.get("show_id") != '0'):
            tasks_filter_values["artist__id"] = request.POST.get("show_id")
        tasks_filter_values["manager__id"] = get_manager_from_site_user(request.POST.get("siteuser")).id
    else:
        try:
            company = Company.objects.get(id=request.POST.get("company_id"))
        except:
            return JsonResponse(create_response("error", "Не найдена компания с таким идентификатором"))
        tasks_filter_values["company__id"] = company.id
        if(request.POST.get("show_id") != "0"):
            tasks_filter_values["artist__id"] = request.POST.get("show_id")
        elif(user_type == "admin"):
            tasks_filter_values["artist__in"] = get_allowed_shows_for_city(_company = company.id)
        else:
            tasks_filter_values["artist__in"] = get_manager_allowed_shows_for_city(request, company.city.id)
        if(user_type != "admin"):
            tasks_filter_values["manager__id"] = get_current_manager(request).id
    if (paginator):
        pages_count = calc_pages_count(elements_on_page, Task.objects.filter(**tasks_filter_values).count())
        return JsonResponse(create_response("data", "", {"page_count": pages_count}))
    else:
        current_page = request.POST.get("page")
        if (current_page):
            current_page = int(current_page)
        last_index = current_page * elements_on_page
        first_index = last_index - elements_on_page

        tasks = list(Task.objects.filter(**tasks_filter_values).order_by("-datetime").values("id", "artist__name", "manager__siteuser__alias", "datetime", "statsdt", "description", "artist__color", "done", "doneDateTime"))[first_index:last_index]
        return JsonResponse(create_response("data", "", tasks))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def tasks_get_task(request):
    user_type = get_current_user_type(request)
    options = get_user_options_from_db(request)
    if(user_type != "admin"):
        if(Task.objects.filter(id = request.POST.get("id"), manager= get_current_manager(request)).count() == 0):
            return JsonResponse(create_response("error", "Невозможно просмотреть чужую задачу"))

    task = Task.objects.filter(id=request.POST.get("id")).values("id", "done", "description", "comment", "manager__siteuser__id",
                                                                         "manager__siteuser__alias", "datetime", "company__id",
                                                                         "company__name", "artist__id", "artist__name",
                                                                         "artist__color", "type", "statsdt", "doneDateTime", "assigned_by__name")
    if (task.count() == 0):
        return JsonResponse(create_response("error", "Не найдена задача с таким идентификатором"))
    response = {}
    response["admin_mode"] = options["admin_mode"]
    response["task_data"] = list(task)[0]
    manager_work_filter_values = {}

    if (get_current_user_type(request) == "admin"):
        allowed_shows = Artist.objects.all().values_list("id", flat=True)
    else:
        allowed_shows = CMSILink.objects.filter(company__id = task[0]["company__id"], manager=get_current_manager(request)).values("show_id").distinct()
        manager_work_filter_values["artist__in"] = allowed_shows

    manager_work_filter_values["company__id"] = task[0]["company__id"]

    if(Call.objects.filter(type="company", **manager_work_filter_values).count() == 0):
        response["last_call"] = False
    else:
        response["last_call"] = list(Call.objects.filter(type="company", **manager_work_filter_values).order_by("-datetime")[:1].values("id"))[0]
    if(task[0]["type"] == "event"):
        response["event_id"] = Event.objects.filter(Q(taskDay__id = request.POST.get("id")) | Q(taskWeek__id = request.POST.get("id")) | Q(taskMonth__id = request.POST.get("id"))).values_list("id", flat=True)[0]
    return JsonResponse(create_response("data", "", response))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def tasks_get_in_daterange(request):
    # Вовращаемое значение, в формате JSON. Контейнер
    site_user = get_current_site_user(request, "object")
    current_user_type = get_current_user_type(request)
    result = {}
    filter_values = {}
    current_manager_id = 0

    count_on_page = 20

    paginator = request.POST.get('paginator')

    search_string = ""

    cal_type = request.POST.get("cal_type")
    cal_option = request.POST.get("cal_option")


    if (request.POST.get("search")):
        search_string = request.POST.get("search")
    city_id = int_convertor_error_json_response(request.POST.get("city"))
    show_id = int_convertor_error_json_response(request.POST.get("show"))

    # Если есть диапазоны, они указываются, в противном случае - указывается менеджер, для полного списка мероприятий
    if(site_user.options.hide_done_tasks):
        filter_values["done"] = False
    if (request.POST.get("to") != "false"):
        filter_values["datetime__lte"] = request.POST.get("to")
    if (request.POST.get("from") != "false"):
        filter_values["datetime__gte"] = request.POST.get("from")
    else:
        if (cal_type == "fulllist"):
            month_count = site_user.options.full_list_tasks_month_count
            filter_values["datetime__gte"] = datetime.now() - monthdelta(month_count)
    if (request.POST.get("page")):
        shift = (int_convertor_error_json_response(request.POST.get("page")) - 1) * count_on_page
        tasks_from = shift
        tasks_to = shift + count_on_page
    else:
        tasks_from = 0
        tasks_to = 5000


    # Параметры фильтрации по ролям
    pairs = None
    if (current_user_type == "admin"):
        if(site_user.options.only_own_tasks_for_admin):
            try:
                current_manager_id = get_current_manager(request).id
            except:
                current_manager_id = 0
            filter_values["manager__id"] = current_manager_id
        if (city_id != 0):
            filter_values["company__city__id"] = city_id
            if (show_id == 0):
                filter_values["artist__id__in"] = get_worked_shows(city_id, _list=True)
            else:
                filter_values["artist__id"] = show_id
        else:
            if(show_id != 0):
                pairs = get_allowed_show_city_pairs(request, show_id)
            else:
                pairs = get_allowed_show_city_pairs(request)
    elif (current_user_type == "manager"):
        # Если нет менеджера, указывается несуществующий
        try:
            current_manager_id = get_current_manager(request).id
        except:
            current_manager_id = 0

        filter_values["manager__id"] = current_manager_id

        if (city_id != 0):
            filter_values["company__city__id"] = city_id
            if (show_id == 0):
                filter_values["artist__id__in"] = get_manager_allowed_shows_for_city(request, city_id)
            else:
                filter_values["artist__id"] = show_id
        else:
            if(show_id != 0):
                pairs = get_allowed_show_city_pairs(request, show_id)
            else:
                pairs = get_allowed_show_city_pairs(request)
    else:
        return JsonResponse(create_response("error", "Не определен текущий тип пользователя"))

    tasks = Task.objects.filter(**filter_values)

    if(search_string != ""):
        tasks = tasks.filter((Q(company__name__icontains = search_string) | Q(company__adress__icontains = search_string)))

    if (pairs):
        show_city_Qfilters_and_list = []
        for pair in pairs:
            Qfiltred = [Q(x) for x in pair]
            show_city_Qfilters_and_list.append(functools.reduce(operator.and_, Qfiltred))
        Qfiltred = [Q(x) for x in show_city_Qfilters_and_list]
        show_city_Qfilters_or_list = functools.reduce(operator.or_, Qfiltred)
        tasks = tasks.filter(show_city_Qfilters_or_list)


    if (paginator):
        data = {}
        data["page_count"] = calc_pages_count(count_on_page, tasks.count())
        return JsonResponse(create_response("data", "", data))

    else:
        try:
            tasks = replace_datetime_to_string(list(tasks.order_by("datetime").order_by("datetime").values("id", "company__id", "company__name", "manager__siteuser__alias", "company__city__name", "company__adress",
                                            "artist__id", "artist__name", "artist__color", "manager__id", "description", "datetime", "done", "doneDateTime", "comment", "type")[tasks_from:tasks_to]))
        except:
            return JsonResponse(create_response("info",
                                                "Слишком много задач для запроса, уточните параметры: город, шоу"))
    #Calculate expired tasks
    if("datetime__gte" in filter_values):
        filter_values.pop("datetime__gte")
    filter_values["done"] = False
    if(current_manager_id != 0):
        filter_values["manager__id"] = current_manager_id

    current_date_time = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
    filter_values["datetime__lte"] = current_date_time

    non_qfiltred_expired_tasks = Task.objects.filter(**filter_values)

    if(pairs):
        show_city_Qfilters_and_list = []
        for pair in pairs:
            Qfiltred = [Q(x) for x in pair]
            show_city_Qfilters_and_list.append(functools.reduce(operator.and_, Qfiltred))
        Qfiltred = [Q(x) for x in show_city_Qfilters_and_list]
        show_city_Qfilters_or_list = functools.reduce(operator.or_, Qfiltred)
        expired_tasks = non_qfiltred_expired_tasks.filter(show_city_Qfilters_or_list)
    else:
        expired_tasks = non_qfiltred_expired_tasks


    result["events"] = tasks
    result["expired_tasks"] = list(expired_tasks.order_by("datetime").values("id", "company__id", "company__name", "manager__siteuser__alias", "company__city__name", "company__adress",
                                            "artist__id", "artist__name", "artist__color", "manager__id", "description", "datetime", "done", "doneDateTime", "comment", "type"))
    return JsonResponse(create_response("data", "", result, request))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def tasks_get_frame_counters(request):
    site_user = get_current_site_user(request, "object")

    current_user_type = get_current_user_type(request)
    result = {}
    filter_values = {}
    current_manager_id = 0


    city_id = int_convertor_error_json_response(request.POST.get("city"))
    show_id = int_convertor_error_json_response(request.POST.get("show"))

    if (site_user.options.hide_done_tasks):
        filter_values["done"] = False
    # Параметры фильтрации по ролям
    pairs = None
    if (current_user_type == "admin"):
        if (site_user.options.only_own_tasks_for_admin):
            try:
                current_manager_id = get_current_manager(request).id
            except:
                current_manager_id = 0
            filter_values["manager__id"] = current_manager_id
        if (city_id != 0):
            filter_values["company__city__id"] = city_id
            if (show_id == 0):
                filter_values["artist__id__in"] = get_worked_shows(city_id, _list=True)
            else:
                filter_values["artist__id"] = show_id
        else:
            if (show_id != 0):
                pairs = get_allowed_show_city_pairs(request, show_id)
            else:
                pairs = get_allowed_show_city_pairs(request)
    elif (current_user_type == "manager"):
        # Если нет менеджера, указывается несуществующий
        try:
            current_manager_id = get_current_manager(request).id
        except:
            current_manager_id = 0

        filter_values["manager__id"] = current_manager_id

        if (city_id != 0):
            filter_values["company__city__id"] = city_id
            if (show_id == 0):
                filter_values["artist__id__in"] = get_manager_allowed_shows_for_city(request, city_id)
            else:
                filter_values["artist__id"] = show_id
        else:
            if (show_id != 0):
                pairs = get_allowed_show_city_pairs(request, show_id)
            else:
                pairs = get_allowed_show_city_pairs(request)
    else:
        return JsonResponse(create_response("error", "Не определен текущий тип пользователя"))

    non_qfiltred_tasks_for_counters = Task.objects.filter(**filter_values)
    if (pairs):
        show_city_Qfilters_and_list = []
        for pair in pairs:
            Qfiltred = [Q(x) for x in pair]
            show_city_Qfilters_and_list.append(functools.reduce(operator.and_, Qfiltred))
        Qfiltred = [Q(x) for x in show_city_Qfilters_and_list]
        show_city_Qfilters_or_list = functools.reduce(operator.or_, Qfiltred)
        tasks_for_counters = non_qfiltred_tasks_for_counters.filter(show_city_Qfilters_or_list)
    else:
        tasks_for_counters = non_qfiltred_tasks_for_counters

    truncate_date = connection.ops.date_trunc_sql('day', 'datetime')
    extra_tasks = tasks_for_counters.extra({'day': truncate_date})
    result["counters"] = list(extra_tasks.values('day').annotate(Count('pk')).order_by('day'))
    return JsonResponse(create_response("data", "", result, request))

@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def events_load_presentators_list(request):
    city_id = int_convertor_error_0(request.POST.get("city_id"))
    show_id = int_convertor_error_0(request.POST.get("show_id"))
    allowed_presentators_filter_params = {}

    if(city_id != 0):
        allowed_presentators_filter_params["city__id"] = city_id

    if (get_current_user_type(request) == "manager"):
        if(show_id != 0):
            allowed_presentators_filter_params["shows__id__in"] = [show_id,]
        else:
            allowed_presentators_filter_params["shows__id__in"] = get_manager_allowed_shows_for_city(request, city_id)
    else:
        if (show_id != 0):
            allowed_presentators_filter_params["shows__id__in"] = [show_id, ]
        else:
            allowed_presentators_filter_params["shows__id__in"] = get_allowed_shows_for_city(city_id).values_list("id", flat=True)
    try:
        allowed_presentators = list(Presentator.objects.filter(id__in=PCSLink.objects.filter(**allowed_presentators_filter_params).values_list(
            "presentator__id", flat=True)).order_by("siteuser__name").values("id", "siteuser__alias"))
    except:
        return JsonResponse(create_response("error", "Ошибка при получении данных артистов мероприятий"))
    return JsonResponse(create_response("data", "", allowed_presentators, _nodatalabel=True))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def events_list(request):
    user_type = get_current_user_type(request)
    site_user = get_current_site_user(request)
    elements_on_page = 13
    paginator = request.POST.get("paginator")


    events_filter_values = {}
    if (request.POST.get("stats")):
        events_filter_values["startTime__gte"] = request.POST.get("from")
        events_filter_values["startTime__lte"] = request.POST.get("to")
        if(request.POST.get("city__id") != '0'):
            events_filter_values["company__city__id"] = request.POST.get("city__id")
        if(request.POST.get("show_id") != '0'):
            events_filter_values["artist__id"] = request.POST.get("show_id")
        events_filter_values["manager__id"] = get_manager_from_site_user(request.POST.get("siteuser")).id
        if(get_current_manager(request).id != events_filter_values["manager__id"] and user_type == "manager"):
            return JsonResponse(create_response("info", "Невозможно просматривать список мероприятий чужого менеджера"))
    else:
        try:
            company = Company.objects.get(id=request.POST.get("company_id"))
        except:
            return JsonResponse(create_response("error", "Не найдена компания с таким идентификатором"))
        if (user_type == "manager" and (not company.id in get_manager_allowed_companies(request))):
            return JsonResponse(create_response("error", "Запрещен просмотр истории мероприятий к чужой компании"))
        events_filter_values["company__id"] = company.id
        if (request.POST.get("show_id") != "0"):
            events_filter_values["artist__id"] = request.POST.get("show_id")
        elif (user_type == "admin"):
            events_filter_values["artist__in"] = get_allowed_shows_for_city(_company=company.id)
        else:
            events_filter_values["artist__in"] = get_manager_allowed_shows_for_city(request, company.city.id)
    if (paginator):
        pages_count = calc_pages_count(elements_on_page, Event.objects.filter(**events_filter_values).count())
        return JsonResponse(create_response("data", "", {"page_count": pages_count}))
    else:
        current_page = request.POST.get("page")
        if (current_page):
            current_page = int(current_page)
        last_index = current_page * elements_on_page
        first_index = last_index - elements_on_page

        unprocessed_events = Event.objects.filter(**events_filter_values).order_by("-startTime").values(*events_get_values_list_for_calc_statuses())[first_index:last_index]
        events = events_fill_events_statuses(unprocessed_events[first_index:last_index], request)
        allowed_events_ids = []
        for event in events:
            if(event["status"] != "alien"):
                allowed_events_ids.append(event["id"])

        allowed_events = Event.objects.filter(id__in = allowed_events_ids).values("id", "startTime", "statsdt", "resultSum", "manager__siteuser__id")
        for event in events:
            if(event["manager__siteuser__id"] == site_user):
                event["own"] = True
            else:
                event["own"] = False
            for allowed_event in allowed_events:
                if(event["id"] == allowed_event["id"]):
                    event["data"] = allowed_event

        return JsonResponse(create_response("data", "", events))
@user_is_auth_aj_checker
def events_get_event_data(request):
    current_user_type = get_current_user_type(request)
    current_user = get_current_site_user(request, "object")

    now_date_time = datetime.now()
    result = {}
    response = {}

    try:
        event = Event.objects.get(id = request.POST.get("id"))
        event_id = event.id
    except:
        return JsonResponse(create_response("error", "Не удалось найти мероприятие с таким идентификатором"))
    if(current_user_type == "manager" or current_user_type == "admin"):
        manager = get_current_manager(request).id
        if(not current_user_type == "admin"):
            if (not(event.manager.id == manager)):
                manager_options = get_user_options_from_db(request)
                if(not(event.company.city.id in get_manager_allowed_cities(manager) and manager_options["can_show_allowed_cities_events"])):
                    return JsonResponse(create_response("info", "Невозможно просмотреть чужое мероприятие"))
    elif(current_user_type == "presentator"):
        events_limit = get_user_options_from_db(request)["presentator_show_events_limit"]
        if(event.startTime > datetime.now()):
            if((event.startTime - datetime.now()).total_seconds()/60/60/24/7 > events_limit):
                return JsonResponse(create_response("info", "Информация о мероприятии появится за " + str(events_limit) + " недели до начала мероприятия"))
    ###Частичное получение данных - только данные компании для артиста либо только данные о звонках
    if(request.POST.get("only_artist_data")):
        try:
            if(request.POST.get("manager_data") != "false"):
                if (user_is_admin(request.user) or Event.objects.filter(id=request.POST.get("id"), manager=get_current_manager(request)).count() > 0):
                    data = Event.objects.filter(id=request.POST.get("id")).values("company__id",
                                                                                  "company__name",
                                                                                  "company__city__name",
                                                                                  "company__ctype",
                                                                                  "company__adress",
                                                                                  "company__contacts",
                                                                                  "company__telephone",
                                                                                  "company__site",
                                                                                  "company__email",
                                                                                  "company__comment"
                                                                                  )[0]
                    result = create_response("data", "Успешная загрузка данных", data)
                else:
                    result = create_response("info", "Невозможно просматривать данные чужих компаний")
            else:
                if(user_is_admin(request.user) or user_is_presentator(request.user)):
                    data = Event.objects.filter(id = request.POST.get("id")).values("companyName", "companyAdress", "companyContacts", "company__city__name")[0]
                    result = create_response("data", "Успешная загрузка данных", data)
                else:
                    if(Event.objects.filter(id = request.POST.get("id"), manager = get_current_manager(request)).count() > 0):
                        data = Event.objects.filter(id = request.POST.get("id")).values("companyName", "companyAdress", "companyContacts", "company__city__name")[0]
                        result = create_response("data", "Успешная загрузка данных", data)
                    else:
                        result = create_response("info", "Невозможно просматривать данные чужих компаний")
        except:
            result = create_response("error", "Неизвестная ошибка при получении данных мероприятия. Перезагрузите страницу и попробуйте снова, при повторном возникновении ошибки обратитесь к администратору сайта")
        return JsonResponse(result)
    if(request.POST.get("only_calls_data")):
        calls_data = {}
        event_data = Event.objects.filter(id = request.POST.get("id")).values("callDay", "callWeek", "callMonth", "taskDay", "taskWeek", "taskMonth")[0]
        if(event_data["callDay"]):
            calls_data["day"] = "done"
        elif(event_data["taskDay"]):
            calls_data["day"] = "need"
        else:
            calls_data["day"] = "notneed"

        if (event_data["callWeek"]):
            calls_data["week"] = "done"
        elif (event_data["taskWeek"]):
            calls_data["week"] = "need"
        else:
            calls_data["week"] = "notneed"

        if (event_data["callMonth"]):
            calls_data["month"] = "done"
        elif (event_data["taskMonth"]):
            calls_data["month"] = "need"
        else:
            calls_data["month"] = "notneed"

        return JsonResponse(create_response("data", "", calls_data))
    ####################################################################################################################################


    if(current_user_type == "manager" or current_user_type == "admin"):
        event = Event.objects.filter(id=request.POST.get("id")).values("company__name", "company__id", "manager__id", "manager__siteuser__alias", "manager__siteuser__id",
                                                                       "company__city__name", "id", "companyName",
                                                                       "companyContacts", "companyAdress", "artist__id",
                                                                       "artist__name", "artist__color",
                                                                       "startTime", "crash", "crashBool", "sumTransfered", "resultSum",
                                                                       "price", "childCount", "percent", "removed",
                                                                       "removedDescription", "note", "sumPercent", "artistNote", "statsdt")

        converted_event = replace_datetime_to_string((event))[0]
        #Добавление значений Filled некоторых параметров - для определения статуса мероприятия
        if(converted_event["sumTransfered"] != None and converted_event["sumTransfered"] != 0):
            converted_event["sumTransferedFilled"] = True
        else:
            converted_event["sumTransferedFilled"] = False
        if (converted_event["resultSum"] != None and converted_event["resultSum"] != 0):
            converted_event["resultSumFilled"] = True
        else:
            converted_event["resultSumFilled"] = False


        response = replace_datetime_to_string((event))[0]

        converted_event["own"] = True
        converted_event["allowed"] = True
        response["status"] = events_fill_event_status(response, now_date_time, current_user_type)
        result = create_response("data", "", response, request)
    elif(current_user_type == "presentator"):
        event = Event.objects.filter(id = request.POST.get("id")).values("company__city__name", "id",
                                                                         "companyName", "companyContacts", "companyAdress",
                                                                         "artist__id", "artist__name", "artist__color",
                                                                         "manager__siteuser__alias", "startTime",
                                                                         "price", "childCount", "resultSum", "sumTransfered",
                                                                         "crashBool", "crash", "removed", "removedDescription", "note", "percent", "artistNote", "statsdt")
        response = replace_datetime_to_string((event))[0]
        response["own"] = True
        response["allowed"] = True
        #Добавление значений Filled некоторых параметров - для определения статуса мероприятия
        if(response["sumTransfered"] != None and response["sumTransfered"] != 0):
            response["sumTransferedFilled"] = True
        else:
            response["sumTransferedFilled"] = False
        if (response["resultSum"] != None and response["resultSum"] != 0):
            response["resultSumFilled"] = True
        else:
            response["resultSumFilled"] = False
        response["status"] = events_fill_event_status(response, now_date_time, current_user_type)

        result = create_response("data", "", response, request)


    # calls data
    event_for_call_data = Event.objects.get(id=request.POST.get("id"))
    calls_data = {}
    response["events_calls_data"] = calls_data

    # month
    call_data = {}
    if (event_for_call_data.callMonth):
        call_data["status"] = "done"
        call_data["id"] = event_for_call_data.callMonth.id
        call_data["result"] = event_for_call_data.callMonth.comment
    elif (event_for_call_data.taskMonth):
        if (current_user_type == "presentator"):
            call_data["status"] = "presentatorwait"
        else:
            call_data["status"] = "need"
            call_data["id"] = event_for_call_data.taskMonth.id
            call_data["result"] = event_for_call_data.taskMonth.description
    else:
        call_data["status"] = "notneed"
    calls_data["month"] = call_data

    # week
    call_data = {}
    if (event_for_call_data.callWeek):
        call_data["status"] = "done"
        call_data["id"] = event_for_call_data.callWeek.id
        call_data["result"] = event_for_call_data.callWeek.comment
    elif (event_for_call_data.taskWeek):
        if (current_user_type == "presentator"):
            call_data["status"] = "presentatorwait"
        else:
            call_data["status"] = "need"
            call_data["id"] = event_for_call_data.taskWeek.id
            call_data["result"] = event_for_call_data.taskWeek.description
    else:
        call_data["status"] = "notneed"
    calls_data["week"] = call_data

    # day
    call_data = {}
    if (event_for_call_data.callDay):
        call_data["status"] = "done"
        call_data["id"] = event_for_call_data.callDay.id
        call_data["result"] = event_for_call_data.callDay.comment
    elif (event_for_call_data.taskDay):
        if(current_user_type == "presentator"):
            call_data["status"] = "presentatorwait"
        else:
            call_data["status"] = "need"
            call_data["id"] = event_for_call_data.taskDay.id
            call_data["result"] = event_for_call_data.taskDay.description
    else:
        call_data["status"] = "notneed"
    calls_data["day"] = call_data

    #Сообщения чата - последнее сообщение артиста и менеджера, число непрочитанных сообщений
    last_artist_chat_message = ChatMessage.objects.filter(table_link_id = request.POST.get("id"), table = get_table_object_from_name("Event"), sender__type = "p").order_by("-datetime")
    if(last_artist_chat_message.count() > 0):
        response["last_artist_chat_message"] = last_artist_chat_message.values("message", "sender__alias")[0]
    else:
        response["last_artist_chat_message"] = ""

    last_manager_chat_message = ChatMessage.objects.filter(Q(sender__type = "m") | Q(sender__type = "a"), table_link_id = request.POST.get("id"), table = get_table_object_from_name("Event")).order_by("-datetime")
    if(last_manager_chat_message.count() > 0):
        response["last_manager_chat_message"] = last_manager_chat_message.values("message", "sender__alias")[0]
    else:
        response["last_manager_chat_message"] = ""

    response["unread_messages_count"] = ChatMessage.objects.filter(table_link_id = request.POST.get("id"), table = get_table_object_from_name("Event")).exclude(viewers__id = current_user.id).count()

    allowed_tables = ["Call", "Event"]
    allowed_linked_tables = ["Event"]
    allowed_actions = []
    if (current_user_type == "admin"):
        allowed_actions = ["add", "edit", "remove", "view", "transfer", "transference", "permission"]
    else:
        allowed_actions = ["add", "edit", "remove", "view", "transfer", "transference"]

    response['logs_count'] = ChangeFieldLog.objects.filter(Q(table_link_id=event_id) | Q(link_to_object_id=event_id), Q(
        link_to_object_table__name__in=allowed_linked_tables) | Q(table__name__in=allowed_tables), changeType__name__in = allowed_actions).exclude(whoViewChange__id=current_user.id).count()
    response["options"] = get_user_options_from_db(request)
    return JsonResponse(result)
@user_is_auth_aj_checker
def events_edit_event(request):
    event_id = int_convertor_error_json_response(request.POST.get("id"))
    try:
        check_event = Event.objects.get(id=event_id)
    except:
        return JsonResponse(create_response("error", "Мероприятие с таким идентификатором не найдено"))

    current_user_type = get_current_user_type(request)
    if(current_user_type == "manager"):
        current_manager = get_current_manager(request)
        if(current_manager.id != check_event.manager.id):
            return JsonResponse(create_response("info", "Невозможно редактировать чужое мероприятие"))
    if(request.POST.get("crash") or request.POST.get("remove") or request.POST.get("transfer")):
        if (current_user_type == "presentator"):
            return JsonResponse(create_response("error", "Только менеджеры и администраторы могут редактировать состояние мероприятия"))
        else:
            if(request.POST.get("crash")):
                if(request.POST.get("abort_crash")):
                    check_event.crashBool = False
                    write_to_change_log(check_event.id, "mark", "Event", "crashBool",
                                        "Отмечена отмена слета мероприятия",
                                        get_current_site_user(request))
                    success_text = "Успешная отметка отмены слета мероприятия"
                else:
                    check_event.crashBool = True
                    check_event.crash = request.POST.get("crashText")
                    write_to_change_log(check_event.id, "mark", "Event", "crashBool",
                                        "Отмечен слет мероприятия. Причина: " + check_event.crash,
                                        get_current_site_user(request))
                    success_text = "Успешная отметка слета мероприятия"
                    if(bool_convertor_error_json_response(request.POST.get("new_task"))):
                        try:
                            task = Task(
                                manager=get_current_manager(request),
                                company=check_event.company,
                                artist=check_event.artist,
                                datetime=request.POST.get('task_datetime'),
                                description=request.POST.get('task_description'),
                                comment=request.POST.get('task_comment'),
                                done=False
                            )
                            task.save()
                            success_text += " и добавление задачи"
                        except:
                            return JsonResponse(create_response("error", "Ошибка при добавлении задачи"))
                        write_to_change_log(task.id, "add", "Task", "description", task.description,
                                    get_current_site_user(request), check_event.id, "Company")
                check_event.save()

                return JsonResponse(create_response("success",
                                                    success_text))
            elif (request.POST.get("remove")):
                if(not bool_convertor_error_json_response(request.POST.get("remove_confirm"))):
                    return JsonResponse(create_response("error", "Удаление не подтверждено. Для подтверждения используйте переключатель"))

                check_event.removed = True
                check_event.removedDescription = request.POST.get("removedDescriptionText")

                check_event.save()

                tasks_to_remove = []
                if(check_event.taskDay):
                    tasks_to_remove.append(check_event.taskDay.id)
                if (check_event.taskWeek):
                    tasks_to_remove.append(check_event.taskWeek.id)
                if (check_event.taskMonth):
                    tasks_to_remove.append(check_event.taskMonth.id)
                Task.objects.filter(id__in = tasks_to_remove).delete()

                write_to_change_log(check_event.id, "remove", "Event", "company__name",
                                    check_event.company.name,
                                    get_current_site_user(request))

                return JsonResponse(create_response("success",
                                                    "Успешное удаление мероприятия"))
            elif (request.POST.get("transfer")):
                old_date_time = check_event.startTime

                check_event.startTime = request.POST.get("newDateTime")
                check_event.save()

                write_to_change_log(check_event.id, "transfer", "Event", "startTime",
                                    old_date_time,
                                    get_current_site_user(request))

                return JsonResponse(create_response("success",
                                                    "Успешный перенос мероприятия на новую дату"))
    if(request.POST.get("only_call_data")):
        day = bool_convertor_error_json_response(request.POST.get("day"))
        week = bool_convertor_error_json_response(request.POST.get("week"))
        month = bool_convertor_error_json_response(request.POST.get("month"))

        current_manager = get_current_manager(request)

        task_data = {}
        task_data["company"] = check_event.company
        task_data["artist"] = check_event.artist
        task_data["manager"] = current_manager
        task_data["comment"] = "Компания: " + check_event.company.name + ". Адрес: " + check_event.company.city.name + ", " + check_event.company.adress
        task_data["type"] = "event"
        task_data["assigned_by"] = get_site_user_from_manager(current_manager.id)
        task_data = convert_ids_to_objects("Task", task_data)

        if(day):
            if(check_event.callDay):
                pass
            elif(check_event.taskDay):
                pass
            else:
                task_data["datetime"] = check_event.startTime - timedelta(days=1)
                task_data["description"] = "Отзвон за день"
                day_task = Task(**task_data)
                day_task.save()
                check_event.taskDay = day_task
        else:
            if(check_event.taskDay):
                Task.objects.filter(id = check_event.taskDay.id).delete()
                check_event.taskDay = None
        check_event.save()

        if (week):
            if (check_event.callWeek):
                pass
            elif (check_event.taskWeek):
                pass
            else:
                task_data["datetime"] = check_event.startTime - timedelta(days=7)
                task_data["description"] = "Отзвон за неделю"
                week_task = Task(**task_data)
                week_task.save()
                check_event.taskWeek = week_task
        else:
            if (check_event.taskWeek):
                Task.objects.filter(id=check_event.taskWeek.id).delete()
                check_event.taskWeek = None
        check_event.save()

        if (month):
            if (check_event.callMonth):
                pass
            elif (check_event.taskMonth):
                pass
            else:
                task_data["datetime"] = check_event.startTime - timedelta(days=30)
                task_data["description"] = "Отзвон за месяц"
                month_task = Task(**task_data)
                month_task.save()
                check_event.taskMonth = month_task
        else:
            if (check_event.taskMonth):
                Task.objects.filter(id=check_event.taskMonth.id).delete()
                check_event.taskMonth = None
        check_event.save()


        try:
            return JsonResponse(create_response("success", "Успешное изменение настроек отзвонов"))
        except:
            return JsonResponse(create_response("error", "Ошибка при изменении настроек отзвонов"))
    if(request.POST.get("only_manager_percent")):
        current_user_options = get_user_options_from_db(request)
        try:
            if(current_user_type == "admin" or (current_user_type == "manager" and current_user_options["can_change_event_percents"])):
                check_event.sumPercent = int_convertor_error_false(request.POST.get("managerPercent"))
                check_event.save()
                return JsonResponse(create_response("success", "Процент менеджера успешно изменен"))
            else:
                return JsonResponse(create_response("info", "Невозможно изменить процент менеджера - нет доступа"))
        except:
            return JsonResponse(create_response("error", "Ошибка при изменении процента менеджера"))
    if(request.POST.get("artist_data")):
        if (current_user_type == "presentator"):
            return JsonResponse(create_response("error", "Артист не может изменить данные компании мероприятия"))
        else:
            try:
                artist_company_data = json.loads(request.POST.get("data"))

                old_values = Event.objects.filter(id = check_event.id).values("companyName", "companyAdress", "companyContacts")[0]

                for artist_data in artist_company_data:
                    check_change_value(old_values[artist_data], artist_company_data[artist_data], artist_data, "Event", check_event.id, request)
                Event.objects.filter(id=event_id).update(**artist_company_data)
                return JsonResponse(create_response("success", "Успешное изменение данных для артиста мероприятия"))
            except:
                return JsonResponse(create_response("error", "Ошибка при изменении данных для артиста мероприятия"))
    if(request.POST.get("new_show")):
        if(current_user_type == "presentator"):
            return JsonResponse(create_response("error", "Артист не может изменить шоу мероприятия"))
        else:
            try:
                new_artist_data = json.loads(request.POST.get("data"))
                #Запись в лог
                old_event_data = Event.objects.filter(id = check_event.id).values("artist__name")[0]

                Event.objects.filter(id=event_id).update(**new_artist_data)

                write_to_change_log(check_event.id, "edit", "Event", "artist__name",
                                    old_event_data["artist__name"],
                                    get_current_site_user(request))
                return JsonResponse(create_response("success", "Успешное изменение шоу мероприятия"))
            except:
                return JsonResponse(create_response("error", "Ошибка при изменении шоу мероприятия"))
    if (request.POST.get("new_manager")):
        if (current_user_type != "admin"):
            return JsonResponse(create_response("error", "Только администратор может изменить менеджера мероприятия"))
        else:
            try:
                # new_event_artist = Artist.objects.get(id=new_event_data["artist_id"])
                # new_event_company = Company.objects.get(id=new_event_data["company_id"])
                # cmsi = CMSILink.objects.filter(manager=new_event_data["manager_id"], show=new_event_artist)
                # if (cmsi.count() == 0):
                #     new_link = CMSILink(
                #         manager=Manager.objects.get(id=new_event_data["manager_id"]),
                #         show=Artist.objects.get(id=new_event_data["artist_id"])
                #     )
                #     new_link.save()
                #     new_link.company.add(new_event_company)
                # else:
                #     cmsi[0].company.add(new_event_company.id)
                # response_text = "Успешное изменение менеджера мероприятия. Компания добавлена менеджеру в список клиентов"
                # write_to_change_log(old_event_data["id"], "edit", "Event", "manager__alias", Event.objects.get(id = old_event_data["id"]).manager.alias, get_current_site_user(request))


                new_data = json.loads(request.POST.get("data"))

                write_to_change_log(check_event.id, "transference", "Event", "manager__siteuser__name",
                                    check_event.manager.siteuser.name,
                                    get_current_site_user(request))

                Event.objects.filter(id=event_id).update(**new_data)

                return JsonResponse(create_response("success", "Успешное изменение менеджера мероприятия"))
            except:
                return JsonResponse(create_response("error", "Ошибка при изменении менеджера мероприятия"))

    response_text = "Успешное редактирование мероприятия"
    event = Event.objects.filter(id = request.POST.get("id"))
    data = json.loads(request.POST.get("data"))
    if("percent" in data):
        data["percent"] = data["percent"].replace("%", "")
        if(data["percent"] == ""):
            data["percent"] = None
    if("price" in data):
        if(data["price"] == ""):
            data["price"] = None
    if("childCount" in data):
        if(data["childCount"] == ""):
            data["childCount"] = None
    if("resultSum" in data):
        if(data["resultSum"] == ""):
            data["resultSum"] = None
    if("sumTransfered" in data):
        if(data["sumTransfered"] == ""):
            data["sumTransfered"] = None
    #Возможные поля для редактирования в этой вьюшке
    allowed_fields = []

    try:
        manager = get_current_manager(request)
    except:
        manager = 0

    if (current_user_type == "admin"):
        allowed_fields = ["price", "childCount", "resultSum", "percent", "sumTransfered", "artist", "manager",
                          "removed", "removedDescription", "removedDate", "note", "artistNote"]
    elif(current_user_type == "presentator"):
        presentator = get_current_presentator(request)
        allowed_try = PCSLink.objects.filter(presentator__id=presentator.id,
                                             city__in=City.objects.filter(id=event[0].company.city.id),
                                             shows__in=Artist.objects.filter(id=event[0].artist.id))
        if (allowed_try.count() == 0):
            result = create_response("error", "Невозможно редактировать чужое мероприятие")
            JsonResponse(result)
        else:
            allowed_fields = ["price", "childCount", "resultSum", "sumTransfered", "percent", "artistNote"]
    elif(current_user_type == "manager"):
        allowed_fields = ["price", "childCount", "sumTransfered", "resultSum", "note", "percent"]



    #Удаление полей, которые нельзя редактировать
    changed_data = data.copy()
    for field in changed_data:
        if(not field in allowed_fields):
            data.pop(field)
    #Запись старых значений для лога
    old_event_data = list(event.values())[0]
    #Запись новых данных
    try:
        event.update(**data)
    except:
        return  create_response("error", "Ошибка при редактировании мероприятия, данные не сохранены")

    #Отслеживаемые поля для логов
    followed_fields = ["price", "percent", "childCount", "resultSum", "sumTransfered", "note", "percent"]

    #Запись измененных данных в лог
    new_event_data = event.values()[0]
    for key in data:
        if key in followed_fields:
            if(new_event_data[key] != old_event_data[key]):
                if(old_event_data[key]) == None:
                    old_event_data[key] = ""
                write_to_change_log(old_event_data["id"], "edit", "Event", key, old_event_data[key], get_current_site_user(request))

    result = create_response("success", response_text)
    return JsonResponse(result)
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def events_add_event(request):
    company_id = request.POST.get("company")
    artist = request.POST.get("artist")
    user_options = get_user_options_from_db(request)
    site_user = get_current_site_user(request)

    current_manager = get_current_manager(request)
    try:
        company = Company.objects.get(id = company_id)
    except:
        return JsonResponse(create_response("error", "Ошибка при добавлении мероприятия: не выбрано учреждение"))
    new_event_data = {}
    new_event_data["company"] = company_id
    new_event_data["companyName"] = request.POST.get("companyName")
    new_event_data["companyAdress"] = request.POST.get("companyAdress")
    new_event_data["companyContacts"] = request.POST.get("companyContacts")
    new_event_data["note"] = request.POST.get("note")
    new_event_data["artist"] = artist
    if(request.POST.get("price") != ""):
        new_event_data["price"] = request.POST.get("price")
    if(request.POST.get("childCount") != ""):
        new_event_data["childCount"] = request.POST.get("childCount")
    if(request.POST.get("percent") != ""):
        new_event_data["percent"] = request.POST.get("percent")
    new_event_data["startTime"] = request.POST.get("startTime")
    new_event_data["manager"] = current_manager
    new_event_data["sumPercent"] = current_manager.eventPercent
    try:
        new_event_data = convert_ids_to_objects("Event", new_event_data)
        new_event = Event(**new_event_data)
        new_event.save()
        write_to_change_log(new_event.id, "add", "Event", "company__name", new_event.company.name, get_current_site_user(request),
                            new_event.company.id, "Company")
    except:
        return JsonResponse(create_response("error", "Ошибка при сохранении мероприятия. Перезагрузите страницу и попробуйте снова, при повторной ошибке обратитесь к администратору сервера"))

    if(request.POST.get("attentionCallMonthBool") == "true" or request.POST.get("attentionCallWeekBool") == "true" or request.POST.get("attentionCallDayBool") == "true"):
        task_data = {}
        task_data["company"] = company_id
        task_data["artist"] = artist
        task_data["manager"] = current_manager
        task_data["comment"] = "Компания: " + company.name + ". Адрес: " + company.city.name + ", " + company.adress
        task_data["type"] = "event"
        task_data["assigned_by"] = get_site_user_from_manager(current_manager.id)
        task_data = convert_ids_to_objects("Task", task_data)
        if (request.POST.get("attentionCallMonthBool") == "true"):
            task_data["description"] = "Отзвон за месяц"
            task_data["datetime"] = datetime.strptime(new_event.startTime, "%Y-%m-%d %H:%M") - timedelta(days=30)

            month_task = Task(**task_data)
            month_task.save()
            new_event.taskMonth = month_task
        if (request.POST.get("attentionCallWeekBool") == "true"):
            task_data["description"] = "Отзвон за неделю"
            task_data["datetime"] = datetime.strptime(new_event.startTime, "%Y-%m-%d %H:%M") - timedelta(days=7)
            week_task = Task(**task_data)
            week_task.save()
            new_event.taskWeek = week_task
        if (request.POST.get("attentionCallDayBool") == "true"):
            task_data["description"] = "Отзвон за день"
            task_data["datetime"] = datetime.strptime(new_event.startTime, "%Y-%m-%d %H:%M") - timedelta(days=1)
            day_task = Task(**task_data)
            day_task.save()
            new_event.taskDay = day_task
        new_event.save()
    return JsonResponse(create_response("success", "Успешное добавление мероприятия"))

@user_is_auth_aj_checker
@user_is_presentator_aj_checker
def events_add_presentator_event(request):
    presentator = get_current_presentator(request)
    presentator_event_data = {}
    presentator_event_data["date"] = convert_date_from_select2(request.POST.get("date"))
    if(PresentatorEvent.objects.filter(presentator__id = presentator.id, date = convert_date_from_select2(request.POST.get("date"))).count() != 0):
        return JsonResponse(create_response("info", "На эту дату уже назначено событие, для начала удалите его"))

    full_day = bool_convertor_error_json_response(request.POST.get("full_day"))
    presentator_event_data["presentator"] = presentator.id
    presentator_event_data["fullday"] = full_day

    if(not full_day):
        startTime = convert_time_from_select2(request.POST.get("startTime"))
        endTime = convert_time_from_select2(request.POST.get("endTime"))
        if((endTime-startTime).total_seconds() < 0):
            return JsonResponse(create_response("error", "Диапазон времени не может быть отрицательным"))
        presentator_event_data["startTime"] = startTime.time()
        presentator_event_data["endTime"] = endTime.time()

    presentator_event_data["description"] = request.POST.get("description")
    presentator_event_data["comment"] = request.POST.get("comment")
    presentator_event_data = convert_ids_to_objects("PresentatorEvent", presentator_event_data)
    try:
        presentator_event = PresentatorEvent(**presentator_event_data)
        presentator_event.save()
        return JsonResponse(create_response("success", "Успешная отметка занятого времени"))
    except:
        return JsonResponse(create_response("error", "Ошибка при отметке занятого времени"))
@user_is_auth_aj_checker
def events_remove_presentator_event(request):
    current_user_type = get_current_user_type(request)
    if(current_user_type == "manager"):
        return JsonResponse(create_response("info", "Менеджеры не могут удалять мероприятия артистов"))
    try:
        PresentatorEvent.objects.filter(id = request.POST.get("id")).delete()
        return JsonResponse(create_response("success", "Успешное удаление отметки занятого времени"))
    except:
        return JsonResponse(create_response("error", "Ошибка при удалении отметки занятого времени"))
@user_is_auth_aj_checker
def events_get_presentator_event_data(request):
    try:
        presentator_event = list(PresentatorEvent.objects.filter(id = request.POST.get("id")).values("presentator__siteuser__alias", "startTime", "endTime", "date", "fullday", "description", "comment", "statsdt"))[0]
        return JsonResponse(create_response("data", "", presentator_event, request))
    except:
        return JsonResponse(create_response("error", "Ошибка при получении данных занятого времени"))
@user_is_auth_aj_checker
def events_get_in_daterange(request):
    #Вовращаемое значение, в формате JSON. Контейнер
    site_user = get_current_site_user(request, "object")
    current_user_type = get_current_user_type(request)
    result = {}
    filter_values = {}

    count_on_page = 20

    current_manager_id = None
    only_this_company = bool_convertor_error_json_response(request.POST.get("only_this_company"))

    city_id = int_convertor_error_0(request.POST.get("city"))
    show_id = int_convertor_error_0(request.POST.get("show"))

    current_presentator_id = int_convertor_error_0(request.POST.get("presentator"))

    only_own = bool_convertor_error_json_response(request.POST.get("only_own"))

    cal_type = request.POST.get("cal_type")
    cal_option = request.POST.get("cal_option")

    paginator = request.POST.get('paginator')
    search_string = ""
    if(request.POST.get("search")):
        search_string = request.POST.get("search")
    if(only_this_company):
        filter_values["company__id"] = request.POST.get("company_id")
    #Если есть диапазоны, они указываются, в противном случае - указывается менеджер, для полного списка мероприятий
    if(request.POST.get("to") != "false"):
        filter_values["startTime__lte"] = request.POST.get("to")
    if(request.POST.get("from") != "false"):
        filter_values["startTime__gte"] = request.POST.get("from")
    else:
        if (cal_type == "fulllist"):
            month_count = site_user.options.full_list_events_month_count
            filter_values["startTime__gte"] = datetime.now() - monthdelta(month_count)
    if(request.POST.get("page")):
        shift = (int(request.POST.get("page")) - 1) * count_on_page
        events_from = shift
        events_to = shift + count_on_page
    else:
        events_from = 0
        events_to = 900


    show_city_Qfilters_or_list = None

    #Параметры по ролям

    # Параметры фильтрации по ролям
    pairs = None
    exlude_filter_values = {}
    if (current_user_type == "admin"):
        if(current_presentator_id and current_presentator_id != 0):
            current_manager_id = 0
            pairs = get_allowed_show_city_pairs(request, False, current_presentator_id)
        else:
            current_manager_id = 0
            if (city_id != 0):
                filter_values["company__city__id"] = city_id
                if (show_id == 0):
                    filter_values["artist__id__in"] = get_worked_shows(city_id, _list=True)
                else:
                    filter_values["artist__id"] = show_id
            else:
                if (show_id != 0):
                    pairs = get_allowed_show_city_pairs(request, show_id)
                else:
                    pairs = get_allowed_show_city_pairs(request)
    elif (current_user_type == "manager"):
        current_manager_id = get_current_manager(request).id
        # Проверка на вхождение в доступные города
        if (city_id != 0):
            if (cal_type != "fulllist"):
                if (not (city_id in get_manager_allowed_cities(current_manager_id))):
                    return JsonResponse(create_response("info", "Нет доступа к мероприятиям этого города"))
        if (current_presentator_id and current_presentator_id != 0 and cal_type != "fulllist" and not only_own):
            pairs = get_allowed_show_city_pairs(request, False, current_presentator_id)
        else:
            if (city_id != 0):
                filter_values["company__city__id"] = city_id
                if (show_id == 0):
                    filter_values["artist__id__in"] = get_manager_allowed_shows_for_city(request, city_id)
                else:
                    filter_values["artist__id"] = show_id
            else:
                if (show_id != 0):
                    pairs = get_allowed_show_city_pairs(request, show_id)
                else:
                    pairs = get_allowed_show_city_pairs(request)
            if (only_own or cal_type == "fulllist"):
                filter_values["manager__id"] = current_manager_id
        exlude_filter_values["removed"] = True
    elif (current_user_type == "presentator"):
        current_manager_id = 0
        presentator = get_current_presentator(request)
        if (city_id != 0):
            filter_values["company__city__id"] = city_id
            if (show_id == 0):
                filter_values["artist__id__in"] = get_presentator_allowed_shows_for_city(presentator.id, city_id)
            else:
                filter_values["artist__id"] = show_id
        else:
            if (show_id != 0):
                pairs = get_allowed_show_city_pairs(request, show_id)
            else:
                pairs = get_allowed_show_city_pairs(request)
        exlude_filter_values["removed"] = True
    else:
        return JsonResponse(create_response("error", "Не определен текущий тип пользователя"))

    try:
        events = Event.objects.filter(**filter_values).exclude(**exlude_filter_values)
        if(search_string != ""):
            events = events.filter((Q(company__name__icontains = search_string) | Q(company__adress__icontains = search_string)))
        if(pairs):
            show_city_Qfilters_and_list = []
            for pair in pairs:
                Qfiltred = [Q(x) for x in pair]
                show_city_Qfilters_and_list.append(functools.reduce(operator.and_, Qfiltred))
            Qfiltred = [Q(x) for x in show_city_Qfilters_and_list]
            show_city_Qfilters_or_list = functools.reduce(operator.or_, Qfiltred)
            events = events.filter(show_city_Qfilters_or_list)
        if(paginator):
            data = {}
            data["page_count"] = calc_pages_count(count_on_page, events.count())
            return JsonResponse(create_response("data", "", data))
    except:
        return JsonResponse(create_response("error", "Ошибка при получении списка мероприятий"))
    try:
        events = replace_datetime_to_string(list(events.order_by("startTime").values("id", "company__id", "companyName", "companyAdress", "company__city__name", "artist__id", "artist__name", "artist__color", "manager__id", "manager__siteuser__alias", "manager__siteuser__id", "startTime", "crashBool", "sumTransfered", "resultSum", "attentionCallWeekBool", "attentionCallWeekDone", "attentionCallDayBool", "attentionCallDayDone", "removed")[events_from:events_to]))
    except:
        return JsonResponse(create_response("info", "Слишком много мероприятий для запроса, уточните параметры: город, шоу, артист"))

    events_ids = []
    for event in events:
        events_ids.append(event["id"])
        event["log_count"] = 0
        if(city_id == 0):
            event["companyAdress"] = event["company__city__name"]+ ", " + event["companyAdress"]
            event["companyName"] = event["company__city__name"]+ ", " + event["companyName"]

        if(event["sumTransfered"] != None and event["sumTransfered"] != 0):
            event["sumTransferedFilled"] = True
        else:
            event["sumTransferedFilled"] = False
        event.pop("sumTransfered")

        if (event["resultSum"] != None and event["resultSum"] != 0):
            event["resultSumFilled"] = True
        else:
            event["resultSumFilled"] = False
        event.pop("resultSum")

        if (current_user_type == "presentator"):
            events_limit = get_user_options_from_db(request)["presentator_show_events_limit"]
            converted_event_date = datetime.strptime(event["startTime"], "%Y.%m.%d %H:%M")
            if (converted_event_date > datetime.now() and ((converted_event_date - datetime.now()).total_seconds() / 60 / 60 / 24 / 7 > events_limit)):
                event["companyName"] = event["company__city__name"]
                event["companyAdress"] = event["company__city__name"]
                event["own"] = False
                event["allowed"] = False
            else:
                event["own"] = True
                event["allowed"] = True
        elif(current_manager_id == event["manager__id"] or current_manager_id == 0):
            event["own"] = True
            event["allowed"] = True
        else:
            event["own"] = False
            if(site_user.options.can_show_allowed_cities_events):
                event["allowed"] = True
            else:
                event["companyName"] = event["company__city__name"]
                event["companyAdress"] = event["company__city__name"]
                event["manager__siteuser__alias"] = "Менеджер"
                event["allowed"] = False
                event.pop("company__id")
        event.pop("manager__id")
        event.pop("manager__siteuser__id")

        event["status"] = events_fill_event_status(event, datetime.now(), current_user_type)
    if(cal_type == "mycal" or cal_type == "datelist" or cal_type == "fulllist"):
        events_calls_data = Event.objects.filter(id__in = events_ids).values("id", "callDay__id", "callWeek__id", "callMonth__id", "taskDay__id", "taskWeek__id", "taskMonth__id")


        for calls_data in events_calls_data:
            for event in events:
                if (not event["own"]):
                    continue
                if(calls_data["id"] == event["id"]):
                    calls_data_dict = {}
                    event["calls_data"] = calls_data_dict
                    day_call_data = {}
                    week_call_data = {}
                    month_call_data = {}

                    calls_data_dict["day"] = day_call_data
                    calls_data_dict["week"] = week_call_data
                    calls_data_dict["month"] = month_call_data

                    if(calls_data["callDay__id"]):
                        day_call_data["status"] = "done"
                        day_call_data["id"] = calls_data["callDay__id"]
                    elif(calls_data["taskDay__id"]):
                        if (current_user_type == "presentator"):
                            day_call_data["status"] = "presentatorwait"
                        else:
                            day_call_data["status"] = "need"
                            day_call_data["id"] = calls_data["taskDay__id"]
                    else:
                        day_call_data["status"] = "notneed"

                    if (calls_data["callWeek__id"]):
                        week_call_data["status"] = "done"
                        week_call_data["id"] = calls_data["callWeek__id"]
                    elif (calls_data["taskWeek__id"]):
                        if (current_user_type == "presentator"):
                            week_call_data["status"] = "presentatorwait"
                        else:
                            week_call_data["status"] = "need"
                            week_call_data["id"] = calls_data["taskWeek__id"]
                    else:
                        week_call_data["status"] = "notneed"

                    if (calls_data["callMonth__id"]):
                        month_call_data["status"] = "done"
                        month_call_data["id"] = calls_data["callMonth__id"]
                    elif (calls_data["taskMonth__id"]):
                        if(current_user_type == "presentator"):
                            month_call_data["status"] = "presentatorwait"
                        else:
                            month_call_data["status"] = "need"
                            month_call_data["id"] = calls_data["taskMonth__id"]
                    else:
                        month_call_data["status"] = "notneed"
    chat_message_viewed= ChatMessage.objects.filter(table_link_id__in = events_ids, table = get_table_object_from_name("Event")).exclude(viewers__id = site_user.id).values("id", "table_link_id").distinct()
    for event in events:
        if(event["own"]):
            for chat_message in chat_message_viewed:
                if(chat_message["table_link_id"] == event["id"]):
                    if("chat_messages_count" in event):
                        event["chat_messages_count"] = event["chat_messages_count"] + 1
                    else:
                        event["chat_messages_count"] = 1
    allowed_tables = ["Call", "Event"]
    allowed_linked_tables = ["Event"]
    allowed_actions = []
    if (current_user_type == "admin"):
        allowed_actions = ["add", "edit", "remove", "view", "transfer", "transference", "permission"]
    else:
        allowed_actions = ["add", "edit", "remove", "view", "transfer", "transference"]

    logs = ChangeFieldLog.objects.filter(Q(table_link_id__in = events_ids) | Q(link_to_object_id__in = events_ids), Q(
        link_to_object_table__name__in=allowed_linked_tables) | Q(table__name__in=allowed_tables),
        changeType__name__in=allowed_actions).exclude(whoViewChange__id=site_user.id).values("table_link_id", "link_to_object_id")

    for log in logs:
        for event in events:
            if(event["own"]):
                if(log["table_link_id"] == event["id"] or log["link_to_object_id"] == event["id"]):
                    event["log_count"] = event["log_count"] + 1

    result = {}
    result["events"] = events

    need_presentator_events = False
    presentator_events_dt_filters = {}
    presentator_events_values = ["presentator__siteuser__alias", "startTime", "endTime", "date", "fullday", "description", "comment"]
    if(cal_type == "oldcal"):
        need_presentator_events = True
        presentator_events_dt_filters["date__gte"] = convert_date_from_string(request.POST.get("from"))
        presentator_events_dt_filters["date__lte"] = convert_date_from_string(request.POST.get("to"))
    elif(cal_type == "mycal"):
        if(cal_option == "week"):
            presentator_events_dt_filters["date__gte"] = convert_date_from_string(request.POST.get("from"))
            presentator_events_dt_filters["date__lte"] = convert_date_from_string(request.POST.get("to"))
        else:
            presentator_events_dt_filters["date"] = convert_date_from_string(request.POST.get("from"))
        need_presentator_events = True

    elif(cal_type == "datelist" and cal_option == "date"):
        need_presentator_events = True
        presentator_events_dt_filters["date"] = convert_date_from_string(request.POST.get("from"))
    if(need_presentator_events):
        if(current_presentator_id != 0):
            presentator_events = list(PresentatorEvent.objects.filter(presentator__id = current_presentator_id, **presentator_events_dt_filters).order_by("date").values("id", "presentator__siteuser__alias", "startTime", "endTime", "date", "fullday", "description"))
        else:
            presentator_events = []
        # elif(city_id != 0):
        #     presentator_events = list(PresentatorEvent.objects.filter(
        #         presentator__id__in=PCSLink.objects.filter(city__id=city_id).values_list("presentator__id",
        #                                                                                       flat=True), **presentator_events_dt_filters).order_by("date").values("id", "presentator__siteuser__alias", "startTime", "endTime", "date", "fullday", "description"))
        # else:
        #     pairs = get_allowed_show_city_pairs(request)
        #     if(len(pairs) != 0):
        #         pair_city_id = pairs[0][1][1]
        #         for pair in pairs:
        #             if(pair_city_id != pair[1][1]):
        #                 pair_city_id = 0
        #                 break
        #         if(pair_city_id != 0):
        #             presentator_events = list(PresentatorEvent.objects.filter(presentator__id__in = PCSLink.objects.filter(city__id = pair_city_id).values_list("presentator__id", flat=True), **presentator_events_dt_filters).order_by("date").values("id", "presentator__siteuser__alias", "startTime", "endTime", "date", "fullday", "description"))
        #         else:
        #             presentator_events = []

        result["presentator_events"] = presentator_events
    return JsonResponse(create_response("data", "", result, request))
@user_is_auth_aj_checker
def events_get_frame_counters(request):
    current_user_type = get_current_user_type(request)
    result = {}
    filter_values = {}

    city_id = int_convertor_error_0(request.POST.get("city"))
    show_id = int_convertor_error_0(request.POST.get("show"))

    current_presentator_id = int_convertor_error_0(request.POST.get("presentator"))

    only_own = bool_convertor_error_json_response(request.POST.get("only_own"))

    only_this_company = bool_convertor_error_json_response(request.POST.get("only_this_company"))
    if(only_this_company):
        filter_values["company__id"] = request.POST.get("company_id")
    # Параметры фильтрации по ролям
    pairs = None
    if (current_user_type == "admin"):
        if (current_presentator_id and current_presentator_id != 0):
            current_manager_id = 0
            pairs = get_allowed_show_city_pairs(request, False, current_presentator_id)
        else:
            current_manager_id = 0
            if (city_id != 0):
                filter_values["company__city__id"] = city_id
                if (show_id == 0):
                    filter_values["artist__id__in"] = get_worked_shows(city_id, _list=True)
                else:
                    filter_values["artist__id"] = show_id
            else:
                if (show_id != 0):
                    pairs = get_allowed_show_city_pairs(request, show_id)
                else:
                    pairs = get_allowed_show_city_pairs(request)
    elif (current_user_type == "manager"):
        current_manager_id = get_current_manager(request).id
        if (current_presentator_id and current_presentator_id != 0 and not only_own):
            pairs = get_allowed_show_city_pairs(request, False, current_presentator_id)
        else:
            if (city_id != 0):
                filter_values["company__city__id"] = city_id
                if (show_id == 0):
                    filter_values["artist__id__in"] = get_manager_allowed_shows_for_city(request, city_id)
                else:
                    filter_values["artist__id"] = show_id
            else:
                if (show_id != 0):
                    pairs = get_allowed_show_city_pairs(request, show_id)
                else:
                    pairs = get_allowed_show_city_pairs(request)
        if (only_own):
            filter_values["manager__id"] = current_manager_id
    elif(current_user_type == "presentator"):
        presentator = get_current_presentator(request)
        if (city_id != 0):
            filter_values["company__city__id"] = city_id
            if (show_id == 0):
                filter_values["artist__id__in"] = get_presentator_allowed_shows_for_city(presentator.id, city_id)
            else:
                filter_values["artist__id"] = show_id
        else:
            if (show_id != 0):
                pairs = get_allowed_show_city_pairs(request, show_id)
            else:
                pairs = get_allowed_show_city_pairs(request)
    else:
        return JsonResponse(create_response("error", "Не определен текущий тип пользователя"))

    non_qfiltred_events_for_counters = Event.objects.filter(**filter_values)
    if (pairs):
        show_city_Qfilters_and_list = []
        for pair in pairs:
            Qfiltred = [Q(x) for x in pair]
            show_city_Qfilters_and_list.append(functools.reduce(operator.and_, Qfiltred))
        Qfiltred = [Q(x) for x in show_city_Qfilters_and_list]
        show_city_Qfilters_or_list = functools.reduce(operator.or_, Qfiltred)
        events_for_counters = non_qfiltred_events_for_counters.filter(show_city_Qfilters_or_list)
    else:
        events_for_counters = non_qfiltred_events_for_counters

    if (current_user_type == "manager"):
        current_manager_id = get_current_manager(request).id
        events_for_counters = events_for_counters.filter(Q(manager__id=current_manager_id, removed=True) | Q(removed=False))
    elif(current_user_type == "presentator"):
        events_for_counters = events_for_counters.filter(removed = False)
    truncate_date = connection.ops.date_trunc_sql('day', 'startTime')
    extra_tasks = events_for_counters.extra({'day': truncate_date})
    result["counters"] = list(extra_tasks.values('day').annotate(Count('pk')).order_by('day'))


    if (current_presentator_id != 0):
        presentator_events = list(
            PresentatorEvent.objects.filter(presentator__id=current_presentator_id).values("date", "fullday").order_by("date"))
    else:
        presentator_events = []
    # elif (city_id != 0):
    #     presentator_events = list(PresentatorEvent.objects.filter(
    #         presentator__id__in=PCSLink.objects.filter(city__id=city_id).values_list("presentator__id",
    #                                                                                  flat=True)).values("date", "fullday").order_by("date"))
    #
    # else:
    #     pairs = get_allowed_show_city_pairs(request)
    #     if (len(pairs) != 0):
    #         pair_city_id = pairs[0][1][1]
    #         for pair in pairs:
    #             if (pair_city_id != pair[1][1]):
    #                 pair_city_id = 0
    #                 break
    #         if (pair_city_id != 0):
    #             presentator_events = list(PresentatorEvent.objects.filter(
    #                 presentator__id__in=PCSLink.objects.filter(city__id=pair_city_id).values_list("presentator__id",
    #                                                                                               flat=True)).values("date", "fullday").order_by("date"))
    #         else:
    #             presentator_events = []

    result["busy_marks"] = presentator_events
    return JsonResponse(create_response("data", "", result, request))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def events_get_allowed_managers(request):
    event = Event.objects.get(id = request.POST.get("id"))
    allowed_event_managers = list(Manager.objects.filter(id__in = CMSILink.objects.filter(company__id__in = [event.company.id,], show__id = event.artist.id).values_list("manager__id", flat=True)).values("id", "siteuser__name"))
    return JsonResponse(create_response("data", "", allowed_event_managers ))


def events_fill_events_statuses(_events, _request = False, _done_count = False, _siteuser = False):
    """Получение статусов мероприятий исходя из типа пользователя"""
    current_date_time = datetime.now().replace(hour = 0, minute = 0, second = 0)
    if(_request):
        user_type = get_current_user_type(_request)
        siteuser = get_current_site_user(_request, _type="object")
    else:
        user_type = get_current_user_type(_siteuser=_siteuser)
        siteuser = SiteUser.objects.get(id = _siteuser)
    if(user_type == "manager"):
        if(_request):
            manager_id = get_current_manager(_request).id
        else:
            manager_id = get_manager_from_site_user(_siteuser)
    else:
        manager_id = 0
    processed_events = []

    for event in _events:
        event_status = "error"
        processed_event = {}
        processed_event["id"] = event["id"]
        if(user_type == "manager" and event["manager__id"] != manager_id):
            if(siteuser.options.can_show_allowed_cities_events):
                event_status = "allowed"
            else:
                event_status = "alien"
        elif(event["removed"]):
            event_status = "removed"
        elif(event["crashBool"]):
            event_status = "crash"
        else:
            compare_date_time = event["startTime"].replace(hour=0, minute=0, second=0)
            datetime_dif = current_date_time - compare_date_time
            datetime_dif_days = datetime_dif.days
            if (event["sumTransfered"] != None and event["sumTransfered"] != 0):
                event_status = "success"
            elif(event["resultSum"] != None and event["resultSum"] != 0):
                if (datetime_dif_days > 3):
                    event_status = "late_money"
                else:
                    event_status = "wait_money"
            elif(datetime_dif_days < 0):
                event_status = "wait"
            else:
                event_status = "wait_no_result"

        processed_event["status"] = event_status
        processed_event["artist__name"] = event["artist__name"]
        processed_event["artist__color"] = event["artist__color"]
        processed_event["manager__id"] = event["manager__id"]
        processed_event["manager__siteuser__alias"] = event["manager__siteuser__alias"]
        processed_event["manager__siteuser__id"] = event["manager__siteuser__id"]
        processed_events.append(processed_event)

    counter = 0
    if(_done_count):
        for event in processed_events:
            if(event["status"] == "success" or event["status"] == "late_money" or event["status"] == "wait_money"):
                counter = counter + 1
        return counter
    else:
        return processed_events
def events_get_values_list_for_calc_statuses():
    values = ["id", "manager__id", "manager__siteuser__id", "manager__siteuser__alias", "startTime", "removed", "crashBool", "sumTransfered", "resultSum", "artist__color", "artist__name"]
    return values

def events_fill_event_status(_event, _current_date_time, _user_type):
    _current_date_time = _current_date_time.replace(hour = 0, minute = 0, second = 0)
    _compare_date_time = datetime.strptime(_event["startTime"], "%Y.%m.%d %H:%M").replace(hour = 0, minute = 0, second = 0)
    datetime_dif = _current_date_time - _compare_date_time
    datetime_dif_days = datetime_dif.days
    if(not _event["allowed"]):
        return "alien"
    elif (_event["removed"]):
        return "removed"
    elif (_event["crashBool"]):
        return "crash"
    elif(_event["sumTransferedFilled"]):
        return  "success"
    elif(_event["resultSumFilled"]):
        if(datetime_dif_days  > 3):
            return "late_money"
        else:
            return "wait_money"
    elif(_event["resultSumFilled"]):
        return "success"
    elif(datetime_dif_days < 0):
        return "wait"
    else:
        return "wait_no_result"
    return "unknown_status"

@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def manager_get_manager_companies_search(request):
    user_type = get_current_user_type(request)
    if(user_type == "admin"):
        allowed_companies = Company.objects.filter(city__enabled = True).values_list("id", flat=True)
    else:
        allowed_companies = get_manager_allowed_companies(request)
    if(request.GET.get("id") != "" and request.GET.get("id") != "False" and request.GET.get("id") != None):
        user_type = get_current_user_type(request)
        if(user_type == "admin" or (int(request.GET.get("id")) in get_manager_allowed_companies((request)))):
            finded_companies = Company.objects.filter(id = request.GET.get("id")).values("id", "name", "city__name", "adress", "telephone")
        else:
            finded_companies = []
    else:
        search_string = request.GET.get("search_string")
        filter_values = {}
        if(request.GET.get("city") != "0"):
            filter_values["city__id"] = request.GET.get("city")
        filter_values["id__in"] = allowed_companies
        finded_companies = Company.objects.filter((Q(name__icontains = search_string) | Q(adress__icontains =search_string) | Q(city__name__icontains = search_string)), **filter_values).values("id", "name", "city__name", "adress", "telephone")


    if(len(finded_companies) == 0):
        result = create_response("error", "Ничего не найдено")
    else:
        data = list(finded_companies)
        result = create_response("data", "Успешная загрузка данных", data)
    return JsonResponse(result)

#Control===============================================================================================================================
#Control_users


@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_users_list(request):
    #тип пользователей
    type = ""
    status = ""
    search = ""
    paginator = request.POST.get("paginator")
    if(request.POST.get("page")):
        page = int(request.POST.get("page"))
    filter_params = {}
    count_on_page = 20

    if(request.POST.get("type")):
        type = request.POST.get("type")
    else:
        return JsonResponse(create_response("error", "Ошибка при отображении списка пользователей. Не удалось определить выбранный тип пользователей."))
    if(request.POST.get("status")):
        if(request.POST.get("status") == "Активныеизаблокированные"):
            filter_params["deleted"] = False
        elif(request.POST.get("status") == "Активные"):
            filter_params["deleted"] = False
            filter_params["active"] = True
        elif(request.POST.get("status") == "Заблокированные"):
            filter_params["deleted"] = False
            filter_params["active"] = False
        elif(request.POST.get("status") == "Удаленные"):
            filter_params["deleted"] = True
    else:
        return JsonResponse(create_response("error", "Ошибка при отображении списка пользователей. Не удалось определить выбранный статус пользователей."))
    if(request.POST.get("search") and request.POST.get("search") != ""):
        search = request.POST.get("search")
    else:
        search = ""

    site_users = SiteUser.objects.filter((Q(alias__icontains = search) | Q(name__icontains = search) | Q(user__username__icontains = search)), user__in = User.objects.filter(groups__name = type), **filter_params)

    data = {}
    if(paginator):
        if(site_users.count() != 0):
            data["page_count"] = calc_pages_count(count_on_page, site_users.count())
        else:
            data["page_count"] = 0
        return JsonResponse(create_response("data", "Успешная загрузка данных", data))
    low_range = get_low_range_value(page, count_on_page)
    site_users = SiteUser.objects.filter(
        (Q(alias__icontains=search) | Q(name__icontains=search) | Q(user__username__icontains=search)),
        user__in=User.objects.filter(groups__name=type), **filter_params).exclude(user__username="archy").order_by(
        "name")[low_range:low_range + count_on_page]
    data = list(site_users.values("id", "name", "alias", "user__username", "password", "email", "active", "deleted"))

    return JsonResponse(create_response("data", "", data))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_users_add(request):
    type = request.POST.get("type")
    mode = request.POST.get("mode")
    id = request.POST.get("id")
    choosen_site_user = SiteUser.objects.filter(id = id)

    if(request.POST.get("firstName")):
        first_name = request.POST.get("firstName")
        last_name = request.POST.get("lastName")

        full_name = (first_name.strip() + " " + last_name.strip())
    else:
        name = request.POST.get("name")
        full_name = request.POST.get("name")
    alias = request.POST.get("alias")
    email = request.POST.get("email")
    adress = request.POST.get("adress")
    contacts = request.POST.get("contacts")
    login = request.POST.get("login")
    password = request.POST.get("password")
    percent = int_convertor_error_0(request.POST.get("percent"))

    check_user = User.objects.filter(username = login)
    if(mode == "add"):
        if(check_user.count() != 0 and choosen_site_user.count() == 0):
            return JsonResponse(create_response("error", "Пользователь с таким логином уже присутствует в системе. Используйте другой логин"))

        check_site_user = SiteUser.objects.filter(name = full_name)
        if(check_site_user.count() != 0):
            if(choosen_site_user.count() > 0):
                if(not (choosen_site_user[0].name == check_site_user[0].name)):
                    return JsonResponse(create_response("error", "Пользователь с таким именем и фамилией уже присутствует в системе. Перепроверьте данные"))
            else:
                return JsonResponse(create_response("error", "Пользователь с таким именем и фамилией уже присутствует в системе. Перепроверьте данные"))

        check_site_user = SiteUser.objects.filter(alias = alias)
        if(check_site_user.count() != 0):
            if(choosen_site_user.count() > 0):
                if(not (choosen_site_user[0].alias == check_site_user[0].alias)):
                    return JsonResponse(create_response("error", "Пользователь с такими псевдонимом уже присутствует в системе. Перепроверьте данные"))
            else:
                return JsonResponse(create_response("error", "Пользователь с такими псевдонимом уже присутствует в системе. Перепроверьте данные"))


    #Менеджеры
    if(type == "Менеджеры"):

        #checkers
        if(percent < 1 or percent > 99):
            return JsonResponse(create_response("error", "Значение процента менеджера должно лежать в диапазоне от 1 до 99"))


        #Редактирование
        if(choosen_site_user.count() > 0):
            site_user_model = SiteUser.objects.get(id = id)
            manager = Manager.objects.get(siteuser=site_user_model)
            user = User.objects.get(id = site_user_model.user.id)
            if(site_user_model.password != password):
                site_user_model.password = check_change_value(site_user_model.password, password, "password", "SiteUser", site_user_model.id, request)
                user.set_password(password)
                user.save()

            site_user_model.name = check_change_value(site_user_model.name, full_name, "name", "SiteUser", site_user_model.id, request)
            site_user_model.alias = check_change_value(site_user_model.alias, alias, "alias", "SiteUser", site_user_model.id, request)
            site_user_model.email = check_change_value(site_user_model.email, email, "email", "SiteUser", site_user_model.id, request)
            site_user_model.adress = check_change_value(site_user_model.adress, adress, "adress", "SiteUser", site_user_model.id, request)
            site_user_model.contacts = check_change_value(site_user_model.contacts, contacts, "contacts", "SiteUser", site_user_model.id, request)

            site_user_model.save()
            manager.eventPercent = int(check_change_value(manager.eventPercent, percent, "eventPercent", "Manager", manager.id, request))
            manager.save()

            return JsonResponse(create_response("success", "Успешное редактирование менеджера"))
        #Добавление
        else:
            #Model User
            new_user = User(
                username = login
            )
            try:
                new_user.set_password(password)
                new_user.save()
            except:
                return JsonResponse(create_response("error", "Ошибка при создании основного объекта пользователя"))
            #add user to "managers" group
            new_user.groups.add(Group.objects.get(name = "Менеджеры"))
            try:
                #Model SiteUser
                new_site_user = SiteUser(
                    user = new_user,
                    alias = alias,
                    name = full_name,
                    email = email,
                    adress = adress,
                    contacts = contacts,
                    password = password,
                    type = "m"
                )
                new_site_user.save()
            except:
                new_user.delete()
                return JsonResponse(create_response("error", "Ошибка при создании объекта расширяемой модели пользователя"))
            try:
                new_manager = Manager(
                    siteuser=new_site_user,
                    eventPercent=percent
                )
                new_manager.save()
            except:
                new_user.delete()
                return JsonResponse(create_response("error", "Ошибка при создании объекта менеджера"))
            #SiteUserOptions
            try:
                options = SiteUserOptions(
                )

                # add users permissions
                #if (request.POST.get("checkCanShowAllCompanies") == "true"):
                    #options.can_show_all_companies = True
                options.save()
                new_site_user.options = options
                new_site_user.save()
            except:
               new_user.delete()
               return JsonResponse(create_response("error", "Ошибка при создании объекта настроек пользователя"))
            #SiteUserAllows
            try:
                pass
                #new_site_user.options.full_access_cities_list = City.objects.filter(
                    #id__in=json.loads(request.POST.get("full_access_cities_list")))
            except:
                new_user.delete()
                return JsonResponse(create_response("error", "Ошибка при создании объекта настроек пользователя"))
            write_to_change_log(new_site_user.id,"add", "SiteUser", "alias", new_site_user.alias, get_current_site_user(request), _buf_str_field1 = "Менеджер")
            return JsonResponse(create_response("success", "Успешное добавление нового менеджера"))
    #Артисты
    elif(type == "Артисты"):
        presentator_permissions = json.loads(request.POST.get("permissions"))

        #Редактирование
        if(choosen_site_user.count() > 0):
            site_user_model = SiteUser.objects.get(id = id)
            user = User.objects.get(id = site_user_model.user.id)
            presentator = Presentator.objects.get(siteuser__id = site_user_model.id)

            if(site_user_model.password != password):
                site_user_model.password = check_change_value(site_user_model.password, password, "password", "SiteUser", site_user_model.id, request)
                user.set_password(password)

            site_user_model.name = check_change_value(site_user_model.name, full_name, "name", "SiteUser", site_user_model.id, request)
            site_user_model.alias = check_change_value(site_user_model.alias, alias, "alias", "SiteUser", site_user_model.id, request)
            site_user_model.email = check_change_value(site_user_model.email, email, "email", "SiteUser", site_user_model.id, request)
            site_user_model.adress = check_change_value(site_user_model.adress, adress, "adress", "SiteUser", site_user_model.id, request)
            site_user_model.contacts = check_change_value(site_user_model.contacts, contacts, "contacts", "SiteUser", site_user_model.id, request)
            site_user_model.save()

            presentator.save()

            SiteUserOptions.objects.filter(id = site_user_model.options.id).update(presentator_show_events_limit = request.POST.get("presentator_show_events_limit"))
            try:
                PCSLink.objects.filter(presentator = presentator).delete()
                for permission in presentator_permissions:
                    pcsi_link = PCSLink(
                        presentator=presentator,
                        city = City.objects.get(id = permission["city"])
                    )
                    pcsi_link.save()
                    pcsi_link.shows = Artist.objects.filter(id__in = permission["shows"])
            except:
                PCSLink.objects.filter(presentator=presentator).delete()
                return JsonResponse(create_response("error", "Ошибка при создании доступов презентатора"))

            return JsonResponse(create_response("success", "Успешное редактирование артиста"))

        else:
            #Model User
            new_user = User(
                username = login
            )
            new_user.set_password(password)
            new_user.save()
            new_user.groups.add(Group.objects.get(name = "Артисты"))
            try:
                #Model SiteUser
                new_site_user = SiteUser(
                    user = new_user,
                    alias = alias,
                    name = name,
                    email = email,
                    adress = adress,
                    contacts = contacts,
                    password = password,
                    type="p"
                )
                new_site_user.save()
            except:
                new_user.delete()
                return JsonResponse(create_response("error", "Ошибка при создании объекта расширяемой модели пользователя"))

            #Model Presentator
            try:
                new_presentator = Presentator(
                    siteuser = new_site_user
                )
                new_presentator.save()
            except:
                new_user.delete()
                return JsonResponse(create_response("error", "Ошибка при создании объекта презентатора"))

            try:
                for permission in presentator_permissions:
                    pcsi_link = PCSLink(
                        presentator=new_presentator,
                        city = City.objects.get(id = permission["city"])
                    )
                    pcsi_link.save()
                    pcsi_link.shows = Artist.objects.filter(id__in = permission["shows"])
            except:
                new_user.delete()
                PCSLink.objects.filter(presentator=new_presentator).delete()
                return JsonResponse(create_response("error", "Ошибка при создании доступов презентатора"))
            # SiteUserOptions
            try:
                options = SiteUserOptions(
                    presentator_show_events_limit = request.POST.get("presentator_show_events_limit")
                )
                options.save()
                new_site_user.options = options
                new_site_user.save()
            except:
                new_user.delete()
                return JsonResponse(create_response("error", "Ошибка при создании объекта настроек пользователя"))

            write_to_change_log(new_site_user.id,"add", "SiteUser", "alias", new_site_user.alias, get_current_site_user(request), _buf_str_field1 = "Артист")
            return JsonResponse(create_response("success", "Успешное добавление нового артиста"))
    else:
        response = create_response("error", "Не удалось определить тип добавляемого пользователя")

    return JsonResponse(create_response("info", "Не удалось определить тип добавляемого пользователя"))

@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_users_user_data(request):
    id = request.POST.get("id")
    type = request.POST.get("type")
    link = request.POST.get("link")
    if(link == "manager"):
        try:
            id = get_site_user_from_manager(id).id
        except:
            return JsonResponse(create_response("error", "Не удается найти выбранного пользователя по коду менеджера. Перезагрузите страницу и попробуйте еще раз."))
    if(type == "Менеджеры"):
        try:
            site_users = SiteUser.objects.filter(id = id)
            if(site_users.count() == 0):
                return JsonResponse(create_response("error", "Не удается найти выбранного пользователя. Перезагрузите страницу и попробуйте еще раз."))
            site_user_model = site_users[0]
            manager = list(
                Manager.objects.filter(siteuser=site_user_model).values("siteuser__id", "siteuser__user__username",
                                                                        "siteuser__alias", "siteuser__name",
                                                                        "siteuser__email", "siteuser__adress",
                                                                        "siteuser__contacts",
                                                                        "siteuser__password", "eventPercent"))[0]
            user_permissions = list(User.objects.get(id = site_user_model.user.id).user_permissions.all().values_list("codename", flat=True))
            res_dict = {
                "user_data": manager,
                "options": get_user_options_from_db(request, site_user_model.user)
            }
        except:
            return JsonResponse(create_response("error", "Ошибка при получении данных пользователя. Перезагрузите страницу и попробуйте еще раз."))
        return JsonResponse(create_response("data", "Успешная загрузка данных", res_dict))
    elif(type == "Артисты"):
        try:
            site_users = SiteUser.objects.filter(id = id)
            if(site_users.count() == 0):
                return JsonResponse(create_response("error", "Не удается найти выбранного пользователя. Перезагрузите страницу и попробуйте еще раз."))
            site_user_model = site_users[0]
            site_user = list(site_users.values("alias", "name", "email", "adress", "contacts", "password"))[0]
            presentator = list(Presentator.objects.filter(siteuser__id = site_user_model.id).values("siteuser__id", "siteuser__user__username", "eventPercent"))[0]
            presentator.update(site_user)
            presentator_model = Presentator.objects.get(siteuser = site_user_model)
            unprocessed_presentator_permissions = PCSLink.objects.filter(presentator=presentator_model).values("city__id", "shows")
            presentator_permissions = []
            cities_ids = []
            artists_ids = []
            for permission in unprocessed_presentator_permissions:
                if(not permission["city__id"] in cities_ids):
                    cities_ids.append(permission["city__id"])
                if(not permission["shows"] in artists_ids):
                    artists_ids.append(permission["shows"])

            for city in cities_ids:
                processed_permission = {}
                processed_permission["city_id"] = city
                processed_permission["shows"] = []
                for permission in unprocessed_presentator_permissions:
                    if(city == permission["city__id"]):
                        processed_permission["shows"].append(permission["shows"])
                presentator_permissions.append(processed_permission)

            artists_data = list(Artist.objects.filter(id__in = artists_ids).values("id", "name", "color"))
            cities_data = list(City.objects.filter(id__in = cities_ids).values("id", "name"))

            #Заполнение данных городов и артистов
            for permission in presentator_permissions:
                for city in cities_data:
                    if(permission["city_id"] == city["id"]):
                        permission["city_name"] = city["name"]
                for show in permission["shows"]:
                    for show_data in artists_data:
                        if(show == show_data["id"]):
                            permission["shows"].append(show_data)
                for show in artists_ids:
                    if(show in permission["shows"]):
                        permission["shows"].remove(show)

            user_permissions = list(User.objects.get(id = site_user_model.user.id).user_permissions.all().values_list("codename", flat=True))

            options = get_user_options_from_db(_user = site_user_model.user)
            res_dict = {
                "user_data": presentator,
                "permissions": user_permissions,
                "presentator_permissions": presentator_permissions,
                "events_limit": options["presentator_show_events_limit"]
            }
        except:#Exception:
            return JsonResponse(create_response("error", "Ошибка при получении данных пользователя. Перезагрузите страницу и попробуйте еще раз."))
        return JsonResponse(create_response("data", "Успешная загрузка данных", res_dict))
    else:
        return JsonResponse(create_response("error", "Не удалось определить тип добавляемого пользователя"))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_users_user_options_data(request):
    site_user_id = request.POST.get("id")
    result = {}
    try:
        site_user = SiteUser.objects.get(id = site_user_id)
    except:
        return JsonResponse(create_response("error", "Не найден пользователь с таким идентификатором"))
    user_type = get_current_user_type(_siteuser = site_user_id)
    if(user_type == "presentator"):
        return JsonResponse(create_response("info", "Для артистов не предусмотрены дополнительные опции"))
    result["user_data"] = list(SiteUser.objects.filter(id = request.POST.get("id")).values("id", "alias"))[0]
    result["user_options"] = list(SiteUserOptions.objects.filter(id = SiteUser.objects.get(id = request.POST.get("id")).options.id).values())[0]
    result["full_list_cities"] = list(SiteUserOptions.objects.get(id = SiteUser.objects.get(id = request.POST.get("id")).options.id).full_access_cities_list.all().values_list("id", flat=True))
    result["all_cities"] = list(City.objects.filter(id__in = get_worked_cities()).values("id", "name"))
    return JsonResponse(create_response("data", "", result))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_users_user_options_edit(request):
    try:
        site_user = SiteUser.objects.get(id = request.POST.get("id"))
    except:
        return JsonResponse(create_response("error", "Пользователь не найден"))
    try:
        new_options = json.loads(request.POST.get("options"))
        allowed_cities = json.loads(request.POST.get("allowed_cities"))
    except:
        return JsonResponse(create_response("error", "Ошибка при парсинге запроса"))
    try:
        SiteUserOptions.objects.filter(id = site_user.options.id).update(**new_options)
        site_user.options.full_access_cities_list = City.objects.filter(id__in = allowed_cities)
    except:
        return JsonResponse(create_response("error", "Ошибка при обновлении настроек пользователя"))

    return JsonResponse(create_response("success", "Успешное обновление настроек пользователя"))
@user_is_auth_aj_checker
def control_users_user_individual_options_data(request):
    site_user = get_current_site_user(request)
    user_type = get_current_user_type(_siteuser = site_user)
    result = {}
    if(user_type == "presentator"):
        values_list = ["logout_request",]
    else:
        values_list = ["logout_request", "company_page_calendar", "hide_done_tasks", "larger_font", "scrolltop_show", "only_own_tasks_for_admin"]
    result["user_data"] = list(SiteUser.objects.filter(id = site_user).values("id", "alias"))[0]
    result["user_options"] = list(SiteUserOptions.objects.filter(id = SiteUser.objects.get(id = site_user).options.id).values(*values_list))[0]
    return JsonResponse(create_response("data", "", result))
@user_is_auth_aj_checker
def control_users_user_individual_options_edit(request):
    try:
        loaded_options = json.loads(request.POST.get("options"))
    except:
        return JsonResponse(create_response("error", "Ошибка при парсинге запроса"))

    site_user = get_current_site_user(request, _type="object")
    user_type = get_current_site_user(request)

    if (user_type == "presentator"):
        values_list = ["logout_request",]
    else:
        values_list = ["logout_request", "company_page_calendar", "hide_done_tasks", "larger_font", "scrolltop_show", "only_own_tasks_for_admin"]
    new_options = {}
    for option in loaded_options:
        if option in values_list:
            new_options[option] = loaded_options[option]
    try:
        SiteUserOptions.objects.filter(id = site_user.options.id).update(**new_options)
    except FileNotFoundError:
        return JsonResponse(create_response("error", "Ошибка при обновлении настроек пользователя"))
    return JsonResponse(create_response("success", "Успешное обновление настроек пользователя"))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_users_send_data(request):
    options = get_site_options()

    name = request.POST.get("name")
    username = request.POST.get("username")
    password = request.POST.get("password")
    comment = request.POST.get("comment")

    sended_email = request.POST.get("email")

    msg_content = '<p>Пользователь: <span style="font-weight: bold; color: green;">' + name + '</span></p>'
    msg_content += '<p>Логин: <span style="font-weight: bold; color: blue;">' + username + '</span></p>'
    msg_content += '<p>Пароль: <span style="font-weight: bold; color: red;">' + password + '</span></p>'
    if(comment != ""):
        msg_content += '<br></br>'
        msg_content += '<p>Комментарий администратора:</p>'
        msg_content += '<p style="font-weight: bold">' + comment + '</p>'
    msg_content += '<br></br>'
    msg_content += '<p>-    -    -</p>'
    msg_content += '<br></br>'
    msg_content += '<p>С Уважением, команда сайта <a href="http://шоуконтроль.рф">Шоуконтроль.рф</a> </p>'
    msg_content += '<p>При возникновении вопросов по сайту, можно воспользоваться формой обратной связи на сайте в разделе ' \
                   '"Помощь", либо отправив письмо по адресам электронной почты: <br></br>' + 'Администратор: ' + \
                   options["main_manager_email"] + "<br></br>" \
                   + 'Разработчик: ' + options["developer_email"] + '</p>'
    msg = connections.make_message_body_text("Учетные данные для доступа на сайт (" + name + ")", options["informer_login"]+ "@yandex.ru", sended_email, name, msg_content)
    try:
        connections.send_email(options["informer_out_server"], options["informer_login"], options["informer_password"], options["informer_login"] + "@yandex.ru", sended_email, msg)
        return JsonResponse(create_response("success", "Успешная отправка учетных данных пользователю"))
    except:
        return JsonResponse(create_response("error", "Ошибка при отправке сообщения. Попробуйте еще раз"))

@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_users_lock(request):#Блокировка пользователя сайта
    site_users = SiteUser.objects.filter(id = request.POST.get("id"))
    if(site_users.count() == 0):
        return JsonResponse(create_response("error", "Пользователь не найден. Попробуйте еще раз, при повторном возникновении проблемы, обратитесь к администратору."))

    #Получение объектов пользователя
    site_user = site_users[0]
    user = User.objects.get(id = site_user.user.id)

    status = request.POST.get("status")
    response_text = ""

    #Перевод поля is_active модели User в False необходимо для активации внутренних механизмов Django - пользователь выходит из системы. Без удаления данных сессии это не работает
    if(status == "active"):
        site_user.active = False
        #Удаление данных сессии указанного пользователя
        for session in Session.objects.all():
            if(session.get_decoded().get("_auth_user_id") and int(session.get_decoded().get("_auth_user_id"))  == int(user.id)):
                session.delete()
                break
        response_text = "Пользователь \"" + site_user.name + "\" успешно заблокирован"
    else:
        site_user.active = True
        response_text = "Пользователь \"" + site_user.name + "\" успешно разблокирован"
    site_user.save()
    user.save()
    return JsonResponse(create_response("success", response_text))

@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_users_fast_pass_change(request):
    site_users = SiteUser.objects.filter(id = request.POST.get("id"))
    if(site_users.count() == 0):
        return JsonResponse(create_response("error", "Пользователь не найден. Попробуйте еще раз, при повторном возникновении проблемы, обратитесь к администратору"))

    #Получение объектов пользователя
    site_user = site_users[0]
    user = User.objects.get(id = site_user.user.id)
    new_password = request.POST.get("password")
    try:
        user.set_password(new_password)
        site_user.password = new_password
        user.save()
        site_user.save()
    except:
        return JsonResponse(create_response("error", "Ошибка при смене пароля. Попробуйте еще раз, при повторном возникновении проблемы, обратитесь к администратору"))

    return JsonResponse(create_response("success", "Успешная смена пароля пользователя"))

@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_users_remove(request):
    if (not request.POST.get('remove_confirm') == "true"):
        return JsonResponse(create_response("error", "Удаление не подтверждено. Для подтверждения используйте переключатель"))
    site_users = SiteUser.objects.filter(id = request.POST.get("id"))
    if(site_users.count() == 0):
        return JsonResponse(create_response("error", "Пользователь не найден. Попробуйте еще раз, при повторном возникновении проблемы, обратитесь к администратору"))

    #Получение объектов пользователя
    site_user = site_users[0]
    user = User.objects.get(id = site_user.user.id)
    try:
        site_user.deleted = True
        for session in Session.objects.all():
            if (session.get_decoded().get("_auth_user_id") and int(session.get_decoded().get("_auth_user_id")) == int(user.id)):
                session.delete()
                break
        #user.is_active = False
        user.save()
        site_user.save()

        manager = get_manager_from_site_user(site_user.id)
        if(manager):
            Task.objects.filter(manager = manager).delete()
            CMSILink.objects.filter(manager = manager).delete()
    except:
        return JsonResponse(create_response("error", "Ошибка при удалении пользователя. Попробуйте еще раз, при повторном возникновении проблемы, обратитесь к администратору"))
    return JsonResponse(create_response("success", "Успешное удаление пользователя"))

#Control_stats
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def control_stats_allowed_cities(request):
    type = request.POST.get("type")
    if(type == "manager"):

        manager = get_manager_from_site_user(int(request.POST.get("user")))
        if(get_current_user_type(request) == "manager"):
            if(get_current_manager(request).id != manager.id):
                return JsonResponse(create_response("info", "Невозможно получить города другого менеджера"))
        else:
            if (int_convertor_error_0(request.POST.get("user")) == 0):
                return JsonResponse(create_response("data", "", list(City.objects.filter(id__in = get_worked_cities()).values("id", "name"))))
        try:
            cities_list = get_manager_allowed_cities(manager.id)
        except:
            return JsonResponse(create_response("error", "Ошибка при определении доступов выбранного пользователя. Пользователь не найден. Попробуйте еще раз, при повторном возникновении проблемы, обратитесь к администратору"))

        return JsonResponse(create_response("data", "Успешная загрузка данных", list(City.objects.filter(id__in = cities_list).values("id", "name"))))
    elif(type == "artist"):
        pass
    else:
        return JsonResponse(create_response("error", "Ошибка при определении типа выбранного пользователя. Попробуйте еще раз, при повторном возникновении проблемы, обратитесь к администратору"))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def control_stats_allowed_artists(request):
    type = request.POST.get("type")
    city = request.POST.get("city")
    if(type == "manager"):
        try:
            if (type == "manager"):
                manager = get_manager_from_site_user(int(request.POST.get("user")))
                if (get_current_user_type(request) == "manager"):
                    if (get_current_manager(request).id != manager.id):
                        return JsonResponse(create_response("info", "Невозможно получить шоу другого менеджера"))
            if(int(city) == 0):
                artists_list = CMSILink.objects.annotate(companies_count=Count('company')).filter(manager = manager.id, companies_count__gt=0).values_list("show", flat=True)
            else:
                artists_list = get_manager_allowed_shows_for_city(manager.id, city)
        except:
            return JsonResponse(create_response("error","Ошибка при определении доступов выбранного пользователя. Пользователь не найден. Попробуйте еще раз, при повторном возникновении проблемы, обратитесь к администратору"))
        return JsonResponse(create_response("data", "Успешная загрузка данных", list(Artist.objects.filter(id__in = artists_list).values("id", "name"))))
    elif(type == "artist"):
        pass
    else:
        return JsonResponse(create_response("error", "Ошибка при определении типа выбранного пользователя. Попробуйте еще раз, при повторном возникновении проблемы, обратитесь к администратору"))

@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def control_stats_get_stats_data(request):
    type = request.POST.get("type")
    latin_type = ""
    user = request.POST.get("user")
    city = request.POST.get("city")
    artist = request.POST.get("artist")
    if(city == ''):
        return JsonResponse(create_response("info", "Город не выбран. Выберите город, для которого нужна статистика."))
    if(artist == ''):
        return JsonResponse(create_response("info", "Артист не выбран. Выберите артиста, для которого нужна статистика."))
    artist = request.POST.get("artist")
    date_from = request.POST.get("from")
    date_to = request.POST.get("to")

    if(type == "Менеджер"):
        latin_type = "manager"
        full_stats_dict = {}
        site_user = SiteUser.objects.get(id = user)
        manager = Manager.objects.get(siteuser = site_user)

        if (get_current_user_type(request) == "manager"):
            if (get_current_manager(request).id != manager.id):
                return JsonResponse(create_response("info", "Невозможно получить шоу другого менеджера"))

        companies_count_filters = {}
        calls_count_filters = {}
        tasks_count_filters = {}
        events_count_filters = {}

        city = int(city)
        artist = int(artist)

        if(city != 0):
            companies_count_filters["company__city__name"] = City.objects.get(id = city)
            calls_count_filters["company__city"] = City.objects.get(id = city)
            tasks_count_filters["company__city"] = City.objects.get(id = city)
        if(artist != 0):
            companies_count_filters["show"] = Artist.objects.get(id = artist)
            calls_count_filters["artist"] = Artist.objects.get(id = artist)
            tasks_count_filters["artist"] = Artist.objects.get(id = artist)


        #Companies stats
        companies_count_filters["manager"] = manager

        choosen_companies = list(CMSILink.objects.filter(**companies_count_filters).distinct().values_list("company__id", flat=True))
        full_stats_dict["companies_count"]= len(choosen_companies)

        current_count = len(choosen_companies)
        logs_count = 0
        last_log = 0
        counter = 0
        logs_list = []
        #обход ограничения TooManySQL Variables
        while(current_count > 0):
            sql_vars_limit = 900
            logs = ChangeFieldLog.objects.filter(table__name = "Company", datetime__gte = date_from, datetime__lte = date_to, whoChange = site_user, table_link_id__in = choosen_companies[sql_vars_limit*counter:sql_vars_limit*(counter+1)]).values_list("datetime", flat=True)
            logs_list.extend(logs)

            counter += 1
            current_count -= sql_vars_limit
            logs_count += len(logs)

        if(logs_count != 0):
            last_log = sorted(logs_list)[-1]
            full_stats_dict["companies_date"] = convert_datetime_to_javascript(last_log)
        else:
            full_stats_dict["companies_date"] = convert_datetime_to_javascript(datetime.now().replace(minute=0, hour=0, second=0, year=1950, month=1, day=1))


        #Calls stats
        calls_count_filters["manager"] = manager
        calls_count_filters["datetime__gte"] = date_from
        calls_count_filters["datetime__lte"] = date_to
        calls = Call.objects.filter(**calls_count_filters).order_by("-datetime")

        full_stats_dict["calls_count"] = calls.count()
        if(calls.count() != 0):
            full_stats_dict["calls_date"] = convert_datetime_to_javascript(calls[0].statsdt)
            full_stats_dict["last_call"] = list(calls.values("id", "company__name"))[0]
        else:
            full_stats_dict["calls_date"] = convert_datetime_to_javascript(datetime.now().replace(minute=0, hour=0, second=0, year=1950, month=1, day=1))
            full_stats_dict["last_call"] = False
        #Tasks stats
        tasks_count_filters["manager"] = manager
        tasks_count_filters["statsdt__gte"] = date_from
        tasks_count_filters["statsdt__lte"] = date_to
        tasks = Task.objects.filter(**tasks_count_filters).order_by("-datetime")

        full_stats_dict["tasks_count"] = tasks.count()
        if(tasks.count() != 0):
            full_stats_dict["tasks_date"] = convert_datetime_to_javascript(tasks[0].datetime)
            full_stats_dict["last_task"] = list(tasks.values("id", "company__name"))[0]
        else:
            full_stats_dict["tasks_date"] = convert_datetime_to_javascript(datetime.now().replace(minute=0, hour=0, second=0, year=1950, month=1, day=1))
            full_stats_dict["last_task"] = False
        #done tasks:
        tasks_count_filters.pop("statsdt__gte")
        tasks_count_filters.pop("statsdt__lte")

        tasks_count_filters["doneDateTime__gte"] = date_from
        tasks_count_filters["doneDateTime__lte"] = date_to

        full_stats_dict["tasks_done_count"] = Task.objects.filter(**tasks_count_filters).order_by("-datetime").count()

        #Events stats
        events_count_filters["manager"] = manager
        events_count_filters["startTime__gte"] = date_from
        events_count_filters["startTime__lte"] = date_to

        events = Event.objects.filter(**events_count_filters).order_by("-startTime")
        manager_salary = events.aggregate(total=Sum(F('resultSum') * F('sumPercent') * 0.01))["total"]
        if(manager_salary):
            full_stats_dict["manager_salary"] = manager_salary
        else:
            full_stats_dict["manager_salary"] = 0
        full_stats_dict["events_count"] = events.count()

        if(events.count() != 0):
            full_stats_dict["events_date"] = convert_datetime_to_javascript(events[0].startTime)
            full_stats_dict["last_event"] = list(events.values("id"))[0]
        else:
            full_stats_dict["events_date"] = convert_datetime_to_javascript(datetime.now().replace(minute=0, hour=0, second=0, year=1950, month=1, day=1))
            full_stats_dict["last_event"] = False

        full_stats_dict["events_done_count"] = events_fill_events_statuses(events.values(*events_get_values_list_for_calc_statuses()), _request=False,
                                    _done_count=True, _siteuser=site_user.id)
        events_count_filters["statsdt__gte"] = date_from
        events_count_filters["statsdt__lte"] = date_to
        events_count_filters.pop("startTime__gte")
        events_count_filters.pop("startTime__lte")
        events = Event.objects.filter(**events_count_filters).order_by("-startTime")
        full_stats_dict["events_add_count"] = events.count()
        #Facts stats
        current_count = len(choosen_companies)
        logs_count = 0
        last_log = 0
        counter = 0
        logs_list = []

        #обход ограничения TooManySQL Variables
        while(current_count > 0):
            sql_vars_limit = 300
            logs = ChangeFieldLog.objects.filter(Q(table__name = "Company", whoChange = site_user, datetime__gte = date_from, datetime__lte = date_to, table_link_id__in = choosen_companies[sql_vars_limit*counter:sql_vars_limit*(counter+1)]) | Q(table__name = "Company", whoChange = site_user, datetime__gte = date_from, datetime__lte = date_to, link_to_object_id__in = choosen_companies[sql_vars_limit*counter:sql_vars_limit*(counter+1)])).values_list("datetime", flat=True)
            logs_list.extend(logs)
            counter += 1
            current_count -= sql_vars_limit
            logs_count += len(logs)

        if(logs_count != 0):
            last_log = sorted(logs_list)[-1]
            full_stats_dict["facts_date"] = convert_datetime_to_javascript(last_log)
        else:
            full_stats_dict["facts_date"] = convert_datetime_to_javascript(datetime.now().replace(minute=0, hour=0, second=0, year=1950, month=1, day=1))
        full_stats_dict["facts_count"] = logs_count

        #Activity
        full_stats_dict["last_activity"] = convert_datetime_to_javascript(site_user.lastActivity)
        work_time = SiteUserActivity.objects.filter(siteUser = site_user, date__gte = datetime.strptime(date_from, '%Y-%m-%d %H:%M').date(), date__lte = datetime.strptime(date_to, '%Y-%m-%d %H:%M').date()).aggregate(Sum('activityTime'))["activityTime__sum"]
        if(work_time):
            full_stats_dict["work_time"] = work_time
        else:
            full_stats_dict["work_time"] = 0


        #Stats_butons
        #Last company log
        #Обход TooManySqlVariables
        current_count = len(choosen_companies)
        logs_count = 0
        last_log = 0
        counter = 0
        logs_list = []
        while(current_count > 0):
            sql_vars_limit = 900
            logs = ChangeFieldLog.objects.filter(table__name = "Company", datetime__gte = date_from, datetime__lte = date_to, whoChange = site_user, table_link_id__in = choosen_companies[sql_vars_limit*counter:sql_vars_limit*(counter+1)]).values_list("id", "datetime")
            logs_list.extend(logs)

            counter += 1
            current_count -= sql_vars_limit
            logs_count += len(logs)
        if(logs_count != 0):
            full_stats_dict["last_company_log"] = sorted(logs_list, key=lambda x: x[1])[-1][0]
        else:
            full_stats_dict["last_company_log"] = 0



        return JsonResponse(create_response("data", "Успешная загрузка данных.", full_stats_dict))

    elif(type == "Артист"):
        pass
    else:
        return JsonResponse(create_response("error", "Ошибка при определении типа пользователя. Обновите страницу и попытайтесь снова."))
    return JsonResponse(create_response("info", "Вход в скрипт"))

#Control_lists_cities
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_lists_cities(request):
    count_on_page = 20
    if (request.POST.get("search") and request.POST.get("search") != ""):
        search = request.POST.get("search")
    else:
        search = ""
    if(request.POST.get("paginator")):
        cities = City.objects.filter(name__icontains=search)

        data = {}
        page_count = 0
        if (cities.count() != 0):
            page_count = calc_pages_count(count_on_page, cities.count())
        return JsonResponse(create_response("data", "", {"page_count": page_count}))


    page = int(request.POST.get("page"))
    cities = City.objects.filter(name__icontains = search).order_by("-enabled", "name")
    low_range = get_low_range_value(page, count_on_page)
    data = list(cities.values("id", "name")[low_range:low_range+count_on_page])

    allowed_cities = get_worked_cities()
    for city in data:
        if(city["id"] in allowed_cities):
            city["enabled"] = True
        else:
            city["enabled"] = False
    return JsonResponse(create_response("data", "Успешная загрузка данных", data))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_lists_city_add(request):
    name = request.POST.get("name")
    try:
        shows = json.loads(request.POST.get("shows"))
    except:
        return JsonResponse(create_response("error", "Ошибка при преобразовании доступных шоу города"))
    cities_check = City.objects.filter(name = name)
    if(cities_check.count() != 0):
        return JsonResponse(create_response("error", "Город с таким названием уже есть в базе. Перепроверьте данные."))
    else:
        city = City(
            name = request.POST.get("name")
        )
        city.save()
        try:
            city.worked_shows = Artist.objects.filter(id__in = shows)
        except:
            city.delete()
            return JsonResponse(create_response("error", "Ошибка при сохранении города"))
        return JsonResponse(create_response("success", "Город " + name + " был успешно добавлен в базу."))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_lists_city_remove(request):
    id = request.POST.get("id")
    cities_checker = City.objects.filter(id = id)
    if(cities_checker.count() == 0):
        return JsonResponse(create_response("error", "Ошибка. Выбранный город не найден. Перезагрузите страницу и попробуйте еще раз."))
    else:
        city_companies = Company.objects.filter(city__id = cities_checker[0].id)
        if(city_companies.count() > 0):
            return JsonResponse(create_response("error", "Ошибка. В выбранном городе есть компании. Сначала необходимо удалить их."))
        else:
            city = cities_checker[0]
            city_name = city.name
            city.delete()
            return JsonResponse(create_response("success", "Город <" + city_name + "> успешно удалён из базы."))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_lists_city_edit(request):
    id = request.POST.get("id")
    name = request.POST.get("name")
    try:
        city = City.objects.get(id = id)
    except:
        return JsonResponse(create_response("error","Город с таким идентификатором не найден в базе"))
    try:
        shows = json.loads(request.POST.get("shows"))
    except:
        return JsonResponse(create_response("error", "Ошибка при преобразовании доступных шоу города"))
    cities_check = City.objects.filter(name = name).exclude(name = city.name)
    if(cities_check.count() != 0):
        return JsonResponse(create_response("error", "Город с таким названием уже есть в базе. Перепроверьте данные."))
    else:
        cities_check = City.objects.filter(id = id)
        city = cities_check[0]
        old_name = city.name
        city.name = name
        city.save()
        try:
            city.worked_shows = Artist.objects.filter(id__in = shows)
        except:
            return JsonResponse(create_response("error", "Ошибка при настройке доступных шоу города"))
        try:
            removed_shows = Artist.objects.exclude(id__in = shows).values_list("id", flat=True)
            cmsi_links = CMSILink.objects.filter(company__city__id=city.id, show__id__in = removed_shows).distinct()
            for link in cmsi_links:
                link.company.remove(*Company.objects.filter(city__id=city.id))
            mark_users_cities_changed()
            pcsi_links = PCSLink.objects.filter(city__id=city.id, shows__id__in=removed_shows)
            for link in pcsi_links:
                link.shows.remove(*Artist.objects.filter(id__in = removed_shows))
        except:
            return JsonResponse(create_response("error", "Ошибка при очистке доступов"))
        return JsonResponse(create_response("success", "Название города было успешно изменено с " + old_name + " на " + name + "."))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_lists_city_change_status(request):
    id = request.POST.get("id")
    try:
        city = City.objects.get(id = id)
    except:
        return JsonResponse(create_response("error", "Не найден город с этим идентификатором"))
    success_text = ""
    try:
        if(city.enabled):
            city.enabled = False
            if(request.POST.get("lock_confirm") != "true"):
                return JsonResponse(create_response("error", "Прочитайте уведомление и подтвердите блокировку переключателем"))
            cmsi_links = CMSILink.objects.filter(company__city__id =city.id).distinct()
            for link in cmsi_links:
                link.company.remove(*Company.objects.filter(city__id = city.id))
            mark_users_cities_changed()
            PCSLink.objects.filter(city__id = city.id).delete()

            #Удаление городов из общего списка пользователей
            user_options_with_allowed_cities = SiteUserOptions.objects.filter(full_access_cities_list__id = city.id)
            for option in user_options_with_allowed_cities:
                option.full_access_cities_list.remove(city)
            success_text = "Город успешно заблокирован"
        else:
            city.enabled = True
            success_text = "Город успешно разблокирован"
        city.save()
    except:
        return JsonResponse(create_response("error", "Ошибка при блокировке/разблокировке города"))
    return JsonResponse(create_response("success", success_text))
#Control_lists_artists
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_lists_artist_paginator(request):

    filter_params = {}
    count_on_page = 20

    if(request.POST.get("search") and request.POST.get("search") != ""):
        search = request.POST.get("search")
    else:
        search = ""

    artists = Artist.objects.filter(name__icontains = search)

    data = {}
    if(artists.count() != 0):
        data["page_count"] = calc_pages_count(count_on_page, artists.count())
    else:
        data["page_count"] = 0
    return JsonResponse(create_response("data", "Успешная загрузка данных", data))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_lists_artist_data(request):
    count_on_page = 20
    page = int(request.POST.get("page"))

    if (request.POST.get("search") and request.POST.get("search") != ""):
        search = request.POST.get("search")
    else:
        search = ""
    artists = Artist.objects.filter(name__icontains = search).order_by("name")
    low_range = get_low_range_value(page, count_on_page)
    data = list(artists.values("id", "name", "color")[low_range:low_range+count_on_page])
    return JsonResponse(create_response("data", "Успешная загрузка данных", data))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_lists_artist_add(request):
    name = request.POST.get("name")
    color = request.POST.get("color")
    artists_check = Artist.objects.filter(name = name)
    if(artists_check.count() != 0):
        return JsonResponse(create_response("error", "Шоу с таким названием уже есть в базе. Перепроверьте данные."))
    else:
        artist = Artist(
            name = name,
            color = color
        )
        artist.save()
        return JsonResponse(create_response("success", "Шоу<" + name + "> было успешно добавлено в базу."))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_lists_artist_remove(request):
    id = request.POST.get("id")
    artists_checker = Artist.objects.filter(id = id)
    if(artists_checker.count() == 0):
        return JsonResponse(create_response("error", "Ошибка. Выбранное шоу не найдено. Перезагрузите страницу и попробуйте еще раз."))
    else:
        artist_events = Event.objects.filter(artist__id = artists_checker[0].id)
        artist_tasks = Task.objects.filter(artist__id = artists_checker[0].id)
        artist_calls = Call.objects.filter(artist__id = artists_checker[0].id)
        if(artist_events.count() > 0 or artist_tasks.count() > 0 or artist_calls.count() > 0):
            return JsonResponse(create_response("error", "Ошибка. Есть связаннные с шоу мероприятия, задачи или звонки. Сначала необходимо удалить их. Свяжитесь с администратором"))
        else:
            artist = artists_checker[0]
            artist_name = artist.name
            artist.delete()
            return JsonResponse(create_response("success", "Шоу <" + artist_name+ "> успешно удалёно из базы."))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_lists_artist_edit(request):
    id = request.POST.get("id")
    name = request.POST.get("name")
    color = request.POST.get("color")
    current_artist_name = ""
    artists_id_check = Artist.objects.filter(id = id)
    if(artists_id_check.count() == 0):
        return JsonResponse(create_response("error", "Шоу с таким идентификатором не найдено в базе. Перезагрузите страницу и попробуйте снова."))
    else:
        current_artist_name = artists_id_check[0].name
    artists_check = Artist.objects.filter(name = name)
    if(artists_check.count() != 0 and current_artist_name != name):
        return JsonResponse(create_response("error", "Шоу с таким названием уже есть в базе. Перепроверьте данные."))
    else:
        artist = artists_id_check[0]
        old_name = artist.name
        artist.name = name
        artist.color = color
        artist.save()
        return JsonResponse(create_response("success", "Шоу было изменено. Название шоу было успешно изменено с <" + old_name + "> на <" + name + ">."))

@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_aggregator(request):
    city_id = request.POST.get("city_id")
    type = request.POST.get("type")
    paginator = request.POST.get('paginator')
    page = request.POST.get('page')

    count_on_page = 20

    query_string = ""
    query_params = []

    if(type == "all"):
        query_params = [city_id, city_id, city_id]
        query_string = "SELECT DB_call.id as id, DB_call.statsdt as statsdt, 'c' as t FROM DB_call INNER JOIN DB_company \
            ON DB_call.company_id= DB_company.id WHERE DB_company.city_id = %s UNION ALL SELECT DB_task.id as id, DB_task.statsdt as datetime, 't' as t \
             FROM DB_task INNER JOIN DB_company ON DB_task.company_id= DB_company.id WHERE DB_company.city_id = %s \
              UNION ALL SELECT DB_event.id as id, DB_event.statsdt as datetime, 'e' as t \
             FROM DB_event INNER JOIN DB_company ON DB_event.company_id = DB_company.id WHERE DB_company.city_id = %s ORDER BY statsdt DESC"
    elif(type == "calls"):
        query_params = [city_id,]
        query_string = "SELECT DB_call.id as id, DB_call.statsdt as statsdt, 'c' as t FROM DB_call INNER JOIN DB_company \
                    ON DB_call.company_id = DB_company.id WHERE DB_company.city_id = %s ORDER BY statsdt DESC"
    elif(type == "events"):
        query_params = [city_id,]
        query_string = "SELECT DB_event.id as id, DB_event.statsdt as statsdt, 'e' as t FROM DB_event INNER JOIN DB_company \
                            ON DB_event.company_id= DB_company.id WHERE DB_company.city_id = %s ORDER BY statsdt DESC"
    elif(type == "tasks"):
        query_params = [city_id,]
        query_string = "SELECT DB_task.id as id, DB_task.statsdt as statsdt, 't' as t FROM DB_task INNER JOIN DB_company \
                          ON DB_task.company_id= DB_company.id WHERE DB_company.city_id = %s ORDER BY statsdt DESC"

    data = raw_query(query_string, query_params, _dict=True)

    if (paginator):
        return JsonResponse(create_response("data", "Успешная загрузка данных", make_page_dict_response(count_on_page, data.__len__())))
    else:
        ranges = get_ranges(page, count_on_page)

        ranged_data = data[ranges["from"]:ranges["to"]]
        calls = Call.objects.filter(id__in = list(map(lambda y: y["id"], list(filter(lambda x: x["t"] == "c", ranged_data))))).values("id", "company__id", "company__name", "manager__id",  "manager__siteuser__name", "manager__siteuser__alias", "artist__color", "artist__name", "comment", "datetime")
        tasks = Task.objects.filter(id__in = list(map(lambda y: y["id"], list(filter(lambda x: x["t"] == "t", ranged_data))))).values("id", "company__id", "company__name", "manager__id", "manager__siteuser__name", "manager__siteuser__alias", "artist__color", "artist__name", "description", "datetime")
        events = Event.objects.filter(id__in = list(map(lambda y: y["id"], list(filter(lambda x: x["t"] == "e", ranged_data))))).values("id", "company__id", "company__name", "manager__id", "manager__siteuser__name", "manager__siteuser__alias", "artist__color", "artist__name", "startTime", "note")

        for call in calls:
            for elem in ranged_data:
                if(elem["id"] == call["id"] and elem["t"] == "c"):
                    elem["data"] = call
        for task in tasks:
            for elem in ranged_data:
                if(elem["id"] == task["id"] and elem["t"] == "t"):
                    elem["data"] = task
                    elem["data"]["comment"] = elem["data"]["description"]
                    elem["data"].pop("description")
        for event in events:
            for elem in ranged_data:
                if(elem["id"] == event ["id"] and elem["t"] == "e"):
                    elem["data"] = event
                    elem["data"]["datetime"] = elem["data"]["startTime"]
                    elem["data"]["comment"] = elem["data"]["note"]
                    elem["data"].pop("note")
                    elem["data"].pop("startTime")

        # managers = Manager.objects.all().values("id", "siteuser")
        # site_users = SiteUser.objects.all().values("id", "user__username", "name", "alias")
        # linked_users = []
        # for manager in managers:
        #     for site_user in site_users:
        #         linked = {}
        #         if(manager["username"] == site_user["user__username"]):
        #             linked["manager__id"] = manager["id"]
        #             linked["site_user__id"] = site_user["id"]
        #             linked["name"] = site_user["name"]
        #             linked["alias"] = site_user["alias"]
        #             linked_users.append(linked)
        #
        # for elem in ranged_data:
        #     for user in linked_users:
        #         if("manager__id" in elem["data"] and elem["data"]["manager__id"] == user["manager__id"]):
        #             elem["data"]["name"] = user["name"]
        #             elem["data"]["alias"] = user["alias"]

        result_dict = {}
        result_dict["events"] = ranged_data
        return JsonResponse(create_response("data", "Успешная загрузка данных", result_dict))


#Control SiteOptions
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_options_load(request):
    try:
        options = SiteOptions.objects.all().values("main_manager_email", "developer_email", "sippoint_login", "sippoint_password")[0]
        result = []
        for option in options:
            option_elem = {}
            option_elem["field"] = option
            option_elem["verbose"] = SiteOptions._meta.get_field_by_name(option)[0].verbose_name
            option_elem["value"] = options[option]
            result.append(option_elem)
        result.sort(key = lambda x: x["field"])
    except:
        return JsonResponse(create_response("error", "Ошибка при получении настроек сайта"))
    return JsonResponse(create_response("data", "", result))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_options_save(request):
    try:
        data = json.loads(request.POST.get("data"))
        SiteOptions.objects.all().update(**data)
    except:
        return JsonResponse(create_response("error", "Ошибка при сохранении настроек сайта"))
    return JsonResponse(create_response("success", "Настройки сайта успешно сохранены"))
#Control Unload
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def control_unload_events(request):
    elements_on_page = 20

    city = int_convertor_error_0(request.POST.get("city"))
    user = int_convertor_error_0(request.POST.get("user"))

    date_from = request.POST.get("from")
    date_to = request.POST.get("to")

    sort_by = request.POST.get("sort")
    sort_list = []
    if(sort_by == "city"):
        sort_list = ["company__city__name", "manager__siteuser__name", "startTime"]
    elif(sort_by == "manager"):
        sort_list = ["manager__siteuser__name", "company__city__name" , "startTime"]
    elif (sort_by == "dt"):
        sort_list = ["startTime", "manager__siteuser__name", "company__city__name"]
    events_filter_values = {}

    events_filter_values["startTime__gte"] = datetime.strptime(date_from, '%Y-%m-%d %H:%M').replace(hour=0, minute=0, second=0, microsecond=0)
    events_filter_values["startTime__lte"] = datetime.strptime(date_to, '%Y-%m-%d %H:%M').replace(hour=23, minute=59, second=59, microsecond=999)

    if(city != 0):
        events_filter_values["company__city__id"] = city
    if(user != 0):
        events_filter_values["manager__id"] = get_manager_from_site_user(user).id

    events = Event.objects.filter(**events_filter_values).order_by(*sort_list)
    events = events.filter(Q(crashBool=False) & Q(removed=False))
    if(request.POST.get("unload")):
        if(events.count() == 0):
            return  HttpResponse("Нет данных для выгрузки")
        file_path = os.getcwd() + "/DB" + "/GeneratedFiles/" + "unload.xlsx"
        #file_path = "C:\\!TMP\\test.xlsx"

        book = xlsxwriter.Workbook(file_path)
        sheet = book.add_worksheet("Отчет")

        headerFormat = book.add_format()
        headerFormat.set_pattern(1)  # This is optional when using a solid fill.
        headerFormat.set_bg_color('#f5f5f5')
        headerFormat.set_border(2)

        city_name_format = book.add_format({'align': 'center', "border": 2, 'bold': True, 'font_size': 16})
        data_cell_format = book.add_format({"border": 1})
        percents_cell_format = book.add_format({"border": 1, "align": "right"})
        number_cell_format = book.add_format({"border": 1, "align": "left"})

        sheet.set_column(0, 0, 5)  # Ширина столбца
        sheet.set_column(1, 1, 20)  # Ширина столбца
        sheet.set_column(2, 2, 25)  # Ширина столбца
        sheet.set_column(3, 3, 40)  # Ширина столбца
        sheet.set_column(4, 4, 20)  # Ширина столбца
        sheet.set_column(5, 5, 20)  # Ширина столбца
        sheet.set_column(6, 6, 20)  # Ширина столбца
        sheet.set_column(7, 7, 20)  # Ширина столбца
        sheet.set_column(8, 8, 20)  # Ширина столбца

        sheet.write(0, 0, "№", headerFormat)
        sheet.write(0, 1, "Дата", headerFormat)
        sheet.write(0, 2, "Город", headerFormat)
        sheet.write(0, 3, "Учреждение", headerFormat)
        sheet.write(0, 4, "Менеджер", headerFormat)
        sheet.write(0, 5, "Итоговая сумма", headerFormat)
        sheet.write(0, 6, "Перечислено", headerFormat)
        sheet.write(0, 7, "Процент учреждению", headerFormat)
        sheet.write(0, 8, "Процент менеджера", headerFormat)


        events = events.values("startTime", "company__city__name", "company__name", "manager__siteuser__name", "resultSum", "sumTransfered", "percent", "sumPercent")

        y_coord = 1
        current_item = 0
        if (sort_by == "city"):
            current_item = events[0]["company__city__name"]
        elif (sort_by == "manager"):
            current_item = events[0]["manager__siteuser__name"]
        counter = 1
        for event in events:
            if (sort_by == "city"):
                if(current_item != event["company__city__name"]):
                    sheet.write(y_coord, 0, "")
                    current_item = event["company__city__name"]
                    y_coord += 1
            elif (sort_by == "manager"):
                if (current_item != event["manager__siteuser__name"]):
                    sheet.write(y_coord, 0, "")
                    current_item = event["manager__siteuser__name"]
                    y_coord += 1

            sheet.write(y_coord, 0, counter, number_cell_format)
            sheet.write(y_coord, 1, event["startTime"].strftime("%d.%m.%y"), data_cell_format)
            sheet.write(y_coord, 2,  event["company__city__name"], data_cell_format)
            sheet.write(y_coord, 3, event["company__name"], data_cell_format)
            sheet.write(y_coord, 4, event["manager__siteuser__name"], data_cell_format)
            sheet.write(y_coord, 5, event["resultSum"], data_cell_format)
            sheet.write(y_coord, 6, event["sumTransfered"], data_cell_format)
            sheet.write(y_coord, 7, str(int(event["percent"])) + "%", percents_cell_format)
            sheet.write(y_coord, 8, str(int(event["sumPercent"])) + "%", percents_cell_format)

            y_coord += 1
            counter += 1

        book.close()


        wrapper = FileWrapper(open(file_path, "rb"))
        response = HttpResponse(wrapper, content_type='application/zip')

        filename = "Report from " + str(events_filter_values["startTime__gte"].strftime("%d.%m.%y")) + " to " + str(events_filter_values["startTime__lte"].strftime("%d.%m.%y")) + ".xlsx"
        response['Content-Disposition'] = "attachment; filename=" + filename#List.xlsx"

        return response
    if(request.POST.get("paginator")):
        pages_count = calc_pages_count(elements_on_page, events.count())
        return JsonResponse(create_response("data", "", {"page_count": pages_count}))
    else:
        current_page = int_convertor_error_0(request.POST.get("page"))
        last_index = current_page * elements_on_page
        first_index = last_index - elements_on_page

        events_statuses = events_fill_events_statuses(events.values(*events_get_values_list_for_calc_statuses())[first_index:last_index], request)

        events = events.values("id", "company__city__name", "company__name", "manager__siteuser__name", "artist__color", "artist__name", "startTime", "resultSum", "sumTransfered")[first_index:last_index]

        for event in events:
            for unprocessed_event in events_statuses:
                if(event["id"] == unprocessed_event["id"]):
                    event["status"] = unprocessed_event["status"]
    return JsonResponse(create_response("data", "", list(events)))


#Artists=====================================================================================

@user_is_auth_aj_checker
def artist_allowed_artists_list(request):
    user_type = get_current_user_type(request)

    if(user_type == "manager"):
        company = Company.objects.get(id = request.POST.get("company"))
        manager = get_current_manager(request)
        allowed_artists_id = CMSILink.objects.filter(company = company, manager = manager).values("show__id")
        artists = list(Artist.objects.filter(id__in = allowed_artists_id).values("id", "color", "name"))
    elif(user_type == "admin"):
        worked_shows = get_worked_shows(_city_id = Company.objects.get(id = request.POST.get("company")).city.id, _list=True)
        artists = list(Artist.objects.filter(id__in = worked_shows).values("id", "color", "name"))
    result = create_response("data", "Успешная загрузка данных", artists)
    return JsonResponse(result)
#############################################################################################
#MainPage####################################################################################
@user_is_auth_aj_checker
def main_load_allowed_shows_to_header(request):
    user_type = get_current_user_type(request)
    city = City.objects.filter(id = request.POST.get("id"))
    if(request.POST.get("full_list")):
        full_list = True
    else:
        full_list = False
    if(city.count() == 0):
        city = False
    if(user_type == "admin"):
        shows = list(Artist.objects.all().values())
        if(city):
            shows = list(Artist.objects.filter(id__in = city[0].worked_shows.all().values_list("id", flat=True)).values())
        else:
            shows = list(Artist.objects.filter(id__in=City.objects.filter(enabled=True).distinct().values_list("worked_shows__id", flat=True)).values())
    elif(user_type == "manager"):
        manager = get_current_manager(request)#Для удобства используем переменную для идентификации менеджера
        if(full_list):
            if request.POST.get('id') == '0': #Если передалось значение "0" - значит выбраны все города
                shows = list(Artist.objects.filter(id__in=City.objects.filter(id__in = get_allowed_cities(request, _full_list=True, _clients_list=True, _id_list=True)).distinct().values_list("worked_shows__id", flat=True)).values())
            else:#В ином случае передается идентификатор города
                shows = list(get_worked_shows(city[0].id).values())
        else:
            if request.POST.get('id') == '0': #Если передалось значение "0" - значит выбраны все города
                links = CMSILink.objects.filter(manager = manager)
                shows_id = []
                for link in links:
                    if(link.company.count() > 0):
                        shows_id.append(link.show.id)
                shows = list(Artist.objects.filter(id__in = shows_id).values())#Получаем список доступных шоу менеджера для выбранного города
            else:#В ином случае передается идентификатор города
                city = City.objects.filter(id = request.POST.get("id"))#Получаем идентификатор города из пост-запроса
                shows = list(Artist.objects.filter(id__in = (CMSILink.objects.filter(manager = manager, company__city = city).values("show").distinct())).values().distinct())#Получаем список
                #доступных шоу менеджера для выбранного города
    elif(user_type == "presentator"):
        presentator = get_current_presentator(request)
        if request.POST.get('id') == '0':  # Если передалось значение "0" - значит выбраны все города
            links = PCSLink.objects.filter(presentator=presentator)
            shows_id = []
            for link in links:
                if(link.shows):
                    for show in link.shows.all():
                        shows_id.append(show.id)
            shows = list(Artist.objects.filter(
                id__in=shows_id).values())  # Получаем список доступных шоу менеджера для выбранного города
        else:  # В ином случае передается идентификатор города
            shows = list(Artist.objects.filter(
                id__in=PCSLink.objects.filter(
                    presentator=presentator, city__id = request.POST.get('id')).values("shows__id")).values().distinct())  # Получаем список
    return JsonResponse(create_response("data", "Успешная загрузка данных", shows), safe=False)
@user_is_auth_aj_checker
def main_change_admin_mode(request):
    """Включение и отключение режима администратора (ajax request)"""
    siteuser = SiteUser.objects.get(user = request.user)
    new_mode = request.POST.get("mode")
    success_text = ""
    error_text = ""
    if(siteuser.type != "a"):
        return JsonResponse(create_response("error", "Для изменения режима работы администратора, необходимо быть администратором"))
    if(new_mode == "true"):
        siteuser.options.admin_mode = True
        success_text = "Успешное включение режима администратора. Через 3 секунды страница перезагрузится."
        error_text = "Ошибка при включении режима администратора"
    else:
        siteuser.options.admin_mode = False
        success_text = "Успешное включение режима менеджера. Через 3 секунды страница перезагрузится."
        error_text = "Ошибка при включении режима менеджера"
    try:
        siteuser.options.save()
        return JsonResponse(create_response("info", success_text))
    except:
        return JsonResponse(create_response("error", error_text))

#Logs
@user_is_auth_aj_checker
def logs_history(request):
    allowed_tables = []
    allowed_linked_tables= []
    elements_on_page = 13
    element_id = request.POST.get("element_id")
    paginator = request.POST.get('paginator')
    request_type = request.POST.get('type')
    site_user = get_current_site_user(_request = request)
    options = get_user_options_from_db(request)
    if(options["admin_mode"]):
        allowed_actions = ["add", "edit", "remove", "view", "transfer", "transference", "transference_back" "permission", "mark"]
    else:
        allowed_actions = ["add", "edit", "remove", "transference", "transfer", "mark"]
    if(request_type == "companies_history"):
        allowed_tables = ["Company", "Call", "Task", "Event"]
        allowed_linked_tables = ["Company"]
    elif(request_type == "event_history"):
        allowed_tables = ["Event", "Call"]
        allowed_linked_tables = ["Event"]
    elif(request_type == "stats"):
        logs_filter_values = {}
        logs_filter_values["datetime__gte"] = request.POST.get("from")
        logs_filter_values["datetime__lte"] = request.POST.get("to")
        logs_filter_values["whoChange__id"] = request.POST.get("siteuser")
        if (get_current_manager(request).id != get_manager_from_site_user(request.POST.get("siteuser")).id and get_current_user_type(request) == "manager"):
            return JsonResponse(create_response("info", "Невозможно просматривать список мероприятий чужого менеджера"))
        if (paginator):
            pages_count = calc_pages_count(elements_on_page, ChangeFieldLog.objects.filter(**logs_filter_values).count())
            return JsonResponse(create_response("data", "", {"page_count": pages_count}))
        else:
            current_page = request.POST.get("page")
            if (current_page):
                current_page = int(current_page)
            last_index = current_page * elements_on_page
            first_index = last_index - elements_on_page
            logs = list(ChangeFieldLog.objects.filter(**logs_filter_values).order_by(
                "-datetime").values("id", "datetime", "whoChange__id", "whoChange__alias", "field__verbose", "value",
                                    "table__name", "changeType__name", "changeType__verbose",
                                    "link_to_object_table__name", "table_link_id", "link_to_object_id")[
                        first_index:last_index])
            logs_ids = []

            for log in logs:
                logs_ids.append(log["id"])

            linked_calls_ids = []
            linked_tasks_ids = []
            linked_events_ids = []
            for log in logs:
                if (log["table__name"] == "Call"):
                    linked_calls_ids.append(log["table_link_id"])
                if (log["table__name"] == "Task"):
                    linked_tasks_ids.append(log["table_link_id"])
                if (log["table__name"] == "Event"):
                    linked_events_ids.append(log["table_link_id"])
            linked_calls = list(
                Call.objects.filter(id__in=linked_calls_ids).values("id", "artist__color", "artist__name"))

            if (len(linked_tasks_ids) != 0):
                linked_tasks = list(
                    Task.objects.filter(id__in=linked_tasks_ids).values("id", "artist__color", "artist__name"))
            else:
                linked_tasks = []
            if (len(linked_events_ids) != 0):
                linked_events = Event.objects.filter(id__in=linked_events_ids).values("id", "artist__color",
                                                                                      "artist__name")
            else:
                linked_events = []
            for log in logs:
                for call in linked_calls:
                    if (log["table__name"] == "Call" and log["table_link_id"] == call["id"]):
                        log["color"] = call["artist__color"]
                        log["artist__name"] = call["artist__name"]
                for task in linked_tasks:
                    if (log["table__name"] == "Task" and log["table_link_id"] == task["id"]):
                        log["color"] = task["artist__color"]
                        log["artist__name"] = task["artist__name"]
                for event in linked_events:
                    if (log["table__name"] == "Event" and log["table_link_id"] == event["id"]):
                        log["color"] = event["artist__color"]
                        log["artist__name"] = event["artist__name"]
            return JsonResponse(create_response("data", "", logs))

    else:
        return JsonResponse(create_response("error", "Не выбран необходимый тип истории логов"))


    if(paginator):
        pages_count = calc_pages_count(elements_on_page, ChangeFieldLog.objects.filter(Q(table_link_id=element_id) | Q(link_to_object_id=element_id), Q(link_to_object_table__name__in=allowed_linked_tables) | Q(table__name__in=allowed_tables), changeType__name__in = allowed_actions).count())
        return JsonResponse(create_response("data", "", {"page_count": pages_count}))

    
    current_page = request.POST.get("page")
    if (current_page):
        current_page = int(current_page)
    last_index = current_page * elements_on_page
    first_index = last_index - elements_on_page
    logs = list(ChangeFieldLog.objects.filter(Q(table_link_id=element_id) | Q(link_to_object_id=element_id), Q(
        link_to_object_table__name__in=allowed_linked_tables) | Q(table__name__in=allowed_tables), changeType__name__in = allowed_actions).order_by(
        "-datetime").values("id", "datetime", "whoChange__id", "whoChange__alias", "field__verbose", "value", "table__name", "changeType__name", "changeType__verbose",
        "link_to_object_table__name", "table_link_id", "link_to_object_id")[first_index:last_index])
    logs_ids = []

    for log in logs:
        logs_ids.append(log["id"])
    unviewed_logs = ChangeFieldLog.objects.filter(id__in = logs_ids).exclude(whoViewChange__id = site_user).values_list("id", flat=True)
    # Создание соединяющего объекта
    throughModel = ChangeFieldLog.whoViewChange.through
    # Создание объединений объектов (дополнение связующей таблицы ключами)
    throughModel.objects.bulk_create([throughModel(changefieldlog_id=x, siteuser_id=site_user) for x in unviewed_logs])

    linked_calls_ids = []
    linked_tasks_ids = []
    linked_events_ids = []
    for log in logs:
        if(log["table__name"] == "Call"):
            linked_calls_ids.append(log["table_link_id"])
        if (log["table__name"] == "Task"):
            linked_tasks_ids.append(log["table_link_id"])
        if (log["table__name"] == "Event"):
            linked_events_ids.append(log["table_link_id"])
    linked_calls = list(Call.objects.filter(id__in = linked_calls_ids).values("id", "artist__color", "artist__name"))

    if(len(linked_tasks_ids) != 0):
        linked_tasks = list(Task.objects.filter(id__in = linked_tasks_ids).values("id", "artist__color", "artist__name"))
    else:
        linked_tasks = []
    if(len(linked_events_ids) != 0):
        linked_events = Event.objects.filter(id__in = linked_events_ids).values("id", "artist__color", "artist__name")
    else:
        linked_events = []
    for log in logs:
        for call in linked_calls:
            if(log["table__name"] == "Call" and log["table_link_id"] == call["id"]):
                log["color"] = call["artist__color"]
                log["artist__name"] = call["artist__name"]
        for task in linked_tasks:
            if(log["table__name"] == "Task" and log["table_link_id"] == task["id"]):
                log["color"] = task["artist__color"]
                log["artist__name"] = task["artist__name"]
        for event in linked_events:
            if(log["table__name"] == "Event" and log["table_link_id"] == event["id"]):
                log["color"] = event["artist__color"]
                log["artist__name"] = event["artist__name"]
    return JsonResponse(create_response("data", "", logs))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def logs_log_data(request):
    id = request.POST.get('id')
    try:
        log = list(ChangeFieldLog.objects.filter(id = id).values("datetime", "whoChange__id", "whoChange__alias", "field__verbose", "value", "table__name", "changeType__name", "changeType__verbose",
        "link_to_object_table__name", "table_link_id", "link_to_object_id"))[0]
    except:
        return JsonResponse(create_response("error", "Не удалось найти лог с этим идентификатором"))
    return JsonResponse(create_response("data", "Успешная загрузка данных", log))
#Users
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def manager_get_manager_own_data(request):
    try:
        manager = get_current_manager(request)
        site_user = get_site_user_from_manager(manager.id)
        manager_data = Manager.objects.filter(id = manager.id).values("id", "eventPercent")[0]
        manager_data["alias"] = site_user.alias
        response = create_response("data", "Успешная загрузка данных", manager_data)
    except:
        response = create_response("error", "Ошибка при получении данных менеджера")
    return JsonResponse(response)
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def manager_manager_list(request):
    site_users = SiteUser.objects.filter((Q(type="m") | Q(type = "a")), deleted = False).values("id")
    try:
        company = Company.objects.get(id = request.POST.get("company_id"))
    except:
        return JsonResponse(create_response("error", "Не найдена компания с таким идентификатором"))
    artists = Artist.objects.filter(id__in = get_allowed_shows_for_city(_city = company.city.id))
    result = {}
    result["siteusers"] = list(site_users.order_by("name").values("id", "name", "alias"))
    result["shows"] = list(artists.order_by("name").values("id", "name"))
    return JsonResponse(create_response("data", "", result))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def manager_add_company_to_manager(request):
    """Добавление одной компании менеджеру"""
    site_user = SiteUser.objects.filter(id = request.POST.get("user"))
    if(site_user.count() == 0):
        return JsonResponse(create_response("error", "Выбранный пользователь не найден. Перезагрузите страницу и попробуйте еще раз."))
    manager = Manager.objects.get(siteuser=site_user)
    company = Company.objects.get(id = request.POST.get("company"))
    shows_id = json.loads(request.POST.get("shows"))
    for show in shows_id:
        link, created = CMSILink.objects.get_or_create(manager = manager, show = Artist.objects.get(id = show))
        if(not link.company.all().filter(id = company.id).exists()):
            link.company.add(company)
            link.save()
    mark_users_cities_changed(site_user[0].id)
    return JsonResponse(create_response("success", "Компания с выбранными шоу была успешно добавлена пользователю."))
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def manager_steal_company(request):
    site_users = SiteUser.objects.filter(id = int(request.POST.get("site_user"))).values_list("id", flat=True)
    site_users_id = []
    for siteuser in site_users:
        site_users_id.append(siteuser)
    manager = get_managers_from_site_users(site_users_id)[0]
    company = Company.objects.get(id = request.POST.get("company"))
    if(request.POST.get("show_id")):
        str_shows_id = list(str(request.POST.get("show_id")))
        shows_id = list([int(i) for i in str_shows_id])
        response_status = "Шоу успешно удалено у менеджера для выбранной компании"
    else:
        shows_dict = list(CMSILink.objects.filter(company = company, manager = manager).values("show__id"))
        shows_id = []
        for key in shows_dict:
            shows_id.append(key["show__id"])
        response_status = "Компания успешно удалена у менеджера"

    for show in shows_id:
        link = CMSILink.objects.filter(company = company, manager = manager, show__id = show)[0]
        if(link.company.all().filter(id = company.id).exists()):
            link.company.remove(company)
            link.save()
    mark_users_cities_changed(site_users[0])
    return JsonResponse(create_response("success", response_status))
@user_is_auth_checker
@user_passes_test(user_is_admin, login_url='/events/')
def manager_allowed_shows(request):
    company_id = request.POST.get("company")
    site_user_id = request.POST.get("site_user")
    try:
        shows = CMSILink.objects.filter(company__id = company_id, manager__id = get_manager_from_site_user(site_user_id).id).values("show__id", "show__name", "show__color")
    except:
        return JsonResponse(create_response("error", "Ошибка при определении шоу менеджера"))
    if(shows.count() == 0):
        return JsonResponse(create_response("error", "Отсутствуют шоу у выбранного учреждения для выбранного менеджера."))
    return JsonResponse(create_response("data", "Успешная загрузка данных", list(shows)))
@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def manager_allowed_shows_for_company(request):
    user_type = get_current_user_type(request)
    try:
        company = Company.objects.get(id = request.POST.get('company'))
    except:
        return JsonResponse(create_response("error", "Не найдена компания с этим идентификатором"))
    if(user_type == "admin"):
        shows = get_allowed_shows_for_city(_company=company.id).values()
    elif(user_type == "manager"):
        shows = Artist.objects.filter(id__in = get_allowed_manager_shows_for_company(get_current_manager(request), request.POST.get("company"))).values("id", "name")
    if(shows.count() == 0):
        return JsonResponse(create_response("info", "Не найдены доступные шоу для выбранного учреждения"))
    else:
        return JsonResponse(create_response("data", "", {"shows":list(shows)}))
#Chat#############################################################################################
def chat_add_message(request):
    message = request.POST.get("message")
    table_link_id = request.POST.get("id")
    table = get_table_object_from_name(request.POST.get("table"))
    sender = get_current_site_user(request, "object")
    chat_message = ChatMessage(
        message = message,
        table_link_id = table_link_id,
        table = table,
        sender = sender
    )
    chat_message.save()
    chat_message.viewers.add(sender)
    response = create_response("success", "Сообщение успешно добавлено в чат")
    return JsonResponse(response)
def chat_show_messages(request):
    current_site_user = get_current_site_user(request)
    table_link_id = request.POST.get("id")
    table = get_table_object_from_name(request.POST.get("table"))



    #Отметка сообщений прочитанными
    #Нахождение непрочитанных сообщений для текущего пользователя
    unviewed_chat_messages_id_list = ChatMessage.objects.filter(table_link_id = table_link_id, table = table).exclude(viewers__id = current_site_user).values_list("id", flat=True)
    senderTest = get_current_site_user(request, "object")

    #Создание соединяющего объекта
    ThroughModel = ChatMessage.viewers.through
    #Создание объединений объектов (дополнение связующей таблицы ключами)
    ThroughModel.objects.bulk_create([ThroughModel(chatmessage_id = x, siteuser_id = current_site_user) for x in unviewed_chat_messages_id_list])
    ###

    #Получение сообщений
    chat_messages = list(ChatMessage.objects.filter(table_link_id = table_link_id, table = table).order_by("-datetime").values("id", "sender__id", "sender__type", "sender__alias", "datetime", "message"))

    for chat_message in chat_messages:
        if(chat_message["sender__id"]== current_site_user):
            chat_message["own"] = True
        else:
            chat_message["own"] = False
        if(chat_message["id"] in  unviewed_chat_messages_id_list):
            chat_message["new"] = True
        else:
            chat_message["new"] = False
    response = create_response("data", "Успешная загрузка данных", chat_messages, _nodatalabel=True)

    return JsonResponse(response)
##################################################################################################
#Help page
def help_send_email_to_admin(request):
    options = get_site_options()

    site_user = get_current_site_user(request, _type="object")
    current_user_type = get_current_user_type(request)

    message = request.POST.get("message")

    to = request.POST.get("to")
    if(to == "administrator"):
        sended_email = options["main_manager_email"]
    elif(to == "developer"):
        sended_email = options["developer_email"]
    else:
        return JsonResponse(create_response("error", "Невозможно определить адресата для отправки"))
    msg_content = '<p>Пользователь: <span style="font-weight: bold; color: green;">' + site_user.name + '</span></p>'
    msg_content += '<p>Логин: <span style="font-weight: bold; color: blue;">' + site_user.user.username + '</span></p>'
    msg_content += '<p>Пароль: <span style="font-weight: bold; color: blue;">' + site_user.password + '</span></p>'

    msg_content += '<br></br>'
    msg_content += '<p>Текст сообщения:</p>'
    msg_content += '<p style="font-weight: bold">' + message + '</p>'

    msg_content += '<br></br>'
    msg_content += '<p>-    -    -</p>'
    msg_content += '<br></br>'
    msg_content += '<p>С Уважением, команда сайта <a href="http://шоуконтроль.рф">Шоуконтроль.рф</a> </p>'
    msg_content += '<p>При возникновении вопросов по сайту, можно воспользоваться формой обратной связи на сайте в разделе ' \
                   '"Помощь", либо отправив письмо по адресам электронной почты: <br></br>' + 'Администратор: ' + options["main_manager_email"] + "<br></br>" \
                   + 'Разработчик: ' + options["developer_email"] + '</p>'
    msg = connections.make_message_body_text("Вопрос по сайту от " + site_user.name,
                                             options["informer_login"] + "@yandex.ru", sended_email, site_user.name, msg_content)
    try:
        connections.send_email(options["informer_out_server"], options["informer_login"], options["informer_password"],
                               options["informer_login"] + "@yandex.ru", sended_email, msg)
        return JsonResponse(create_response("success", "Успешная отправка учетных данных пользователю"))
    except:
        return JsonResponse(create_response("error", "Ошибка при отправке сообщения. Попробуйте еще раз"))



#HELP FUNCTIONS###########################################################################
#Создание объекта ответа для автогенерации уведомлений
def create_response(_status, _text, _data=False, _request=False, _nodatalabel = False):
    result = {}
    result["status"] = _status
    result["text"] = _text
    if(_status == "data"):
        if(_data or _nodatalabel):
            if(_text == "" or _text == None):
                result["text"] = "Успешная загрузка данных"
            if(type(_data) is list):
                result["data"] = {}
                result["data"]["list"] = _data
            else:
                result["data"] = _data
            if(_request):
                result["data"]["user"] = {}
                result["data"]["user"]["type"] = get_current_user_type(_request)
        else:
             result["status"] = "info"
             result["text"] = "Сервер не нашел данных по указанному запросу. Измените параметры запроса."
    return result
#Возвращает объект таблицы TableName на основе поля name
def get_table_object_from_name(_name):
    return TableName.objects.filter(name = _name)[0]
#Заполняет список городов менеджера, для оптимизации работы с базой (при заполнении страниц,
#нужно получить лишь одно поле из записи о менеджере, вместо выполнения сложных запросов в CMSI-Links)

def fill_manager_cities_list(_manager=None):
    try:
        if(not _manager):
            for siteuser in list(SiteUser.objects.filter(Q(type = "m") | Q(type="a")).values_list("id", flat=True)):
                siteuser.options.cities_changed = False
                siteuser.options.save()
                manager = get_manager_from_site_user(siteuser)
                cities = City.objects.filter(id__in=CMSILink.objects.filter(manager__id=manager.id).values("company__city__id"))
                manager.cities_buf = cities
                manager.save()
        else:
            cities = City.objects.filter(id__in = CMSILink.objects.filter(manager = _manager).values("company__city__id"))
            manager = Manager.objects.get(id = _manager)
            manager.cities_buf = cities
            siteuser = get_site_user_from_manager(_manager)

            siteuser.options.cities_changed = False
            siteuser.options.save()
            manager.save()
        return True
    except:
        return False
def mark_users_cities_changed(_site_user=False):
    try:
        if(_site_user):
            site_user = SiteUser.objects.get(id = _site_user)
            site_user.options.cities_changed = True
            site_user.options.save()
        else:
            for site_user in SiteUser.objects.all():
                site_user.options.cities_changed = True
                site_user.options.save()
    except:
        return False
    return True
#Инициализирует списки городов для всех менеджеров
def initialize_managers():
    for manager in Manager.objects.all():
        mark_users_cities_changed(get_site_user_from_manager(manager))
    return True
#Проверяет данные в переданном списке на наличие полей datetime, в случае нахождения заменяют их на строковые значения
def replace_datetime_to_string(_list):
    for list_item in _list:
        for key in list_item:
            if type(list_item[key]) is datetime:
                list_item[key] = list_item[key].strftime("%Y.%m.%d %H:%M")
    return _list
#Возвращает текущего авторизованного менеджера, используется для экономии кода, часто повторяется
def get_current_manager(_request):
    manager = Manager.objects.get(siteuser = SiteUser.objects.get(user=_request.user))
    return manager
#Возвращает текущего авторизованного презентатора
def get_current_presentator(_request):
    presentator = Presentator.objects.get(siteuser = get_current_site_user(_request, "object"))
    return presentator

def get_current_site_user(_request, _type="id"):
    """Возвращает текущего авторизованного SiteUser, используется для экономии кода, часто повторяется"""
    if(_type == "id"):
        return SiteUser.objects.get(user = _request.user).id
    if(_type == "object"):
        return SiteUser.objects.get(user = _request.user)

#Возвращает тип текущего авторизованного пользователя, используется для экономии кода, часто повторяется
def get_current_user_type(_request = False, _siteuser = False, _user = False):
    if(_request):
        user = _request.user
    elif(_siteuser):
        user = SiteUser.objects.get(id = _siteuser).user
    elif(_user):
        user = _user
    else:
        return False
    if(user_is_admin(user)):
        return "admin"
    if(user_is_manager(user)):
        return "manager"
    if(user_is_presentator(user)):
        return "presentator"
#Возвращает доступные шоу для выбранного менеджера в выбранной компании. Если параметр _manager - request, то возвращается текущий авторизованный менеджер
def get_allowed_manager_shows_for_company(_manager, _company_id):
    if type(_manager) == "int":
        manager = Manager.objects.filter(id = _manager)[0].id
    else:
        manager = _manager
    if(user_is_admin(manager.siteuser.user)):
        show_ids = Artist.objects.all().values("id")
    else:
        show_ids = Artist.objects.filter(id__in = CMSILink.objects.filter(manager = manager, company = _company_id).values("show").distinct()).values("id")
    return show_ids
#Возвращает доступные шоу для выбранного менеджера для выбранного города. Если параметр _manager - request, то возвращается текущий авторизованный менеджер
def get_manager_allowed_shows_for_city(_manager, _city):
    if type(_manager) is int:
        manager = Manager.objects.filter(id = _manager)[0].id
    else:
        manager = get_current_manager(_manager).id
    artists = Artist.objects.filter(id__in = CMSILink.objects.filter(manager__id = manager, company__city__id = _city).values("show__id").distinct()).values_list("id", flat=True)
    return artists

#Возвращает доступные шоу для выбранного презентатора для города
def get_presentator_allowed_shows_for_city(_presentator, _city):
    pcslinks = PCSLink.objects.filter(presentator__id=_presentator, city = _city)
    if(pcslinks.count() != 0):
        current_link = pcslinks[0]
    else:
        return []
    return current_link.shows.all().values_list("id", flat=True)
#Возвращает доступные для менеджера города. Если параметр _manager - request, то возвращается текущий авторизованный менеджер
def get_manager_allowed_cities(_manager):
    if type(_manager) is int:
        manager = Manager.objects.filter(id = _manager)[0].id
    else:
        manager = get_current_manager(_manager).id
    cities = City.objects.filter(id__in = Company.objects.filter(id__in = CMSILink.objects.filter(manager__id = manager).values("company__id")).values("city__id").distinct()).values_list("id", flat=True)
    return cities
#Возвращает доступные для менеджера компании. Если параметр _manager - request, то возвращается текущий авторизованный менеджер
def get_manager_allowed_companies(_manager):
    if type(_manager) == int:
        manager = Manager.objects.filter(id = _manager)[0].id
    else:
        manager = get_current_manager(_manager).id
    companies = Company.objects.filter(id__in = CMSILink.objects.filter(manager__id = manager).values("company__id")).values_list("id", flat=True)
    return companies
def get_allowed_shows_for_city(_city = False, _company = False, _cities = False):
    if(_city or (_city == 0 and type(_city) == int)):
        try:
            shows = Artist.objects.filter(id__in = City.objects.get(id = _city).worked_shows.values_list("id", flat=True))
        except:
            shows = Artist.objects.filter(id = 0)
    elif(_company):
        shows = Artist.objects.filter(id__in = City.objects.get(id = Company.objects.get(id = _company).city.id).worked_shows.values_list("id", flat=True))
    elif(_cities):
        sets = []
        for city in _cities:
            sets.append(set(City.objects.get(id = city).worked_shows.values_list("id", flat=True)))
        result_set = set(Artist.objects.all().values_list("id", flat=True))
        for cities_set in sets:
            result_set = result_set & cities_set

        shows = Artist.objects.filter(id__in=list(result_set)).values_list("id", flat=True)
    else:
        shows = Artist.objects.filter(id = 0)
    return shows
#Возвращает пары город-шоу, доступные для менеджера. Используется для глобальной фильтрации. Если параметр _manager - request, то возвращается текущий авторизованный менеджер
def get_allowed_show_city_pairs(_request, _show_id = False, _presentator_id = False):
    user_type = get_current_user_type(_request)

    show_city_pairs = []
    if(user_type == "admin"):
        if(_presentator_id):
            show_city_pairs_unsorted = PCSLink.objects.filter(presentator__id=_presentator_id).values_list(
                "city__id").order_by(
                "city__id").distinct().values("city__id", "shows__id")
            for pair in show_city_pairs_unsorted:
                tmpList = []
                tmpList.append(("artist__id", pair["shows__id"]))
                tmpList.append(("company__city__id", pair["city__id"]))
                show_city_pairs.append(tmpList)
        else:
            cities = get_worked_cities()
            filtred_pairs = City.objects.filter(id__in = cities)
            if(_show_id):
                filtred_pairs = filtred_pairs.filter(worked_shows__id = _show_id)
            show_city_pairs_unsorted = filtred_pairs.values("id", "worked_shows__id")
            for pair in show_city_pairs_unsorted:
                tmpList = []
                tmpList.append(("artist__id", pair["worked_shows__id"]))
                tmpList.append(("company__city__id", pair["id"]))
                show_city_pairs.append(tmpList)
    elif(user_type == "manager"):
        if (_presentator_id):
            show_city_pairs_unsorted = PCSLink.objects.filter(presentator__id=_presentator_id).values_list("city__id").order_by(
                "city__id").distinct().values("city__id", "shows__id")
            for pair in show_city_pairs_unsorted:
                tmpList = []
                tmpList.append(("artist__id", pair["shows__id"]))
                tmpList.append(("company__city__id", pair["city__id"]))
                show_city_pairs.append(tmpList)
        else:
            manager = get_current_manager(_request).id
            filtred_pairs = CMSILink.objects.filter(manager__id = manager).exclude(company = None)
            if (_show_id):
                filtred_pairs = filtred_pairs.filter(show = _show_id)
            show_city_pairs_unsorted = filtred_pairs.values("company__city__id", "show__id").distinct()
            for pair in show_city_pairs_unsorted:
                tmpList = []
                tmpList.append(("artist__id", pair["show__id"]))
                tmpList.append(("company__city__id", pair["company__city__id"]))
                show_city_pairs.append(tmpList)
    elif(user_type == "presentator"):
        pcslinks = PCSLink.objects.filter(presentator = get_current_presentator(_request)).values("city__id", "shows__id")
        for link in pcslinks:
            tmpList = []
            tmpList.append(("artist__id", link["shows__id"]))
            tmpList.append(("company__city__id", link["city__id"]))
            show_city_pairs.append(tmpList)
    return show_city_pairs
#Возвращает связанные объекты Manager для выбранных SiteUser's
def get_managers_from_site_users(_siteusers):
    managers = Manager.objects.filter(siteuser__id__in = _siteusers)
    return managers
#Возвращает связанный объект SiteUser от выбранного id Manager
def get_site_user_from_manager(_manager_id):
    try:
        manager = Manager.objects.get(id = _manager_id)
        return Manager.objects.get(id=_manager_id).siteuser
    except:
        return 0
#Возвращает связанный объект Manager от выбранного id SiteUser
def get_manager_from_site_user(_site_user_id):
    try:
        site_user = SiteUser.objects.get(id = _site_user_id)
        return Manager.objects.get(siteuser = site_user)
    except:
        return False
def get_manager_from_user(_user):
    try:
        siteuser = SiteUser.objects.get(user = _user)
        manager = Manager.objects.get(siteuser = siteuser)
        return manager
    except:
        return 0
def make_cms_link(_siteuser_id, _shows_ids, _companies_ids):
    """Создает связку менеджера шоу и компаний"""
    manager = get_manager_from_site_user(_siteuser_id)
    companies = Company.objects.filter(id__in = _companies_ids)
    try:
        for show in _shows_ids:
            if(CMSILink.objects.filter(manager = manager, show__id = show).count() != 0):
                link = CMSILink.objects.get(manager = manager.id, show__id = show)
                link.company.add(*companies)
            else:
                link = CMSILink(
                    manager = manager,
                    show = Artist.objects.get(id = show)
                )
                link.save()
                link.company.add(*companies)
        return True
    except:
        return False


#Возвращает связанные объекты SiteUser для выбранных Manager's
def get_site_users_from_managers(_managers):
    site_users = SiteUser.objects.filter(user__in = User.objects.filter(username__in = Manager.objects.filter(id__in = _managers).values("username")).values("id"))
    return site_users
# def add_aliases_to_managers()
#Возвращает текущие значения города и шоу, взятые из cookies
def get_current_city_show_filter_values_from_cookies(_request):
    result = {}
    if(_request.COOKIES.get("choosenCity") != "0"):
        result["city"] = _request.COOKIES.get("choosenCity")
    if(_request.COOKIES.get("choosenShow") != "0"):
        result["show"] = _request.COOKIES.get("choosenCity")
    return  result

#Заполнение значений словаря непустыми значениями
def fill_dict(**kwargs):
    res_dict = {}
    for key, value in kwargs:
        if(value != None):
            res_dict[key] = value
    return res_dict
#API imitation
#Получение общих данных базы (города, шоу и т п)
@user_is_auth_aj_checker
def api_all(request):
    data = {}
    count_on_page = 14
    # Доступные компании пользователя
    if (request.POST.get("manager_companies")):
        if(get_current_user_type(request) != "manager"):
            return JsonResponse(create_response("info", "Невозможно получить список компаний менеджера, не являясь менеджером"))
        paginator = request.POST.get("paginator")
        params = {}
        params["manager"] = get_current_manager(request)
        if (request.POST.get("city") != '0'):
            params["company__city__id"] = request.POST.get("city")
        if (request.POST.get("show") != '0'):
            params["show"] = request.POST.get("show")
        companies_ids = CMSILink.objects.filter(**params).distinct().values_list("company__id", flat=True)
        companies = Company.objects.filter(id__in=companies_ids)
        if (paginator):
            if (companies.count() != 0):
                data["page_count"] = calc_pages_count(count_on_page, companies.count())
            else:
                data["page_count"] = 0
        else:
            page = int(request.POST.get("page"))
            low_range = get_low_range_value(page, count_on_page)
            data["list"] = list(
                companies.values("id", "city__name", "name", "adress")[low_range:low_range + count_on_page])
        return JsonResponse(create_response("data", "Успешная загрузка данных", data), safe=False)
    if(request.POST.get("manager_own_data")):
        if(get_current_user_type(request) != "manager"):
            return JsonResponse(create_response("error", "Ошибка при получении данных менеджера: пользователь не является менеджером"))
        else:
            data["manager_own_data"] = list(Manager.objects.filter(id = get_current_manager(request).id).values("siteuser__id", "siteuser__name"))[0]
    if(request.POST.get("cities")):
        data["cities"] = list(City.objects.filter(id__in = get_allowed_cities(request, _full_list=True)).values("id", "name"))
    if(request.POST.get("artists")):
        data["artists"] = list(Artist.objects.all().values("id", "name", "color"))
    if(request.POST.get("city")):
        if(request.POST.get("id")):
            city = City.objects.filter(id = request.POST.get("id"))
            if(request.POST.get("allowed_city_shows")):
                data = list(Artist.objects.filter(id__in = get_allowed_shows_for_city(city[0].id)).values())
            else:
                if(city.count() == 0):
                    data = None
                else:
                    data = city.values("id", "name")[0]
        else:
            data = None
    if(request.POST.get("artist")):
        if(request.POST.get("id")):
            artist = Artist.objects.filter(id = request.POST.get("id"))
            if(artist.count() == 0):
                data = None
            else:
                data = artist.values("id", "name", "color")[0]
        else:
            data = None
    if (request.POST.get("allowed_artists")):
        if (request.POST.get("companies_for_check")):
            allowed_cities = Company.objects.filter(
                id__in=json.loads(request.POST.get("companies_for_check"))).distinct().order_by("city__id").values_list(
                "city__id", flat=True)
            allowed_shows = list(
                Artist.objects.filter(id__in=get_allowed_shows_for_city(_cities=allowed_cities)).values("id", "name",
                                                                                                        "color"))
            data["allowed_shows"] = allowed_shows
    if(request.POST.get("allowed_cities")):
        data["allowed_cities"] = list(get_allowed_cities(_request=request, _clients_list=True, _id_list=True).values("id", "name"))
    if(data == None):
        return JsonResponse(create_response("error", "Ошибка при запросе. Недостаточно параметров или ошибка при получении данных"))
    elif(len(data) == 0):
        return JsonResponse(create_response("error", "Никаких параметров не было запрошено"))
    else:
        return JsonResponse(create_response("data", "Успешная загрузка данных", data), safe=False)

#Получение даннных администратором
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def api_admin(request):
    count_on_page = 14
    data = {}
    try:
        if(request.GET.get("companies") or request.POST.get("companies")):
            filter = {}
            if(request.GET.get('city') != '0' and request.GET.get('city')):
                filter["city__id"] = request.GET.get('city')
            if(request.POST.get('city') != '0' and request.POST.get('city')):
                filter["city__id"] = request.POST.get('city')
            companies = list(Company.objects.filter(**filter).values("id", "name", "adress", "city__name"))
            if(request.GET.get("flat")):
                data = companies
            else:
                data["companies"] = list(Company.objects.filter().values("name", "adress", "city__name"))
        if(request.POST.get("companies_counters")):
            city__id = request.POST.get("city")
            shows = json.loads(request.POST.get("shows"))
            counters = {}
            all_companies_count = len(Company.objects.filter(city__id = city__id).values_list("id", flat=True))
            busy_companies_count = Company.objects.filter(id__in=CMSILink.objects.filter(show__id__in = shows, company__city__id = city__id).values_list("company__id", flat=True).distinct()).count()
            free_companies_count = all_companies_count - busy_companies_count
            counters["all_companies_count"] = all_companies_count
            counters["free_companies_count"] = free_companies_count
            return JsonResponse(create_response("data", "", counters))
        if(request.POST.get("cities")):
            data["cities"] = list(City.objects.filter(enabled=True).values("id", "name"))
        if (request.POST.get("city")):
            if (request.POST.get("id")):
                city = City.objects.filter(id=request.POST.get("id"))
                if (city.count() == 0):
                    data = None
                else:
                    city = list(city.values("id", "name"))[0]
                    data["city_data"] = city
                    allowed_shows = City.objects.get(id=request.POST.get("id")).worked_shows.all()
                    if(allowed_shows):
                        data["allowed_shows"] = list(allowed_shows.values("id", "name", "color"))
                    else:
                        data["allowed_shows"] = []
            else:
                data = None
        if(request.POST.get("allowed_artists")):
            if(request.POST.get("city_for_check")):
                allowed_shows = list(Artist.objects.filter(id__in=get_allowed_shows_for_city(_city=request.POST.get("city_for_check"))).values("id","name","color"))
                data["allowed_shows"] = allowed_shows
            if(request.POST.get("companies_for_check")):
                allowed_cities = Company.objects.filter(id__in = json.loads(request.POST.get("companies_for_check"))).distinct().order_by("city__id").values_list("city__id", flat=True)
                allowed_shows = list(Artist.objects.filter(id__in = get_allowed_shows_for_city(_cities=allowed_cities)).values("id", "name", "color"))
                data["allowed_shows"] = allowed_shows
        if(request.POST.get("artists")):
            data["artists"] = list(Artist.objects.all().values("id", "name"))
        if(request.POST.get("users")):
            data["users"] = list(SiteUser.objects.all().values("id", "name"))
        if(request.POST.get("managers")):
            manager_group = Group.objects.filter(name = "Менеджеры")
            users = User.objects.filter(groups__in = manager_group)
            if(request.POST.get("managers_active")):
                data["managers"] = list(SiteUser.objects.filter(user__in = users, deleted = False).values("id", "alias").order_by("alias"))
            else:
                data["managers"] = list(SiteUser.objects.filter(user__in = users).values("id", "alias").order_by("alias"))
        if(request.POST.get("presentators")):
            presentator_group = Group.objects.filter(name = "Артисты")
            users = User.objects.filter(groups__in = presentator_group)
            if(request.POST.get("presentators_active")):
                data["presentators"] = list(SiteUser.objects.filter(user__in = users, deleted = False).values("id", "alias"))
            else:
                data["presentators"] = list(SiteUser.objects.filter(user__in = users).values("id", "name"))
        #Доступные компании пользователя
        if(request.POST.get("manager_companies")):
            data = {}
            paginator = request.POST.get("paginator")
            params = {}
            params["manager"] = get_managers_from_site_users([request.POST.get("manager")])[0].id
            if(request.POST.get("city") != '0'):
                params["company__city__id"] = request.POST.get("city")
            if(request.POST.get("show") != '0'):
                params["show"] = request.POST.get("show")
            companies_ids = CMSILink.objects.filter(**params).distinct().values_list("company__id", flat=True)
            companies = Company.objects.filter(id__in = companies_ids)
            if(paginator):
                if(companies.count() != 0):
                    data["page_count"] = calc_pages_count(count_on_page, companies.count())
                else:
                    data["page_count"] = 0
            else:
                page = int(request.POST.get("page"))
                low_range = get_low_range_value(page, count_on_page)
                data["list"] = list(companies.values("id", "city__name", "name", "adress")[low_range:low_range+count_on_page])
        #Звонки
        if(request.POST.get("manager_calls")):
            paginator = request.POST.get("paginator")
            params = {}
            params["manager"] = get_managers_from_site_users([request.POST.get("manager")])[0].id
            if (request.POST.get("city") != '0' and request.POST.get("city") != None):
                params["company__city__id"] = request.POST.get("city")
            if (request.POST.get("show") != '0' and request.POST.get("show") != None):
                params["artist"] = request.POST.get("show")
            calls = Call.objects.filter(**params).order_by("-datetime")
            if (paginator):
                if (calls.count() != 0):
                    data["page_count"] = calc_pages_count(count_on_page, len(calls))
                else:
                    data["page_count"] = 0
            else:
                page = int(request.POST.get("page"))
                low_range = get_low_range_value(page, count_on_page)
                data["list"] = list(
                    calls.values("company__id", "company__name", "artist__name", "comment", "datetime")[low_range:low_range + count_on_page])
        #Менеджеры компании
        if(request.POST.get("company_managers")):
            if(request.POST.get("company_id") == ""):
                company = 0
            else:
                company = request.POST.get("company_id")
            if (request.POST.get("show_id") == ""):
                show = 0
            else:
                show = request.POST.get("show_id")
            data["managers"] = list(Manager.objects.filter(id__in = CMSILink.objects.filter(company__id = company, show__id = show).values_list("manager__id", flat=True).order_by("id").distinct()).values("id", 'siteuser__alias'))
        if(len(data) == 0):
            return JsonResponse(create_response("error", "Никаких параментров не было запрошено"))
        else:
            return JsonResponse(create_response("data", "Успешная загрузка данных", data), safe=False)
    except:
        return JsonResponse(create_response("error", "Ошибка при обращении к Admin-API", data), safe=False)
@user_is_auth_aj_checker
@user_is_admin_aj_checker
def individual_script(request):
    siteuser = get_current_site_user(request, _type = "object")
    if(siteuser.user.username != "archy"):
        return JsonResponse(create_response("info", "Только суперадмин может исполнять индивидуальные скрипты"))
    try:
        return JsonResponse(create_response("success", "Успешное исполнение индивидуального скрипта"))
        nahodka_companies = Company.objects.filter(city__id = City.objects.get(name = "Владивосток").id, name__icontains = "Находка")
        nahodka = City.objects.get(name = "Находка")
        for company in nahodka_companies:
            company.city = nahodka
            company.save()
        return JsonResponse(create_response("success", "Успешное исполнение индивидуального скрипта"))
        wb = load_workbook('1.xlsx')
        ws = wb.active
        city = City.objects.get(name="Иркутск")
        merged_ranges = []
        low_range = 0
        range_counter = 0
        for i in range(2, 370):
            if(ws[i][0].value != None):
                if(int_convertor_error_0(ws[i][0].value) != 0):
                    if(range_counter != 0):
                        merged_ranges.append([low_range, range_counter])
                        range_counter = 0
                    low_range = i
                    range_counter = i
            else:
                range_counter += 1
        for merged_range in merged_ranges:
            low_range = merged_range[0]
            high_range = merged_range[1]

            new_company_data = {}
            new_company_data["ctype"] = "ДС"
            new_company_data["name"] = ""
            new_company_data["adress"] = ""
            new_company_data["contacts"] = ""
            new_company_data["telephone"] = ""
            new_company_data["site"] = ""
            new_company_data["email"] = ""
            new_company_data["city"] = city
            for i in range(low_range, high_range + 1):
                if(ws[i][1].value != None):
                    new_company_data["name"] += (" " + str(ws[i][1].value))
                if (ws[i][2].value != None):
                    new_company_data["adress"] += (" " + str(ws[i][2].value))
                if (ws[i][3].value != None):
                    new_company_data["contacts"] += (" " + str(ws[i][3].value))
                if (ws[i][4].value != None):
                    new_company_data["telephone"] += (" " + str(ws[i][4].value))
                if (ws[i][5].value != None):
                    new_company_data["email"] += (" " + str(ws[i][5].value))
                if (ws[i][6].value != None):
                    new_company_data["site"] += str(ws[i][6].value)
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение города Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение г.Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение  г. Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение  г. Иркутска  детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение города  Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение г. Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение г.Иркутска детский", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение  г. Иркутска детский сад", "Детский сад")

            new_company_data["name"] = new_company_data["name"].replace("Муниципальное автономное дошкольное образовательное учреждение центр развития ребенка - детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение города Иркутска  детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение города  Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение  г. Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение г. Иркутска  детский сад", "Детский сад")

            new_company_data["name"] = new_company_data["name"].replace("Муниципальное автономное дошкольное образовательное учреждение города Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение города  Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение детский сад  г. Иркутска", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение г. Иркутска сад", "Детский сад")

            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение города Иркутска", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение города  Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение  г. Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение  г. Иркутска  детский сад", "Детский сад")

            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение города  Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное образовательное учреждение  г. Иркутска  детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение  г. Иркутска  детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение  г. Иркутска  детский сад", "Детский сад")

            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение города  Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение города  Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение  г. Иркутска детский сад", "Детский сад")
            new_company_data["name"] = new_company_data["name"].replace("Муниципальное бюджетное дошкольное образовательное учреждение  г. Иркутска детский сад", "Детский сад")


            new_company = Company(**new_company_data)
            new_company.save()

        wb = load_workbook('2.xlsx')
        ws = wb.active
        for i in range(3, 80):
            if(ws[i][1].value != None):
                new_company_data = {}
                new_company_data["ctype"] = "ШК"
                new_company_data["name"] = ws[i][1].value.replace("МБОУ СОШ", "Школа").replace("МАОУ", "Школа")
                new_company_data["adress"] = ws[i][2].value
                new_company_data["contacts"] = ws[i][3].value
                new_company_data["telephone"] = ws[i][4].value
                new_company_data["site"] = ws[i][6].value
                new_company_data["email"] = ws[i][5].value
                new_company_data["city"] = city

                new_company = Company(**new_company_data)
                new_company.save()
                #if(ws[i][0].value != 0):
                #int_convertor_error_0(ws[i][0].value)
            #print(ws[i][0].value)
            #if (ws[i][0].value == None):
                #continue
        #print(merged_ranges)
        #     ctype = ""
        #     name = str(ws[i][0].value).strip().replace("_x000D_", " ")
        #     adress = str(ws[i][1].value).strip().replace("_x000D_", " ")
        #     contacts = ""
        #     site = str(ws[i][4].value).strip().replace("_x000D_", " ")
        #     email = str(ws[i][3].value).strip().replace("_x000D_", " ")
        #     comment = ""
        #     telephone = str(ws[i][2].value).strip().replace("_x000D_", " ")
        #
        #     if (comment == None):
        #         comment = ""
        #
        #     company = Company(ctype=ctype, city=city, name=name, adress=adress, contacts=contacts, telephone=telephone,
        #                       site=site, email=email, comment=comment)
        #     company.save()
        #
        # #
        # wb = load_workbook('LSV.xlsx')
        # ws = wb.active
        # city = City.objects.get(name="Пермь")
        # for i in range(1, 76):
        #     if (ws[i][1].value == None):
        #         continue
        #     ctype = ""
        #     name = str(ws[i][1].value).strip().replace("_x000D_", " ")
        #     adress = str(ws[i][2].value).strip().replace("_x000D_", " ")
        #     contacts = str(ws[i][5].value).strip().replace("_x000D_", " ")
        #     site = ""
        #     email = str(ws[i][3].value).strip().replace("_x000D_", " ")
        #     comment = ""
        #     telephone = str(ws[i][4].value).strip().replace("_x000D_", " ")
        #
        #     if (comment == None):
        #         comment = ""
        #
        #     company = Company(ctype=ctype, city=city, name=name, adress=adress, contacts=contacts, telephone=telephone,
        #                       site=site, email=email, comment=comment)
        #     company.save()
        #datelimit = datetime.now() + timedelta(days=1)
        # events = Event.objects.filter(startTime__gte=datelimit)
        # for event in events:
        #     if(event.attentionCallDayComment != "" and event.attentionCallDayComment != "None"):
        #         print("fff")
        #         call = Call(
        #             datetime=datetime.now(),
        #             manager=event.manager,
        #             company=event.company,
        #             artist=event.artist,
        #             comment=event.attentionCallDayComment,
        #             type="event"
        #         )
        #         call.save()
        #         event.callDay = call
        #         event.save()
        #     if (event.attentionCallWeekComment != "" and event.attentionCallWeekComment != "None"):
        #         print("fff")
        #         call = Call(
        #             datetime=datetime.now(),
        #             manager=event.manager,
        #             company=event.company,
        #             artist=event.artist,
        #             comment=event.attentionCallWeekComment,
        #             type="event"
        #         )
        #         call.save()
        #         event.callWeek = call
        #         event.save()
        #     if (event.attentionCallMonthComment != "" and event.attentionCallMonthComment != "None"):
        #         print("fff")
        #         call = Call(
        #             datetime=datetime.now(),
        #             manager=event.manager,
        #             company=event.company,
        #             artist=event.artist,
        #             comment=event.attentionCallMonthComment,
        #             type="event"
        #         )
        #         call.save()
        #         event.callMonth = call
        #         event.save()

        # companies = Company.objects.filter(ctype = "ДC")
        # print(companies.count())
        # companies = Company.objects.filter(ctype="ДC").update(ctype="ДС")
        # companies = Company.objects.filter(ctype="ДC")
        # print(companies.count())
        #
        # companies = Company.objects.filter(comment="None")
        # print(companies.count())
        # companies = Company.objects.filter(comment="None").update(comment="")
        # companies = Company.objects.filter(comment="None")
        # print(companies.count())
        #
        # companies = Company.objects.filter(telephone="None")
        # print(companies.count())
        # companies = Company.objects.filter(telephone="None").update(telephone="")
        # companies = Company.objects.filter(telephone="None")
        # print(companies.count())
        #
        # companies = Company.objects.filter(contacts="None")
        # print(companies.count())
        # companies = Company.objects.filter(contacts="None").update(contacts="")
        # companies = Company.objects.filter(contacts="None")
        # print(companies.count())
        #
        # companies = Company.objects.filter(site="None")
        # print(companies.count())
        # companies = Company.objects.filter(site="None").update(site="")
        # companies = Company.objects.filter(site="None")
        # print(companies.count())
        #
        # companies = Company.objects.filter(email="None")
        # print(companies.count())
        # companies = Company.objects.filter(email="None").update(email="")
        # companies = Company.objects.filter(email="None")
        # print(companies.count())
        # #Categorizer
        # ds_patterns = ["детский сад", "детскийсад", "дескийсад", "доу", "детский клуб"]
        # sc_patterns = ["сош", "школа", "оош"]

        # all_companies = Company.objects.all()
        # for company in all_companies:
        #     if (company.ctype == "Детский сад"):
        #         company.ctype = "ДС"
        #         company.save()
        #         continue
        #     if (company.ctype == "Школа"):
        #         company.ctype = "ШК"
        #         company.save()
        #         continue
        #     if (company.ctype == ""):
        #         lower_company_name = company.name.lower()
        #         type_finded = False
        #         for ds_pattern in ds_patterns:
        #             if (lower_company_name.find(ds_pattern) != -1):
        #                 company.ctype = "ДС"
        #                 company.save()
        #                 type_finded = True
        #                 break
        #         if (type_finded):
        #             continue
        #         for sc_pattern in sc_patterns:
        #             if (lower_company_name.find(sc_pattern) != -1):
        #                 company.ctype = "ШК"
        #                 company.save()
        #                 type_finded = True
        #                 break




        # manager = get_manager_from_site_user(SiteUser.objects.get(user__username = "kahovlad").id)
        #datelimit = datetime.now().replace(year=2019, month=4, day=7, hour=23, minute=59, second=59)
        #tasks = Task.objects.filter(manager__id = manager.id, datetime__lte = datelimit, done = False)
        # tasks = Task.objects.filter(manager__id=manager.id, done=False)
        # print(tasks.count())
        # tasks.update(done = True, doneDateTime = datetime.now())
        # print(tasks.count())
        # manager = get_manager_from_site_user(SiteUser.objects.get(user__username="archy").id)
        # datelimit = datetime.now().replace(year=2019, month=4, day=7, hour=23, minute=59, second=59)
        # tasks = Task.objects.filter(manager__id = manager.id, datetime__lte = datelimit, done = False)
        #tasks = Task.objects.filter(manager__id=manager.id, done=False)
        #tasks = Task.objects.filter(manager__id=manager.id, done=False).update(datetime=datetime.now())
        # Event.objects.all().update(statsdt=datetime.now())

        # siteusers = SiteUser.objects.filter(type = "m")
        # for siteuser in siteusers:
        #     siteuser.options.clients_page_companies_count = 20
        #     siteuser.options.full_list_page_companies_count = 20
        #     siteuser.options.save()
        #Task.objects.all().update(statsdt = datetime.now())
        #Task.objects.filter(doneDateTime=None, done = True).update(doneDateTime = datetime.now())
        return JsonResponse(create_response("success", "Успешное исполнение индивидуального скрипта"))
    except FileNotFoundError:
        return JsonResponse(create_response("error", "Ошибка при исполнении индивидуального скрипта"))
def get_worked_cities():
    """Получения списка действующих городов"""
    cities = City.objects.filter(enabled = True).values_list("id", flat=True)
    return cities
def get_allowed_cities(_request = False, _clients_list = False, _full_list = False, _id_list = False, _site_user = False):
    if(_request):
        user = get_current_site_user(_request, _type="object")
    else:
        user = SiteUser.objects.get(id = _site_user)
    if(_request):
        user_type = get_current_user_type(_request = _request)
    else:
        user_type = get_current_user_type(_siteuser = _site_user)
    if(user_type == "admin"):
        cities = City.objects.filter(enabled=True)
    elif(user_type == "manager"):
        manager = get_manager_from_site_user(user.id)
        if(_clients_list):
            allowed_cities = City.objects.filter(id__in = Company.objects.filter(id__in = CMSILink.objects.filter(manager__id = manager.id).values("company__id")).values("city__id").distinct()).values_list("id", flat=True)
        else:
            allowed_cities = []
        if(_full_list):
            full_list_cities = user.options.full_access_cities_list.all().values_list("id", flat=True)
        else:
            full_list_cities = []
        ac_set = set(allowed_cities)
        fl_set = set(full_list_cities)
        cities = City.objects.filter(id__in = (ac_set | fl_set))
    elif(user_type == "presentator"):
        pass
    else:
        return []
    if(_id_list):
        return cities.values_list("id", flat=True)
    else:
        return cities
def get_worked_shows(_city_id = False, _cities_list = False, _list = False):
    if(_city_id):
        city = City.objects.get(id = _city_id)
        if(_list):
            artists =  city.worked_shows.all().values_list("id", flat=True)
        else:
            artists = city.worked_shows.all().values("id", "name", "color")
    elif(_cities_list):
        artists = City.objects.filter(id__in = _cities_list).order_by("worked_shows__id").distinct().values_list("worked_shows__id", flat=True)

    return artists
#CALCULATORS============================================================================================
#Вычисляет число страниц исходя из общего числа элементов и элементов на странице
def calc_pages_count(_count_on_page, _elems_count):
    """Вычисления количества страниц по числу элементов на странице и общему числу элементов"""
    if(_elems_count == 0):
        return 0
    pages_count = 0
    if(_elems_count % _count_on_page != 0 or _elems_count < _count_on_page):
        pages_count += 1
    pages_count += _elems_count // _count_on_page
    return pages_count
#Вычисляет нижнюю границу диапазона в зависимости от выбранной страницы и числа элементов на странице
def get_low_range_value(_page, _count_on_page):
    return (_page - 1) * _count_on_page
def get_ranges(_page, _count_on_page):
    res_dict = {}
    _page = int(_page)
    _count_on_page = int(_count_on_page)
    res_dict["from"] = (_page - 1) * _count_on_page
    res_dict["to"] = res_dict["from"] + _count_on_page
    return res_dict
def make_page_dict_response(_count_on_page, _elements_count):
    res_dict = {}
    if(_elements_count == 0):
        res_dict["page_count"] = 0
    else:
        res_dict["page_count"] = calc_pages_count(_count_on_page, _elements_count)
    return res_dict
#=======================================================================================================

#CONVERTORS#################################################################################
#Преобразует датувремя из объекта datetime в словарь
def convert_datetime_to_javascript(_datetime):
    dict_item = {}
    dict_item["year"] = _datetime.year
    dict_item["month"] = _datetime.month
    dict_item["day"] = _datetime.day
    dict_item["hour"] = _datetime.hour
    dict_item["minute"] = _datetime.minute
    dict_item["second"] = _datetime.second
    return dict_item
#Преобразует датувремя из формата JavaScript в формет datetime. В данный момент не настроено
def convert_datetime_from_javascript(_datetime):
    return datetime.strptime('Jun 1 2005  1:33PM', '%b %d %Y %I:%M%p')
def convert_date_from_select2(_date):
    return datetime.strptime(_date, "%Y-%m-%d").date()
def convert_time_from_select2(_time):
    return datetime.strptime(_time, "%H:%M")
def convert_date_from_string(_date):
    return datetime.strptime(_date, "%Y-%m-%dT%H:%M").date()
#Преобразует идентификатор объекта в базе данных в объект базы данных.
def convert_id_to_object(_table, _field, _value):#Проверяет выбранное поле выбранной таблицы, если это первичный ключ другой таблицы, возвращает объект
    table = django.apps.apps.get_model('DB', _table)
    if(table._meta.get_field(_field).get_internal_type() == "ForeignKey"):
        foreign = table._meta.get_field(_field).rel.to
        if(type(_value) == foreign):
            return _value
        foreignInstanse = foreign.objects.get(id = _value)
        return foreignInstanse
    else:
        return _value
#Преобразует найденные идентификаторы объектов в словаре полей в объекты. Используется для преобразования идентификаторов,
#полученных от клиента, как правило это запросы на сохранение данных
def convert_ids_to_objects(_table, _dict):#Сканирует словарь на типы полей выбранной таблицы, в случае первичного ключа, заменяет в словаре идентификаторы на объекты
    result = {}
    for key in _dict:
        result[key] = convert_id_to_object(_table, key, _dict[key])
    return result
#===========================================================================================

#============================================================================================

#COMMON - Функции общего назначения:
#Изменение одного поля в выбранной таблице

@user_is_auth_aj_checker
@user_is_manager_or_admin_aj_checker
def change_one_field(request):
    allowed_tables = ["Company"]
    object_id = request.POST.get("id")
    table_name = request.POST.get("table")
    field = request.POST.get("field")
    value = request.POST.get("value")
    if (table_name in allowed_tables):
        if(table_name == "Company"):
            if (not user_is_admin(request.user)):
                if(not check_user_permission(request.user, "can_show_all_companies")):
                    if((CMSILink.objects.filter(company__id = object_id, manager__siteuser__user = request.user).count() == 0)):
                        return JsonResponse(create_response("error", "Эта компания запрещена для редактирования с текущими правами доступа"))
        try:
            update_data_dict = {}
            update_data_dict[field] = value
            table = django.apps.apps.get_model('DB', table_name)
            previous_value = table.objects.filter(id = object_id).values_list(field, flat=True)[0]
            table.objects.filter(id = object_id).update(**update_data_dict)
        except:
            return JsonResponse(create_response("error", "Ошибка при редактировании поля"))
    else:
        return JsonResponse(create_response("error", "Эта таблица запрещена для редактирования с текущими правами доступа"))
    write_to_change_log(object_id, "edit", table_name, field, previous_value, get_current_site_user(request))
    return JsonResponse(create_response("success", "Значение успешно изменено"))

@user_is_auth_aj_checker
def get_one_field(request):
    allowed_tables = ["Company"]

    object_id = request.POST.get("id")
    table_name = request.POST.get("table")
    field = request.POST.get("field")

    if (table_name in allowed_tables):
        if(table_name == "Company"):
            if(not user_is_admin(request.user)):
                if(not check_user_permission(request.user, "can_show_all_companies")):
                    if(CMSILink.objects.filter(company__id = object_id, manager__siteuser__user = request.user).count() == 0):
                        return JsonResponse(create_response("error", "Эта компания запрещена для просмотра с текущими правами доступа"))
        try:
            table = django.apps.apps.get_model('DB', table_name)
            object = table.objects.filter(id=object_id).values(field)
            return JsonResponse(create_response("data", "", {"value":object[0][field]}))
        except:
            return JsonResponse(create_response("error", "Ошибка при редактировании поля"))
    else:
        return JsonResponse(create_response("error", "Эта таблица запрещена для просмотра с текущими правами доступа"))
#Записывает данные в лог - ChangeFieldLog
def write_to_change_log(_table_link_id, _changeType, _table, _field, _value, _who_change, _link_to_object_id = 0, _link_to_object_table = None, _buf_str_field1 = None, _buf_str_field2 = None):
    check_existence("TableName", "name", _table) #Проверка наличия нужного значения в базе наименований таблиц
    check_existence("FieldName", "name", _field, _table)#Проверка наличия нужного значения в базе наименований полей
    if(_link_to_object_table):
        check_existence("TableName", "name", _link_to_object_table) #Проверка наличия нужного значения в базе наименований таблиц

    site_user = SiteUser.objects.get(id = _who_change)
    logFields = {}
    logFields["table_link_id"] = _table_link_id
    logFields["changeType"] = ChangeType.objects.get(name = _changeType)
    logFields["table"] = TableName.objects.get(name = _table)
    try:
        logFields["field"] = FieldName.objects.get(name = _field)
    except:
        logFields["field"] = FieldName.objects.get(id = FieldName.objects.filter(name = _field).values_list("id", flat=True)[0])

    logFields["value"] = _value
    logFields["whoChange"] = site_user

    if(_link_to_object_id != 0):
        logFields["link_to_object_id"] = _link_to_object_id
    if(_link_to_object_table):
        logFields["link_to_object_table"] = TableName.objects.get(name = _link_to_object_table)
    if(_buf_str_field1):
        logFields["bufStrField1"] = _buf_str_field1
    if(_buf_str_field2):
        logFields["bufStrField2"] = _buf_str_field2
    log = ChangeFieldLog(**logFields)
    log.save()
    log.whoViewChange.add(site_user)
    return 0
#Проверяет наличие объекта в таблице, если нету - добавляет#check_existence("FieldName", "name", _field)
def check_existence(_table, _field, _value, _checkTableName = False, _values=False):
    table = django.apps.apps.get_model('DB', _table)
    fields = {}
    fields[_field] = _value
    if(_checkTableName and _table == "FieldName"):
        try:
            check_table = django.apps.apps.get_model('DB', _checkTableName)
            field_verbose = check_table._meta.get_field_by_name(_value)[0].verbose_name
            fields["verbose"] = field_verbose
        except:
            pass
    if(_values):
        for key in _values:
            fields[key] = _values[key]
    if(table.objects.filter(**fields).count() < 1):
        object = table(**fields)
        object.save()
    return True
#Проверка изменения поля
def check_change_value(_old_value, _new_value, _fieldName, _modelName, _id, _request):
    if(str(_old_value).strip() != str(_new_value).strip()):
        write_to_change_log(_id, "edit", _modelName, _fieldName, _old_value, get_current_site_user(_request))
        return _new_value
    else:
        return _old_value
#COMBINERS###############################################################################################
def combiners_exe(request):
    combiner_site_user_type()
    return HttpResponse("Success")

#Данные SiteUsers
def combiners_siteusers_data():
    susers = SiteUser.objects.all()
    managers_users = User.objects.filter(groups__name='Менеджеры').values_list("id", flat=True)
    artist_users = User.objects.filter(groups__name='Артисты').values_list("id", flat=True)
    for suser in susers:
        suser.email = ""
        suser.adress = ""
        suser.contacts = ""
        suser.password = ""
        suser.active = True
        suser.lastActivity = datetime.now()

        options = SiteUserOptions()
        options.save()
        suser.options = options



        if(suser.user.id in managers_users):
            suser.type = "m"
        elif(suser.user.id in artist_users):
            suser.type = "p"
        else:
            suser.type = "a"

        suser.save()
    return True
def combiners_manager_siteuser_model_filler():
    managers = Manager.objects.all()
    for manager in managers:
        manager.eventPercent = 15
        siteuser = SiteUser.objects.get(user__username = manager.username)
        manager.siteuser = siteuser
        manager.save()
    return True
def combiners_executed_tasks_done_date_and_assigned_by():
    tasks = Task.objects.all()
    counter = 0
    for task in tasks:
        counter += 1
        if(task.done):
            task.doneDateTime = task.datetime
            task.assigned_by = get_site_user_from_manager(task.manager.id)
            task.save()
    return True
def combiners_city_worked_shows():
    cities = City.objects.all()
    for city in cities:
        city.worked_shows = Artist.objects.all()
        city.save()
def combiner_site_user_alias():
    site_users = SiteUser.objects.all()
    for suser in site_users:

        suser.alias = suser.name
        suser.save()
    return 0
def combiners_pcsi_links():
    for presentator in Presentator.objects.all():
        for city in presentator.cities.all():
            psci_link = PCSLink(
                presentator = presentator,
                city = city
            )
            psci_link.save()
            psci_link.shows = presentator.artists.all()
            psci_link.save()
    return True

def combiners_events_data():
    dtwithoutmonth = datetime.now() - monthdelta(2)
    events = Event.objects.filter(startTime__gte = dtwithoutmonth).order_by("startTime")
    counter = 0
    for event in events:
        counter += 1
        event.companyName = event.company.name
        event.companyAdress = event.company.adress
        event.companyContacts = event.company.contacts
        if(event.companyNote and event.companyNote != ""):
            event.note += " Менеджер: " + event.companyNote
        if (event.artistNote and event.artistNote != ""):
            event.note += " Артист: " + event.artistNote
        event.save()
    return  True
#Activity Monitoring
#
def mark_activity(request):
    site_user = SiteUser.objects.get(user = request.user)
    #Если разница по времени между активностями менее 10 минут, то отмечается в рабочее время, иначе отчет идет заново
    if((datetime.now() - site_user.lastActivity).seconds < 600):
        activity = SiteUserActivity.objects.filter(siteUser = site_user, date = datetime.now().date())
        if(activity.count() > 0):
            activity = activity[0]
            activity.activityTime += (datetime.now() - site_user.lastActivity).seconds
        else:
            activity = SiteUserActivity(
                siteUser = site_user,
                date = datetime.now().date(),
                activityTime = (datetime.now() - site_user.lastActivity).seconds
            )
        activity.save()
    site_user.lastActivity = datetime.now()

    site_user.save()
    return

#первоначальное заполнение опций
def options_filler():

    return
def add_shows_to_cities():
    for city in City.objects.all():
        city.worked_shows = Artist.objects.all()
        city.save()
    return HttpResponse("Q")

#База данных
#Чистый запрос
def raw_query(_query_string, _params_list = [], _dict = False):
    cursor = connection.cursor()

    cursor.execute(_query_string, _params_list)
    if(_dict):
        return dictfetchall(cursor)
    else:
        return cursor.fetchall()

def dictfetchall(cursor):
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]
#+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-


#Potential Functions
def getmailslist(request):


    if request.POST.get("type") == "companies":
        companiesList = (request.POST.get("content")).split(',')
        companies = Company.objects.filter(id__in = companiesList)
    else:
        choosenCity = City.objects.get(id = request.POST.get("content"))
        link = CMSILink.objects.filter(manager = Manager.objects.get(username = request.user.username)).values("company__city__id")
        companies = Company.objects.filter(city__id__in = link)

    mailslist = []
    for company in companies:
        if(company.email != None):
            for mail in company.email.split(';'):
                mailslist.append(mail.strip())
    return render_to_response("html/companies/maillist.html", {'mails':mailslist})
def exportmailslisttoexcel(request):

    if request.POST.get("type") == "companies":
        companiesList = (request.POST.get("content")).split(',')
        companies = Company.objects.filter(id__in = companiesList)

    else:
        choosenCity = City.objects.get(id = request.POST.get("content"))
        link = CMSILink.objects.filter(manager = Manager.objects.get(username = request.user.username)).values("company__city__id")
        companies = Company.objects.filter(city__id__in = link)

    strUserName = str(request.user.username)

    file_path = os.getcwd() + "/DB" + "/GeneratedFiles" + "/" + strUserName + ".xlsx"
    #file_path = "C:\\2\\test.xlsx"

    book = xlsxwriter.Workbook(file_path)
    sheet = book.add_worksheet("Клиенты")


    headerFormat = book.add_format()
    headerFormat.set_pattern(1)  # This is optional when using a solid fill.
    headerFormat.set_bg_color('#808080')

    mailFormat = book.add_format()
    mailFormat.set_text_wrap()

    sheet.set_column(0, 0, 10)#Ширина столбца
    sheet.set_column(1, 1, 20)#Ширина столбца
    sheet.set_column(2, 2, 50)#Ширина столбца
    sheet.set_column(3, 3, 50)#Ширина столбца
    sheet.set_column(4, 4, 40)#Ширина столбца

    sheet.write(0, 0, "Код", headerFormat)
    sheet.write(0, 1, "Город", headerFormat)
    sheet.write(0, 2, "Название", headerFormat)
    sheet.write(0, 3, "Адрес", headerFormat)
    sheet.write(0, 4, "Почта", headerFormat)

    i = 1
    for link in companies:
        sheet.write(i, 0, str(link.id))
        sheet.write(i, 1, str(link.city))
        sheet.write(i, 2, str(link.name))
        sheet.write(i, 3, str(link.adress))

        if(link.email != None):
            maillist = link.email.split(";")
            splittedStr = ""
            for mail in maillist:
                splittedStr += mail.strip() + "\n"
            sheet.write(i, 4, splittedStr, mailFormat)
        i+= 1
    book.close()


    wrapper = FileWrapper(open(file_path, "rb"))
    response = HttpResponse(wrapper, content_type='application/zip')
    response['Content-Disposition'] = "attachment; filename=List.xlsx"

    return response
def testfunc(request):#Вспомогательная функция для выполнения индивидуальных скриптов админа
    if(request.user.username == "archy"):
        case = ""

        if(case == "tester"):
            api_admin(request)

        if(case == "Categorizer"):

            ds_patterns = ["детский сад", "детскийсад", "дескийсад", "доу", "детский клуб"]
            sc_patterns = ["сош","школа", "оош"]

            all_companies = Company.objects.filter()
            for company in all_companies:
                if(company.ctype == "Детский сад"):
                    company.ctype = "ДС"
                    company.save()
                    continue
                if(company.ctype == "Школа"):
                    company.ctype = "ШК"
                    company.save()
                    continue
                if(company.ctype == ""):
                    lower_company_name = company.name.lower()
                    type_finded = False
                    for ds_pattern in ds_patterns:
                        if(lower_company_name.find(ds_pattern) != -1):
                            company.ctype = "ДС"
                            company.save()
                            type_finded = True
                            break
                    if(type_finded):
                        continue
                    for sc_pattern in sc_patterns:
                        if(lower_company_name.find(sc_pattern) != -1):
                            company.ctype = "ШК"
                            company.save()
                            type_finded = True
                            break




        if(case == "AntiDouble"):
            checked_cities_list = ["Омск"]
            removed_companies_id_list = []

            for city in checked_cities_list:
                checked_companies_list = Company.objects.filter(city__name = city)
                for company in checked_companies_list:
                    if(company.id in removed_companies_id_list):
                        continue
                    else:
                        if checked_companies_list.filter(adress__contains = company.adress.strip()).count() > 1:
                            finded_double_companies = checked_companies_list.filter(adress__contains = company.adress.strip()).order_by("-id")
                            stayed_company = finded_double_companies[0]
                            for finded_double_company in finded_double_companies:
                                if finded_double_company.id == stayed_company.id:
                                    continue
                                else:
                                    Task.objects.filter(company = finded_double_company).update(company = stayed_company)
                                    Event.objects.filter(company = finded_double_company).update(company = stayed_company)
                                    Call.objects.filter(company = finded_double_company).update(company = stayed_company)

                                    links = CMSILink.objects.filter(company__id = finded_double_company.id)
                                    for link in links:
                                        link.company.remove(finded_double_company)
                                        link.company.add(stayed_company)
                                        link.save()

                                stayed_company.telephone = check_not_null_value(stayed_company.telephone, finded_double_company.telephone)

                                stayed_company.contacts = check_not_null_value(stayed_company.contacts, finded_double_company.contacts)

                                stayed_company.site = check_not_null_value(stayed_company.site, finded_double_company.site)
                                stayed_company.email = check_not_null_value(stayed_company.email, finded_double_company.email)
                                stayed_company.comment = check_not_null_value(stayed_company.comment, finded_double_company.comment)


                                stayed_company.save()
                                if(not finded_double_company.id in removed_companies_id_list):
                                    removed_companies_id_list.append(finded_double_company.id)
            Company.objects.filter(id__in = removed_companies_id_list).delete()
        if(case == "ChangeCityCase"):#Замена ошибочно указанного города в учреждении
            fromCity = City.objects.get(name = "Красноярск")
            toCity = City.objects.get(name = "Братск")
            searchStrings = ["Братск", "братск", "Вихоревка", "вихоревка", "Усть - Илимск", "Усть-Илимск", "Усть-и", "Усть- И"]
            for cityName in searchStrings:
                companies = Company.objects.filter(city = fromCity, name__icontains = cityName).update(city = toCity)




        if(case == "LoadDataFromExcelCase"): #Экспорт данных из Excel таблиц
            #Мариинск
            """wb = load_workbook('MNSK.xlsx')
            ws = wb.active
            city = City.objects.get(name = "Кемерово")
            for i in range(1, 21):
                if(ws[i][0].value == None):
                    continue
                ctype = ""
                name = str(ws[i][0].value).strip().replace("_x000D_", " ")
                adress = str(ws[i][1].value).strip().replace("_x000D_", " ")
                contacts = ""
                site = str(ws[i][4].value).strip().replace("_x000D_", " ")
                email = str(ws[i][3].value).strip().replace("_x000D_", " ")
                comment = ""
                telephone = str(ws[i][2].value).strip().replace("_x000D_", " ")

                if(comment == None):
                    comment = ""


                company = Company(ctype = ctype, city = city, name = name, adress = adress, contacts = contacts, telephone = telephone, site = site, email = email, comment = comment)
                company.save()

            #
            wb = load_workbook('LSV.xlsx')
            ws = wb.active
            city = City.objects.get(name = "Пермь")
            for i in range(1, 76):
                if(ws[i][1].value == None):
                    continue
                ctype = ""
                name = str(ws[i][1].value).strip().replace("_x000D_", " ")
                adress = str(ws[i][2].value).strip().replace("_x000D_", " ")
                contacts = str(ws[i][5].value).strip().replace("_x000D_", " ")
                site = ""
                email = str(ws[i][3].value).strip().replace("_x000D_", " ")
                comment = ""
                telephone = str(ws[i][4].value).strip().replace("_x000D_", " ")

                if(comment == None):
                    comment = ""


                company = Company(ctype = ctype, city = city, name = name, adress = adress, contacts = contacts, telephone = telephone, site = site, email = email, comment = comment)
                company.save()"""
        if(case == "DoubleCompanyRemover"):
            citiesList = ["Барнаул"]
            forremove = []
            forcheck = []
            for city in citiesList:
                cityCompanies = Company.objects.filter(city__name = city)
                for company in cityCompanies:
                    checked = Company.objects.filter(city__name = city, name__contains = company.name.strip(), adress__contains = company.adress.strip())
                    if checked.count() > 1:
                        choosen = checked[0]
                        for topCompareCompany in checked:
                            for botCompareCompany in checked:
                                if topCompareCompany.id == botCompareCompany.id:
                                    continue
                                else:
                                    if(CMSILink.objects.filter(company__id = topCompareCompany.id).count() > CMSILink.objects.filter(company__id = botCompareCompany.id).count()):
                                        choosen = topCompareCompany
                                    else:
                                        if topCompareCompany.id < botCompareCompany.id:
                                            choosen = topCompareCompany

                        for compareCompany in checked:
                            if compareCompany.id == choosen.id:
                                continue
                            else:
                                if CMSILink.objects.filter(company__id = compareCompany.id):
                                    topManager = CMSILink.objects.filter(company__id = compareCompany.id)[0].manager
                                else:
                                    topManager = None

                                if CMSILink.objects.filter(company__id = choosen.id):
                                    botManager = CMSILink.objects.filter(company__id = compareCompany.id)[0].manager
                                else:
                                    botManager = None

                                if(Event.objects.filter(company = compareCompany.id, startTime__gte = datetime.now()) or (topManager != botManager and topManager != None and botManager != None)):

                                    forcheck.append(compareCompany.id)
                                else:
                                    Event.objects.filter(company = compareCompany.id).update(company = choosen)
                                    Task.objects.filter(company = compareCompany.id).update(company = choosen)
                                    Call.objects.filter(company = compareCompany.id).update(company = choosen)
                                    forremove.append(compareCompany.id)


            Company.objects.filter(id__in = forremove).delete()
            return HttpResponse("Удалено")# %d компаний, спорных %d компаний", forremove.count(), forcheck.count())
                    #for removeCompany in checked:
                        #if removeCompany == choosen:
        if(case == "DoubleCompanyRemoverV2"):
            #citiesList = ["Астрахань", "Барнаул", "Бердск", "Братск", "Волгоград", "Волжский", "Красноярск", "Москва-Восток", "Москва-Запад", "Москва-Юг", "Нижневартовск", "Новокузнецк", "Новосибирск", "Омск", "Петербург", "Саратов", "Тольятти", "Уфа", "Череповец"]
            citiesList = ["TESTCITY",]
            currentStatus = 0


            for i, city in enumerate(citiesList):
                companies = Company.objects.filter(city__name = city)
                eventCounter = 0
                taskCounter = 0
                callCounter = 0
                cityStatus = 0
                deletedCompanies = 0
                for company in companies:
                    doubleChecker = companies.filter(name__contains = company.name.strip(), adress__contains = company.adress.strip())
                    if doubleChecker.count() < 1:
                        continue
                    else:
                        remainCompany = doubleChecker[0]
                        for checkedCompany in doubleChecker:
                            if(checkedCompany == remainCompany):
                                continue
                            else:
                                if (remainCompany.contacts.strip() == ""  or remainCompany.contacts == None) and checkedCompany.contacts.strip() != "":
                                    remainCompany.contacts = checkedCompany.contacts
                                if (remainCompany.comment.strip() == ""  or remainCompany.comment == None) and checkedCompany.comment.strip() != "":
                                    remainCompany.comment = checkedCompany.comment
                                if (remainCompany.email.strip() == "" or remainCompany.email == None) and checkedCompany.email.strip() != "":
                                    remainCompany.email = checkedCompany.email
                                if (remainCompany.site.strip() == "" or remainCompany.site == None) and checkedCompany.site.strip() != "":
                                    remainCompany.site = checkedCompany.site
                                if (remainCompany.telephone.strip() == "" or remainCompany.telephone == None) and checkedCompany.telephone.strip() != "":
                                    remainCompany.telephone = checkedCompany.telephone
                                remainCompany.save()

                                movedEvents = Event.objects.filter(company = checkedCompany).update(company = remainCompany)
                                if(movedEvents):
                                    eventCounter += len(movedEvents)

                                movedTasks = Task.objects.filter(company = checkedCompany).update(company = remainCompany)
                                if(movedTasks):
                                    taskCounter += len(movedTasks)

                                movedCalls = Call.objects.filter(company = checkedCompany).update(company = remainCompany)
                                if(movedCalls):
                                    callCounter += len(movedCalls)


                                managerLinks = CMSILink.objects.filter(company__id = checkedCompany.id)
                                for link in managerLinks:
                                    link.company.remove(checkedCompany)
                                    link.company.add(remainCompany)
                                    link.save()
                            deletedCompany = Company.objects.filter(id = checkedCompany.id)
                            deletedCompany.delete()
                            deletedCompanies += 1
                    cityStatus += 100 / len(companies)
                    print("Текущий прогресс по городу: " + str(cityStatus) + " %")


                print("Удалено " + str(deletedCompanies) + " компаний.")
                print("Город " + city + " проверен. Перемещено " + str(eventCounter) + " мероприятий, " + str(taskCounter) + " задач, " + str(callCounter) + " звонков.")
                currentStatus += 100/ len(citiesList)
                print("Общий прогресс: " + str(currentStatus) + " %")
    else:
        return HttpResponse("НЕАВТОРИЗОВАНО")
    return HttpResponse(case + " SCRIPT END")
def check_not_null_value(value1, value2):
    value1 = value1.strip(" ")
    value2 = value2.strip(" ")
    res = ""
    if(value1 != None and value1 != "" and value1 != value2):
        res = value1
    elif(value2 != None and value2 != "" and value1 != value2):
        res = value2
    if(value1 == value2):
        res = value1
    return res
def int_convertor_error_json_response(_value):
    try:
        return int(_value)
    except:
        return JsonResponse(create_response("error", "Фатальная ошибка: не удалось выполнить преобразование в число"))
def int_convertor_error_0(_value):
    try:
        return int(_value)
    except:
        return 0

def int_convertor_error_false(_value):
    try:
        return int(_value)
    except:
        return False
def bool_convertor_error_json_response(_value):
    if(_value == "true"):
        return True
    elif(_value == "false"):
        return False
    else:
        return False

def time_checker():
    # start_time = time.time()
    # print("--- %s seconds ---" % (time.time() - start_time))
    return  0

